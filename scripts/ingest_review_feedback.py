"""
Feedback Ingestion: Read human-reviewed validation_review.xlsx and produce
confirmed controls, error analysis, and config improvement suggestions.

Reads:
    output/6 - validation/validation_review.xlsx  (human-reviewed)
    output/2 - controls/controls_output.csv       (original extraction)

Outputs:
    output/6 - validation/confirmed_controls.csv
    output/6 - validation/feedback_report.txt
    output/6 - validation/suggested_config_changes.yaml  (if applicable)

Usage:
    python scripts/ingest_review_feedback.py
    python scripts/ingest_review_feedback.py --config dps_config.xlsx
"""

from __future__ import annotations

import csv
import os
import re
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

from shared_utils import (
    ensure_output_dir,
    load_config,
    setup_argparse,
)

# Must match extract_controls.py CSV_COLUMNS
CONTROLS_CSV_COLUMNS = [
    "control_id", "control_name", "baseline",
    "control_description", "supplemental_guidance", "purpose", "scope", "applicability",
    "miscellaneous", "section_header", "source_file", "extraction_source",
    "compliance_date", "published_url",
]

# Column name mapping from validation_review.xlsx → controls_output.csv field
REVIEW_TO_CSV = {
    "Control ID": "control_id",
    "Control Name": "control_name",
    "Baseline": "baseline",
    "Extracted Description": "control_description",
    "Extracted Guidance": "supplemental_guidance",
    "Section Header": "section_header",
    "Source File": "source_file",
    "Extraction Source": "extraction_source",
}

# Column name mapping from "Add Missing Controls" sheet → controls_output.csv field
MISSING_TO_CSV = {
    "Control ID": "control_id",
    "Control Name": "control_name",
    "Baseline": "baseline",
    "Control Description": "control_description",
    "Supplemental Guidance": "supplemental_guidance",
    "Section Header": "section_header",
    "Source File": "source_file",
}


def load_reviewed_workbook(xlsx_path: str) -> tuple[list[dict], list[dict]]:
    """
    Read the reviewed validation_review.xlsx.
    Returns (reviewed_rows, missing_rows) where each row is a dict keyed by header name.
    """
    wb = load_workbook(xlsx_path, data_only=True)

    # --- Validation Review sheet ---
    reviewed_rows = []
    if "Validation Review" in wb.sheetnames:
        ws = wb["Validation Review"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val if val is not None else ""
            if row_dict.get("Control ID"):
                reviewed_rows.append(row_dict)

    # --- Add Missing Controls sheet ---
    missing_rows = []
    if "Add Missing Controls" in wb.sheetnames:
        ws = wb["Add Missing Controls"]
        # Headers are in row 2 (row 1 is the instruction)
        headers = [cell.value for cell in ws[2]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val if val is not None else ""
            # Skip empty rows and the placeholder example row.
            # The placeholder is identified by having BOTH the example control
            # ID and the example description — not just the ID alone, since a
            # real control could legitimately be AC-2.3.
            cid = str(row_dict.get("Control ID", "")).strip()
            desc = str(row_dict.get("Control Description", "")).strip().lower()
            is_placeholder = (
                cid == "AC-2.3"
                and ("example" in desc or "placeholder" in desc or not desc)
            )
            if cid and not is_placeholder:
                missing_rows.append(row_dict)

    wb.close()
    return reviewed_rows, missing_rows


def load_original_controls(csv_path: str) -> list[dict]:
    """Read the original controls_output.csv."""
    controls = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            controls.append(dict(row))
    return controls


def _map_review_row_to_csv(row: dict, field_map: dict) -> dict:
    """Map a review/missing row to controls_output.csv schema."""
    csv_row = {col: "" for col in CONTROLS_CSV_COLUMNS}
    for src_key, dst_key in field_map.items():
        val = row.get(src_key, "")
        csv_row[dst_key] = str(val).strip() if val else ""
    return csv_row


def build_confirmed_controls(
    reviewed_rows: list[dict],
    missing_rows: list[dict],
    original_controls: list[dict],
    output_path: str,
) -> dict:
    """
    Build confirmed_controls.csv by merging:
    - Original controls that are "Correct" or unreviewed (not false positive)
    - Manually added missing controls

    Returns stats dict for the report.
    """
    # Index reviewed rows by (control_id, source_file) → status
    review_status = {}
    for row in reviewed_rows:
        cid = str(row.get("Control ID", "")).strip()
        sf = str(row.get("Source File", "")).strip()
        status = str(row.get("Validation Status", "")).strip()
        if cid:
            review_status[(cid, sf)] = status

    # Categorize original controls
    confirmed = []
    excluded = []
    unreviewed_count = 0
    status_counts = Counter()

    for ctrl in original_controls:
        cid = ctrl.get("control_id", "").strip()
        sf = ctrl.get("source_file", "").strip()
        status = review_status.get((cid, sf), "")
        status_counts[status if status else "(unreviewed)"] += 1

        if status == "Wrong-FalsePositive":
            excluded.append(ctrl)
        else:
            confirmed.append(ctrl)
            if not status:
                unreviewed_count += 1

    # Append manually added missing controls
    added_missing = []
    for row in missing_rows:
        csv_row = _map_review_row_to_csv(row, MISSING_TO_CSV)
        csv_row["extraction_source"] = "Manual-Review"
        confirmed.append(csv_row)
        added_missing.append(csv_row)

    # Write confirmed_controls.csv
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CONTROLS_CSV_COLUMNS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for row in confirmed:
            # Ensure all columns exist
            clean = {col: row.get(col, "") for col in CONTROLS_CSV_COLUMNS}
            writer.writerow(clean)

    return {
        "total_original": len(original_controls),
        "confirmed_count": len(confirmed),
        "excluded_count": len(excluded),
        "unreviewed_count": unreviewed_count,
        "added_missing_count": len(added_missing),
        "added_missing": added_missing,
        "excluded": excluded,
        "status_counts": status_counts,
    }


def analyze_error_patterns(
    reviewed_rows: list[dict],
    missing_rows: list[dict],
    config: dict,
) -> dict:
    """
    Analyze patterns in reviewer feedback to produce actionable insights.
    """
    analysis = {
        "false_positive_patterns": {},
        "description_errors": [],
        "missing_control_patterns": {},
        "flag_accuracy": {},
        "per_document_stats": {},
        "suggestions": [],
    }

    # --- False positive analysis ---
    fp_rows = [r for r in reviewed_rows
                if str(r.get("Validation Status", "")).strip() == "Wrong-FalsePositive"]

    # Group by section header
    fp_by_section = Counter()
    fp_by_prefix = Counter()
    for row in fp_rows:
        section = str(row.get("Section Header", "")).strip()
        cid = str(row.get("Control ID", "")).strip()
        if section:
            fp_by_section[section.lower()] += 1
        # Extract control ID prefix (letters before first digit)
        prefix_match = re.match(r'^([A-Z]{2,4})', cid)
        if prefix_match:
            fp_by_prefix[prefix_match.group(1)] += 1

    analysis["false_positive_patterns"] = {
        "by_section": dict(fp_by_section),
        "by_prefix": dict(fp_by_prefix),
        "total": len(fp_rows),
    }

    # Suggest blacklist entries for common false positive prefixes
    for prefix, count in fp_by_prefix.items():
        if count >= 3:
            analysis["suggestions"].append({
                "type": "blacklist",
                "value": f"{prefix}-*",
                "reason": f"{count} controls with prefix '{prefix}' marked as false positives",
                "action": f"Add '{prefix}-*' to control_extraction.blacklist in dps_config.xlsx",
            })

    # Suggest suspect sections for common false positive sections
    for section, count in fp_by_section.items():
        if count >= 3:
            analysis["suggestions"].append({
                "type": "suspect_section",
                "value": section,
                "reason": f"{count} false positives from section '{section}'",
                "action": f"Add '{section}' to SUSPECT_SECTIONS in validate_controls.py",
            })

    # --- Description / guidance error analysis ---
    desc_errors = [r for r in reviewed_rows
                   if str(r.get("Validation Status", "")).strip() in
                   ("Wrong-Description", "Wrong-Guidance")]
    guidance_flag_correlation = sum(
        1 for r in desc_errors
        if "GUIDANCE_IN_DESC" in str(r.get("Flags", ""))
    )
    if desc_errors:
        analysis["description_errors"] = {
            "total": len(desc_errors),
            "with_guidance_flag": guidance_flag_correlation,
            "correlation_pct": round(guidance_flag_correlation / len(desc_errors) * 100, 1)
                               if desc_errors else 0,
        }
        if guidance_flag_correlation >= 3:
            analysis["suggestions"].append({
                "type": "guidance_keywords",
                "reason": f"{guidance_flag_correlation} description/guidance errors correlated with GUIDANCE_IN_DESC flag",
                "action": "Review guidance_keywords in dps_config.xlsx — may need additional keywords to split description from guidance correctly",
            })

    # --- Missing controls analysis ---
    if missing_rows:
        missing_by_file = Counter()
        missing_by_section = Counter()
        for row in missing_rows:
            sf = str(row.get("Source File", "")).strip()
            sec = str(row.get("Section Header", "")).strip()
            if sf:
                missing_by_file[sf] += 1
            if sec:
                missing_by_section[sec.lower()] += 1

        analysis["missing_control_patterns"] = {
            "total": len(missing_rows),
            "by_file": dict(missing_by_file),
            "by_section": dict(missing_by_section),
        }

        # Check if missing control IDs match the current regex patterns
        ctrl_cfg = config.get("control_extraction", {})
        id_patterns = ctrl_cfg.get("control_id_patterns",
                                    [r'\b[A-Z]{2,4}[-.]?\d{1,3}[-.]\d{2,4}\b'])
        combined_re = re.compile("|".join(f"({p})" for p in id_patterns))

        unmatched_ids = []
        for row in missing_rows:
            cid = str(row.get("Control ID", "")).strip()
            if cid and not combined_re.search(cid):
                unmatched_ids.append(cid)

        if unmatched_ids:
            analysis["suggestions"].append({
                "type": "control_id_pattern",
                "value": unmatched_ids,
                "reason": f"{len(unmatched_ids)} manually added control(s) do not match current regex patterns: {', '.join(unmatched_ids[:5])}",
                "action": "Add a new control_id_pattern to dps_config.xlsx to capture these IDs",
            })

    # --- Flag accuracy: correlate flags with actual errors ---
    flag_hits = Counter()   # flag appeared AND control was wrong
    flag_total = Counter()  # flag appeared at all
    for row in reviewed_rows:
        flags = [f.strip() for f in str(row.get("Flags", "")).split(",") if f.strip()]
        status = str(row.get("Validation Status", "")).strip()
        is_error = status.startswith("Wrong-") or status == "Missing-Content"
        for flag in flags:
            flag_total[flag] += 1
            if is_error:
                flag_hits[flag] += 1

    flag_accuracy = {}
    for flag in flag_total:
        total = flag_total[flag]
        hits = flag_hits.get(flag, 0)
        flag_accuracy[flag] = {
            "total": total,
            "true_positives": hits,
            "precision_pct": round(hits / total * 100, 1) if total > 0 else 0,
        }
    analysis["flag_accuracy"] = flag_accuracy

    # --- Per-document accuracy ---
    doc_stats = {}
    for row in reviewed_rows:
        sf = str(row.get("Source File", "")).strip()
        status = str(row.get("Validation Status", "")).strip()
        if not sf:
            continue
        if sf not in doc_stats:
            doc_stats[sf] = {"total": 0, "correct": 0, "errors": 0, "unreviewed": 0}
        doc_stats[sf]["total"] += 1
        if status == "Correct":
            doc_stats[sf]["correct"] += 1
        elif status.startswith("Wrong-") or status == "Missing-Content":
            doc_stats[sf]["errors"] += 1
        elif not status:
            doc_stats[sf]["unreviewed"] += 1
    analysis["per_document_stats"] = doc_stats

    return analysis


def write_feedback_report(
    stats: dict,
    analysis: dict,
    output_path: str,
) -> None:
    """Write a human-readable feedback report."""
    lines = []
    lines.append("=" * 70)
    lines.append("VALIDATION FEEDBACK REPORT")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 70)
    lines.append("")

    # --- Review Coverage ---
    lines.append("REVIEW COVERAGE")
    lines.append("-" * 40)
    lines.append(f"Original extracted controls:  {stats['total_original']}")
    lines.append(f"Confirmed (kept):             {stats['confirmed_count']}")
    lines.append(f"Excluded (false positives):   {stats['excluded_count']}")
    lines.append(f"Unreviewed (assumed correct):  {stats['unreviewed_count']}")
    lines.append(f"Missing controls added:        {stats['added_missing_count']}")
    lines.append("")

    # --- Status Breakdown ---
    lines.append("STATUS BREAKDOWN")
    lines.append("-" * 40)
    for status, count in sorted(stats["status_counts"].items(), key=lambda x: -x[1]):
        lines.append(f"  {status:30s}  {count}")
    lines.append("")

    # --- Excluded Controls ---
    if stats["excluded"]:
        lines.append("EXCLUDED CONTROLS (False Positives)")
        lines.append("-" * 40)
        for ctrl in stats["excluded"]:
            cid = ctrl.get("control_id", "")
            sf = ctrl.get("source_file", "")
            sec = ctrl.get("section_header", "")
            lines.append(f"  {cid:20s}  {sf}  [{sec}]")
        lines.append("")

    # --- Added Missing Controls ---
    if stats["added_missing"]:
        lines.append("ADDED MISSING CONTROLS")
        lines.append("-" * 40)
        for ctrl in stats["added_missing"]:
            cid = ctrl.get("control_id", "")
            sf = ctrl.get("source_file", "")
            sec = ctrl.get("section_header", "")
            lines.append(f"  {cid:20s}  {sf}  [{sec}]")
        lines.append("")

    # --- False Positive Patterns ---
    fp = analysis.get("false_positive_patterns", {})
    if fp.get("total", 0) > 0:
        lines.append("FALSE POSITIVE PATTERNS")
        lines.append("-" * 40)
        lines.append(f"Total false positives: {fp['total']}")
        if fp.get("by_section"):
            lines.append("  By section:")
            for sec, count in sorted(fp["by_section"].items(), key=lambda x: -x[1]):
                lines.append(f"    {sec:40s}  {count}")
        if fp.get("by_prefix"):
            lines.append("  By control ID prefix:")
            for prefix, count in sorted(fp["by_prefix"].items(), key=lambda x: -x[1]):
                lines.append(f"    {prefix:40s}  {count}")
        lines.append("")

    # --- Description Errors ---
    desc = analysis.get("description_errors", {})
    if isinstance(desc, dict) and desc.get("total", 0) > 0:
        lines.append("DESCRIPTION / GUIDANCE ERRORS")
        lines.append("-" * 40)
        lines.append(f"Total: {desc['total']}")
        lines.append(f"Correlated with GUIDANCE_IN_DESC flag: {desc['with_guidance_flag']} ({desc['correlation_pct']}%)")
        lines.append("")

    # --- Missing Control Patterns ---
    mp = analysis.get("missing_control_patterns", {})
    if mp.get("total", 0) > 0:
        lines.append("MISSING CONTROL PATTERNS")
        lines.append("-" * 40)
        lines.append(f"Total controls added by reviewer: {mp['total']}")
        if mp.get("by_file"):
            lines.append("  By source file:")
            for sf, count in sorted(mp["by_file"].items(), key=lambda x: -x[1]):
                lines.append(f"    {sf:50s}  {count}")
        if mp.get("by_section"):
            lines.append("  By section:")
            for sec, count in sorted(mp["by_section"].items(), key=lambda x: -x[1]):
                lines.append(f"    {sec:40s}  {count}")
        lines.append("")

    # --- Flag Accuracy ---
    flag_acc = analysis.get("flag_accuracy", {})
    if flag_acc:
        lines.append("FLAG ACCURACY (how well each flag predicts actual errors)")
        lines.append("-" * 40)
        lines.append(f"  {'Flag':35s}  {'Flagged':>8s}  {'Errors':>8s}  {'Precision':>10s}")
        for flag, data in sorted(flag_acc.items(), key=lambda x: -x[1]["precision_pct"]):
            lines.append(
                f"  {flag:35s}  {data['total']:>8d}  {data['true_positives']:>8d}  "
                f"{data['precision_pct']:>9.1f}%"
            )
        lines.append("")

    # --- Per-Document Stats ---
    doc_stats = analysis.get("per_document_stats", {})
    if doc_stats:
        lines.append("PER-DOCUMENT ACCURACY")
        lines.append("-" * 40)
        lines.append(f"  {'Source File':50s}  {'Total':>6s}  {'OK':>4s}  {'Err':>4s}  {'N/R':>4s}")
        for sf, data in sorted(doc_stats.items()):
            lines.append(
                f"  {sf:50s}  {data['total']:>6d}  {data['correct']:>4d}  "
                f"{data['errors']:>4d}  {data['unreviewed']:>4d}"
            )
        lines.append("")

    # --- Config Suggestions ---
    suggestions = analysis.get("suggestions", [])
    if suggestions:
        lines.append("=" * 70)
        lines.append("CONFIG IMPROVEMENT SUGGESTIONS")
        lines.append("=" * 70)
        for i, sug in enumerate(suggestions, 1):
            lines.append(f"\n  {i}. [{sug['type'].upper()}]")
            lines.append(f"     Reason:  {sug['reason']}")
            lines.append(f"     Action:  {sug['action']}")
            if "value" in sug:
                val = sug["value"]
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val[:10])
                lines.append(f"     Value:   {val}")
        lines.append("")
    else:
        lines.append("No config improvement suggestions — review data did not reveal clear patterns.")
        lines.append("")

    lines.append("=" * 70)
    lines.append("END OF REPORT")
    lines.append("=" * 70)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_suggested_config(suggestions: list[dict], output_path: str) -> None:
    """Write suggested config changes as a YAML file for easy reference."""
    if not suggestions:
        return

    lines = [
        "# Suggested configuration changes based on validation review feedback",
        f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "#",
        "# These are SUGGESTIONS — review each one before applying.",
        "# Apply changes to dps_config.xlsx (Control Extraction sheet) or",
        "# to validate_controls.py (SUSPECT_SECTIONS list) as noted.",
        "",
    ]

    blacklist_entries = []
    suspect_sections = []
    other_suggestions = []

    for sug in suggestions:
        if sug["type"] == "blacklist":
            blacklist_entries.append(sug["value"])
        elif sug["type"] == "suspect_section":
            suspect_sections.append(sug["value"])
        else:
            other_suggestions.append(sug)

    if blacklist_entries:
        lines.append("# Add these to control_extraction.blacklist in dps_config.xlsx:")
        lines.append("control_extraction:")
        lines.append("  blacklist_additions:")
        for entry in blacklist_entries:
            lines.append(f'    - "{entry}"')
        lines.append("")

    if suspect_sections:
        lines.append("# Add these to SUSPECT_SECTIONS in scripts/validate_controls.py:")
        lines.append("validate_controls:")
        lines.append("  suspect_section_additions:")
        for section in suspect_sections:
            lines.append(f'    - "{section}"')
        lines.append("")

    if other_suggestions:
        lines.append("# Additional suggestions (review manually):")
        for sug in other_suggestions:
            lines.append(f"# - [{sug['type']}] {sug['reason']}")
            lines.append(f"#   Action: {sug['action']}")
        lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    parser = setup_argparse(
        "Ingest human-reviewed validation_review.xlsx and produce "
        "confirmed controls, error analysis, and config suggestions"
    )
    args = parser.parse_args()

    config = load_config(args.config)
    output_cfg = config.get("output", {})
    config_dir = config.get("_config_dir", os.getcwd())

    # Resolve output root
    output_root = output_cfg.get("directory", "./output")
    if not os.path.isabs(output_root):
        output_root = os.path.normpath(os.path.join(config_dir, output_root))

    # Resolve validation directory
    validation_cfg = output_cfg.get("validation", {})
    validation_dir = os.path.join(output_root, validation_cfg.get("directory", "9 - validation"))

    # Resolve review workbook path
    review_file = validation_cfg.get("review_file", "validation_review.xlsx")
    review_path = os.path.join(validation_dir, review_file)

    if not os.path.exists(review_path):
        print(f"ERROR: Reviewed workbook not found at {review_path}")
        print("Run Step 9 first, then review the workbook before running this script.")
        return

    # Resolve original controls CSV
    controls_cfg = output_cfg.get("controls", {})
    controls_dir = os.path.join(output_root, controls_cfg.get("directory", "2 - controls"))
    controls_file = controls_cfg.get("output_file", "controls_output.csv")
    controls_csv_path = os.path.join(controls_dir, controls_file)

    if not os.path.exists(controls_csv_path):
        alt_path = controls_csv_path.replace(".xlsx", ".csv")
        if os.path.exists(alt_path):
            controls_csv_path = alt_path
        else:
            print(f"ERROR: Original controls not found at {controls_csv_path}")
            print("Step 2 (extract_controls.py) must have run first.")
            return

    # --- Load data ---
    print("  Loading reviewed workbook...")
    reviewed_rows, missing_rows = load_reviewed_workbook(review_path)
    print(f"  Found {len(reviewed_rows)} reviewed control(s), {len(missing_rows)} manually added control(s)")

    print("  Loading original controls...")
    original_controls = load_original_controls(controls_csv_path)
    print(f"  Found {len(original_controls)} original control(s)")

    # --- Build confirmed controls ---
    confirmed_path = os.path.join(validation_dir, "confirmed_controls.csv")
    print("  Building confirmed controls...")
    stats = build_confirmed_controls(reviewed_rows, missing_rows, original_controls, confirmed_path)
    print(f"  Confirmed: {stats['confirmed_count']}  Excluded: {stats['excluded_count']}  Added: {stats['added_missing_count']}")

    # --- Analyze error patterns ---
    print("  Analyzing error patterns...")
    analysis = analyze_error_patterns(reviewed_rows, missing_rows, config)

    # --- Write outputs ---
    report_path = os.path.join(validation_dir, "feedback_report.txt")
    write_feedback_report(stats, analysis, report_path)
    print(f"\n  Feedback report written to: {report_path}")

    print(f"  Confirmed controls written to: {confirmed_path}")

    suggestions = analysis.get("suggestions", [])
    if suggestions:
        yaml_path = os.path.join(validation_dir, "suggested_config_changes.yaml")
        write_suggested_config(suggestions, yaml_path)
        print(f"  Config suggestions written to: {yaml_path}")

    # --- Print summary ---
    print()
    print("=" * 60)
    print("FEEDBACK INGESTION SUMMARY")
    print("=" * 60)
    print(f"Original controls:           {stats['total_original']}")
    print(f"Confirmed (kept):            {stats['confirmed_count']}")
    print(f"  - Reviewed as Correct:     {stats['status_counts'].get('Correct', 0)}")
    print(f"  - Unreviewed (kept):       {stats['unreviewed_count']}")
    print(f"  - Missing (added):         {stats['added_missing_count']}")
    print(f"Excluded (false positives):  {stats['excluded_count']}")

    if suggestions:
        print(f"\nConfig suggestions: {len(suggestions)}")
        for sug in suggestions:
            print(f"  - [{sug['type']}] {sug['action']}")

    print(f"\nSee {report_path} for full details.")


if __name__ == "__main__":
    main()
