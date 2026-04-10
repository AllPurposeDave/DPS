"""
Shared utilities for the DPS pre-processing pipeline.

Common functions used across all pipeline scripts.
Supports both dps_config.xlsx (preferred) and dps_config_fallback.yaml (legacy fallback).

REQUIREMENTS:
  pip install python-docx openpyxl pyyaml
  Python 3.8 or later
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import sys
from pathlib import Path
from typing import Optional


def load_config(config_path: Optional[str] = None) -> dict:
    """
    Load DPS configuration from Excel (.xlsx) or YAML (.yaml).

    Search order:
      1. Explicit path passed via --config (any format)
      2. ./dps_config.xlsx  → ../dps_config.xlsx   (Excel preferred)
      3. ./dps_config_fallback.yaml  → ../dps_config_fallback.yaml   (YAML fallback)

    Returns the parsed config dict, or an empty dict if no config is found.
    """
    search_paths = []
    if config_path:
        search_paths.append(config_path)
    # Excel first, then YAML fallback
    for name in ["dps_config.xlsx", "dps_config_fallback.yaml"]:
        search_paths.append(os.path.join(os.getcwd(), name))
        search_paths.append(os.path.join(os.getcwd(), "..", name))

    for path in search_paths:
        if os.path.isfile(path):
            if path.lower().endswith(".xlsx"):
                return load_config_xlsx(path)
            else:
                import yaml
                with open(path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}
                config["_config_dir"] = os.path.dirname(os.path.abspath(path))
                return config

    return {}


# ── Excel config parser ─────────────────────────────────────────────────────

def _coerce_value(val):
    """Coerce Excel cell value to appropriate Python type."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        # Convert float-that-is-int (e.g. 36000.0 → 36000)
        if isinstance(val, float) and val == int(val) and not str(val).startswith("0."):
            return int(val)
        return val
    s = str(val).strip()
    if s == "":
        return None
    if s.upper() == "TRUE":
        return True
    if s.upper() == "FALSE":
        return False
    try:
        f = float(s)
        if f == int(f) and "." not in s:
            return int(s)
        return f
    except ValueError:
        return s


def _expand_dot_keys(flat: dict) -> dict:
    """Expand dot-notation keys into nested dicts. E.g. {'a.b': 1} → {'a': {'b': 1}}."""
    result = {}
    for key, val in flat.items():
        parts = key.split(".")
        d = result
        for part in parts[:-1]:
            d = d.setdefault(part, {})
        d[parts[-1]] = val
    return result


def _read_all_rows(ws):
    """Read all rows from a worksheet as lists of values."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _is_subheader(row):
    """Check if a row is a sub-header (starts with single # but not ##).

    Convention: '# Section Name' = sub-header (block delimiter).
    '## description text' = inline help comment (not a delimiter).
    """
    if not row or not row[0]:
        return False
    val = str(row[0]).strip()
    return val.startswith("#") and not val.startswith("##")


def _split_by_subheaders(rows):
    """Split rows into named blocks at # sub-headers. Returns list of (name, rows) tuples."""
    blocks = []
    current_name = ""
    current_rows = []
    for row in rows:
        if _is_subheader(row):
            if current_rows:
                blocks.append((current_name, current_rows))
            current_name = str(row[0]).strip().lstrip("# ").strip()
            current_rows = []
        else:
            current_rows.append(row)
    if current_rows:
        blocks.append((current_name, current_rows))
    return blocks


def _is_comment(row):
    """Check if a row is a comment/description line (starts with ## or #)."""
    if not row or not row[0]:
        return False
    val = str(row[0]).strip()
    return val.startswith("#")


def _parse_settings_rows(rows):
    """Parse Setting|Value|Description rows into a flat dict (skips sub-headers, comments, and blanks)."""
    result = {}
    for row in rows:
        if not row or not row[0] or _is_comment(row):
            continue
        key = str(row[0]).strip()
        raw = row[1] if len(row) > 1 else None
        # For settings, preserve empty strings (they're meaningful, e.g. fallback_template: "")
        if raw is None:
            continue
        if isinstance(raw, str) and raw.strip() == "":
            result[key] = ""
        else:
            val = _coerce_value(raw)
            if val is not None:
                result[key] = val
    return result


def _parse_list_column(rows, col=0):
    """Extract non-empty values from a single column, skipping sub-headers and comments."""
    result = []
    for row in rows:
        if not row or _is_comment(row):
            continue
        val = row[col] if col < len(row) else None
        if val is not None and str(val).strip():
            result.append(_coerce_value(val) if not isinstance(val, str) else str(val).strip())
    return result


def _parse_map_rows(rows, key_col=0, val_col=1):
    """Parse Key|Value rows into a dict, skipping sub-headers, comments, and blanks."""
    result = {}
    for row in rows:
        if not row or _is_comment(row):
            continue
        key = row[key_col] if key_col < len(row) else None
        val = row[val_col] if val_col < len(row) else None
        if key and str(key).strip():
            result[str(key).strip()] = _coerce_value(val) if val is not None else ""
    return result


# ── Per-sheet parsers ────────────────────────────────────────────────────────

def _parse_input_sheet(ws):
    rows = _read_all_rows(ws)[1:]  # skip header
    blocks = _split_by_subheaders(rows)
    result = {}
    exclude = []
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "directory" in name_lower or "setting" in name_lower:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
        elif "exclude" in name_lower or "pattern" in name_lower or "skip" in name_lower or "file" in name_lower:
            exclude.extend(_parse_list_column(block_rows))
    if exclude:
        result["exclude_patterns"] = exclude
    return result


def _parse_output_sheet(ws):
    rows = _read_all_rows(ws)[1:]  # skip header
    flat = _parse_settings_rows(rows)
    return _expand_dot_keys(flat)


def _parse_sections_sheet(ws):
    rows = _read_all_rows(ws)[1:]  # skip header
    result = {}
    for row in rows:
        if not row or _is_comment(row):
            continue
        category = str(row[0]).strip().lower() if row[0] else None
        keyword = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if category and keyword:
            result.setdefault(category, []).append(keyword)
    return result


def _parse_headings_sheet(ws):
    rows = _read_all_rows(ws)[1:]  # skip header
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "built-in" in name_lower or "builtin" in name_lower or "standard word" in name_lower:
            result["builtin_styles"] = _parse_list_column(block_rows)
        elif "style map" in name_lower or "maps your" in name_lower or ("map" in name_lower and "key" in name_lower):
            result["custom_style_map"] = _parse_map_rows(block_rows)
        elif ("custom" in name_lower or "org" in name_lower) and "map" not in name_lower:
            if "custom_heading_styles" not in result:
                result["custom_heading_styles"] = _parse_list_column(block_rows)
        elif "fake" in name_lower:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
        elif "pattern" in name_lower or "level" in name_lower:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
    return result


def _parse_text_deletions_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "setting" in name_lower:
            result.update(_parse_settings_rows(block_rows))
        elif "section" in name_lower:
            # Parse Section Heading | Delete (TRUE/FALSE) | Description rows
            section_deletions = []
            for row in block_rows:
                if not row or _is_comment(row):
                    continue
                heading = row[0] if row[0] else None
                if heading is None or not str(heading).strip():
                    continue
                heading = str(heading).strip()
                # Skip the sub-table column header row
                if heading.lower() == "section heading":
                    continue
                delete = True
                if len(row) > 1 and row[1] is not None:
                    val = _coerce_value(row[1])
                    if isinstance(val, bool):
                        delete = val
                    elif isinstance(val, str) and val.strip().upper() == "FALSE":
                        delete = False
                desc = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                section_deletions.append({
                    "heading": heading,
                    "delete": delete,
                    "description": desc,
                })
            result["section_deletions"] = section_deletions
        elif "phrase" in name_lower:
            result["phrases"] = _parse_list_column(block_rows)
    # Ensure keys exist even if empty
    result.setdefault("phrases", [])
    result.setdefault("section_deletions", [])
    result.setdefault("remove_table_of_content", False)
    result.setdefault("remove_headers_footers", False)
    result.setdefault("remove_revision_tables", False)
    result.setdefault("flatten_definition_tables", False)
    return result


def _parse_cross_references_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "detection" in name_lower and "setting" in name_lower:
            result.update(_parse_settings_rows(block_rows))
        elif "extraction" in name_lower or "phrase" in name_lower:
            patterns = []
            for row in block_rows:
                if row and row[0] and not _is_comment(row):
                    phrase = str(row[0]).strip()
                    ptype = str(row[1]).strip() if len(row) > 1 and row[1] else "internal"
                    if phrase:
                        patterns.append({"phrase": phrase, "type": ptype})
            if patterns:
                result["extraction_patterns"] = patterns
        elif "profiler" in name_lower or "counting" in name_lower:
            if "profiler_patterns" not in result:
                result["profiler_patterns"] = _parse_list_column(block_rows)
        elif "keyword" in name_lower or "document name" in name_lower:
            result["document_name_keywords"] = _parse_list_column(block_rows)
    return result


def _parse_tables_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    classification = {}
    for row in rows:
        if not row or _is_comment(row):
            continue
        ttype = str(row[0]).strip() if row[0] else None
        keyword = str(row[1]).strip() if len(row) > 1 and row[1] else None
        min_cols = _coerce_value(row[2]) if len(row) > 2 and row[2] else 2
        if ttype and keyword:
            entry = classification.setdefault(ttype, {"keywords": [], "min_columns": min_cols})
            entry["keywords"].append(keyword)
            entry["min_columns"] = min_cols
    return {"classification": classification}


def _parse_classification_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "type a" in name_lower:
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["type_a"] = flat.get("type_a", settings)
        elif "type b" in name_lower:
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["type_b"] = flat.get("type_b", settings)
        elif "type c" in name_lower or "procedure" in name_lower or "keyword" in name_lower:
            keywords = _parse_list_column(block_rows)
            if keywords:
                result["type_c"] = {"procedure_keywords": keywords}
        elif "type d" in name_lower:
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["type_d"] = flat.get("type_d", settings)
    return result


def _parse_profiling_flags_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    flat = _parse_settings_rows(rows)
    return _expand_dot_keys(flat)


def _parse_thresholds_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    return _parse_settings_rows(rows)


def _parse_priority_scoring_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "weight" in name_lower:
            result["weights"] = _parse_settings_rows(block_rows)
        elif "usage" in name_lower or "frequency" in name_lower:
            freq = {}
            for row in block_rows:
                if not row or _is_comment(row):
                    continue
                fname = str(row[0]).strip() if row[0] else ""
                val = _coerce_value(row[1] if len(row) > 1 else None)
                if fname and val is not None:
                    freq[fname] = val
            result["usage_frequency"] = freq
    result.setdefault("usage_frequency", {})
    return result


def _parse_search_terms_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        if "setting" in name.lower():
            result.update(_parse_settings_rows(block_rows))
        elif "term" in name.lower():
            result["terms"] = _parse_list_column(block_rows)
    return result


def _parse_control_extraction_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "common" in name_lower and "setting" in name_lower:
            result.update(_parse_settings_rows(block_rows))
        elif "control id" in name_lower or ("regex" in name_lower and "pattern" in name_lower):
            result["control_id_patterns"] = _parse_list_column(block_rows)
        elif "whitelist" in name_lower or "blacklist" in name_lower:
            wl = []
            bl = []
            for row in block_rows:
                if not row or _is_comment(row) or not row[0]:
                    continue
                val = str(row[0]).strip()
                if val.startswith("whitelist:"):
                    wl.append(val.split(":", 1)[1].strip())
                elif val.startswith("blacklist:"):
                    bl.append(val.split(":", 1)[1].strip())
            result["whitelist"] = wl
            result["blacklist"] = bl
        elif "guidance" in name_lower or "boundary" in name_lower:
            if "guidance_keywords" not in result:
                result["guidance_keywords"] = _parse_list_column(block_rows)
        elif "trigger" in name_lower and ("metadata" in name_lower or "category" in name_lower):
            triggers = {}
            for row in block_rows:
                if not row or _is_comment(row) or not row[0]:
                    continue
                cat = str(row[0]).strip()
                kw = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if cat and kw:
                    triggers.setdefault(cat, []).append(kw)
            if triggers:
                result["metadata_triggers"] = triggers
        elif "heading detection" in name_lower or "advanced" in name_lower:
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["heading_detection"] = flat.get("heading_detection", settings)
        elif "implementation trigger" in name_lower:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
        else:
            # Catch remaining settings
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
    # Ensure list keys default to empty
    result.setdefault("whitelist", [])
    result.setdefault("blacklist", [])
    return result


def _parse_pipeline_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    steps = []
    for row in rows:
        if not row or _is_comment(row):
            continue
        if row[1] is None and row[2] is None:
            continue
        steps.append({
            "name": str(row[1]).strip() if row[1] else "",
            "script": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "enabled": _coerce_value(row[3]) if len(row) > 3 else True,
            "description": str(row[4]).strip() if len(row) > 4 and row[4] else "",
        })
    return {"steps": steps}


def _parse_metadata_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "general" in name_lower:
            result.update(_parse_settings_rows(block_rows))
        elif ("key" in name_lower and "label" in name_lower) or name_lower == "metadata fields":
            # Metadata fields table: Key | Label | Enabled | Source | Value
            fields = []
            for row in block_rows:
                if not row or _is_comment(row) or not row[0]:
                    continue
                field = {"key": str(row[0]).strip()}
                if len(row) > 1 and row[1]:
                    field["label"] = str(row[1]).strip()
                if len(row) > 2 and row[2] is not None:
                    field["enabled"] = _coerce_value(row[2])
                if len(row) > 3 and row[3]:
                    field["source"] = str(row[3]).strip()
                if len(row) > 4 and row[4]:
                    field["value"] = str(row[4]).strip()
                fields.append(field)
            result["fields"] = fields
        elif "url" in name_lower and ("resolution" in name_lower or "lookup" in name_lower):
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["url"] = flat.get("url", settings)
        elif "advanced" in name_lower:
            result.update(_parse_settings_rows(block_rows))
        elif "tag generation" in name_lower:
            settings = _parse_settings_rows(block_rows)
            flat = _expand_dot_keys(settings)
            result["tags"] = flat.get("tags", settings)
        elif "static" in name_lower or ("tag" in name_lower and "all" in name_lower):
            tags = _parse_list_column(block_rows)
            result.setdefault("tags", {})["static_tags"] = tags
    return result


def _parse_docx2md_sheet(ws):
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if ("metadata field" in name_lower and "name" in name_lower) or name_lower == "metadata frontmatter":
            # Block may contain settings before the ## table header, then field rows after
            settings_rows = []
            field_rows = []
            past_table_header = False
            for row in block_rows:
                if row and row[0] and str(row[0]).startswith("##"):
                    past_table_header = True
                    continue
                if past_table_header:
                    field_rows.append(row)
                else:
                    settings_rows.append(row)
            if settings_rows:
                result.update(_parse_settings_rows(settings_rows))
            fields = []
            for row in field_rows:
                if not row or not row[0] or str(row[0]).startswith("#"):
                    continue
                field = {"name": str(row[0]).strip()}
                if len(row) > 1 and row[1]:
                    field["source"] = str(row[1]).strip()
                if len(row) > 2 and row[2] is not None and str(row[2]).strip():
                    field["default"] = str(row[2]).strip()
                fields.append(field)
            if fields:
                result["metadata_fields"] = fields
        else:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
    return result


def _parse_acronym_finder_sheet(ws):
    """Parse the Acronym Finder sheet into the same dict structure as the YAML config."""
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for name, block_rows in blocks:
        name_lower = name.lower()
        if "search" in name_lower:
            result["search"] = _parse_settings_rows(block_rows)
        elif "pattern" in name_lower:
            result["patterns"] = _parse_settings_rows(block_rows)
        elif "ignore" in name_lower:
            result["ignore_list"] = _parse_list_column(block_rows, col=0)
        else:
            settings = _parse_settings_rows(block_rows)
            result.update(settings)
    return result


def _parse_docx2jsonl_sheet(ws):
    """Parse the Docx2jsonl sheet into the same dict structure as the YAML config."""
    rows = _read_all_rows(ws)[1:]
    blocks = _split_by_subheaders(rows)
    result = {}
    for _name, block_rows in blocks:
        settings = _parse_settings_rows(block_rows)
        result.update(settings)
    return result


# ── Main Excel loader ────────────────────────────────────────────────────────

def load_config_xlsx(xlsx_path: str) -> dict:
    """
    Load configuration from a dps_config.xlsx workbook.

    Returns the same nested dict structure as loading from dps_config.yaml,
    so all downstream scripts work without changes.
    """
    from openpyxl import load_workbook as _load_wb

    wb = _load_wb(xlsx_path, read_only=True, data_only=True)

    # Map sheet names to parser functions
    sheet_parsers = {
        "Input": ("input", _parse_input_sheet),
        "Output": ("output", _parse_output_sheet),
        "Sections": ("sections", _parse_sections_sheet),
        "Headings": ("headings", _parse_headings_sheet),
        "Text Deletions": ("text_deletions", _parse_text_deletions_sheet),
        "Cross References": ("cross_references", _parse_cross_references_sheet),
        "Tables": ("tables", _parse_tables_sheet),
        "Classification": ("classification", _parse_classification_sheet),
        "Profiling Flags": ("profiling_flags", _parse_profiling_flags_sheet),
        "Thresholds": ("thresholds", _parse_thresholds_sheet),
        "Priority Scoring": ("priority_scoring", _parse_priority_scoring_sheet),
        "Search Terms": ("search_terms", _parse_search_terms_sheet),
        "Control Extraction": ("control_extraction", _parse_control_extraction_sheet),
        "Pipeline": ("pipeline", _parse_pipeline_sheet),
        "Metadata": ("metadata", _parse_metadata_sheet),
        "Docx2md": ("docx2md", _parse_docx2md_sheet),
        "Docx2jsonl": ("docx2jsonl", _parse_docx2jsonl_sheet),
        "Acronym Finder": ("acronym_finder", _parse_acronym_finder_sheet),
    }

    config = {}
    for sheet_name, (config_key, parser_fn) in sheet_parsers.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                config[config_key] = parser_fn(ws)
            except Exception as e:
                print(f"  WARNING: Error parsing sheet '{sheet_name}': {e}")
                config[config_key] = {}

    wb.close()

    # Store config directory for relative path resolution
    config["_config_dir"] = os.path.dirname(os.path.abspath(xlsx_path))
    return config


def resolve_path(config: dict, relative_path: str) -> str:
    """
    Resolve a relative path from the config against the config file's directory.
    If the path is already absolute, return it as-is.
    """
    if os.path.isabs(relative_path):
        return relative_path
    config_dir = config.get("_config_dir", os.getcwd())
    return os.path.normpath(os.path.join(config_dir, relative_path))


def normalize_doc_name(name: str) -> str:
    """Normalize a document name for matching.

    Strips extension, lowercases, and collapses underscores/hyphens/spaces
    into single spaces so that 'POL-AC-2026-001', 'POL_AC_2026_001', and
    'POL AC 2026 001' all produce the same key.
    """
    stem = os.path.splitext(name)[0]
    return re.sub(r"[_\-\s]+", " ", stem).strip().lower()


def match_doc_name(needle: str, haystack_key: str) -> bool:
    """Check whether two document names refer to the same document.

    Both values are normalized before comparison.  Returns True on exact
    match *or* substring containment (either direction) so that split
    sub-files like 'PolicyDoc - Section 3' still match 'PolicyDoc'.
    """
    a = normalize_doc_name(needle)
    b = normalize_doc_name(haystack_key)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def get_input_dir(config: dict, cli_override: Optional[str] = None) -> str:
    """
    Get the input directory path, with CLI override taking priority.

    Priority: CLI arg > config input.directory > ./input
    """
    if cli_override:
        return cli_override
    input_dir = config.get("input", {}).get("directory", "./input")
    return resolve_path(config, input_dir)


def get_output_dir(config: dict, step_key: str, cli_override: Optional[str] = None) -> str:
    """
    Get the output directory for a specific pipeline step.

    Args:
        config: Parsed config dict
        step_key: Key in config output section (e.g., "profiler", "cross_references")
        cli_override: CLI argument that takes priority over config

    Priority: CLI arg > config output.<step_key>.directory > ./output
    """
    if cli_override:
        return cli_override
    output_root = config.get("output", {}).get("directory", "./output")
    step_subdir = config.get("output", {}).get(step_key, {}).get("directory", "")
    full_path = os.path.join(output_root, step_subdir) if step_subdir else output_root
    return resolve_path(config, full_path)


def setup_argparse(description: str) -> argparse.ArgumentParser:
    """
    Create an argument parser with standard --config, input_dir, and output_dir args.

    All scripts now support:
      python <script>.py --config ../dps_config.yaml         # config-driven
      python <script>.py <input_dir> <output_dir>            # standalone (legacy)
      python <script>.py --config ../dps_config.yaml ./docs  # config + override

    FAILURE POINT: Passing a file path instead of a directory will cause
    downstream glob calls to find nothing and exit with "No .docx files found".
    """
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--config",
        default=None,
        help="Path to dps_config.xlsx or dps_config.yaml (auto-detected if omitted)",
    )
    parser.add_argument(
        "input_dir",
        nargs="?",
        default=None,
        help="Directory containing .docx files (overrides config)",
    )
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=None,
        help="Directory for output files (overrides config)",
    )
    return parser


def iter_docx_files(input_dir: str, config: Optional[dict] = None,
                    exclude_override: Optional[list] = None) -> list:
    """
    Return sorted list of .docx file paths in input_dir.
    Skips files matching exclude patterns from config.

    NOTE: Only scans the top-level directory by default.
    Set input.recursive: true in config to scan sub-folders.

    TWEAK: To scan sub-folders without config, change the pattern to
    include "**" and add recursive=True to glob.glob().

    FAILURE POINT: Returns empty list if input_dir does not exist or
    contains no .docx files. The calling script will print a warning and exit.
    """
    config = config or {}
    input_cfg = config.get("input", {})
    file_pattern = input_cfg.get("pattern", "*.docx")
    recursive = input_cfg.get("recursive", False)
    exclude_patterns = exclude_override if exclude_override is not None else input_cfg.get("exclude_patterns", ["~$"])

    if recursive:
        pattern = os.path.join(input_dir, "**", file_pattern)
        files = glob.glob(pattern, recursive=True)
    else:
        pattern = os.path.join(input_dir, file_pattern)
        files = glob.glob(pattern)

    # Filter out excluded files
    filtered = []
    for f in files:
        basename = os.path.basename(f)
        skip = False
        for exc in exclude_patterns:
            if exc.lower() in basename.lower():
                skip = True
                break
        if not skip:
            filtered.append(f)

    filtered.sort()
    return filtered


def is_heading_style(style) -> bool:
    """
    Check if a paragraph style is a standard Word heading style (Heading 1-9).

    IMPORTANT: This only recognises the exact names "Heading 1" through "Heading 9".
    Custom styles like "Policy Heading 1" return False here — they are handled
    separately by the custom_style_map in dps_config.yaml.

    TWEAK: To also accept "Title" as a heading, change the regex to:
        r"^(Heading \d|Title)$"
    """
    if style is None:
        return False
    name = style.name if hasattr(style, "name") else str(style)
    return bool(re.match(r"^Heading \d$", name))


def get_heading_level(style) -> Optional[int]:
    """
    Return the heading level (1-9) if the style is a standard heading, else None.

    Returns None for custom styles, Normal, Body Text, etc.
    Used by section_splitter.py to find split points.
    """
    if style is None:
        return None
    name = style.name if hasattr(style, "name") else str(style)
    match = re.match(r"^Heading (\d)$", name)
    if match:
        return int(match.group(1))
    return None


def build_custom_style_map(config: dict) -> dict:
    """Build a lowercase-keyed custom style map from config.

    Maps custom Word style names (lowercased) to heading levels.
    Used by docx2md.py and docx2jsonl.py for heading detection.
    """
    hcfg = config.get("headings", {})
    cmap = hcfg.get("custom_style_map", {})
    if cmap:
        return {k.strip().lower(): v for k, v in cmap.items()}
    return {}


def find_parent_heading(paragraphs, index: int) -> str:
    """
    Walk backward from paragraph at `index` to find the nearest
    Heading 1 or Heading 2 paragraph. Returns the heading text,
    or "(No heading found)" if none exists above.

    Used by cross_reference_extractor.py to label which section a
    cross-reference appears in.

    TWEAK: Change `level <= 2` to `level <= 3` to also use Heading 3
    paragraphs as parent section labels.
    """
    for i in range(index - 1, -1, -1):
        level = get_heading_level(paragraphs[i].style)
        if level is not None and level <= 2:
            return paragraphs[i].text.strip()
    return "(No heading found)"


def is_paragraph_bold(paragraph) -> bool:
    """
    Check if a paragraph is bold. Checks both paragraph-level formatting
    and run-level formatting. A paragraph counts as bold if:
    - The paragraph font (via its style) is bold, OR
    - ALL runs in the paragraph are bold (and there is at least one run)

    FAILURE POINT: Mixed bold paragraphs (some runs bold, some not) return
    False because the ALL-runs check fails.

    TWEAK: To treat a paragraph as bold if ANY run is bold (more permissive),
    change the last line to:
        return any(run.bold for run in runs)
    """
    # Check paragraph-level bold (set via the style's font, not individual runs)
    if paragraph.paragraph_format and hasattr(paragraph, "style"):
        pf = paragraph.style.font if hasattr(paragraph.style, "font") else None
        if pf and pf.bold:
            return True

    # Check run-level bold: all runs must be bold
    runs = paragraph.runs
    if not runs:
        return False
    return all(run.bold for run in runs)


def sanitize_filename(text: str, max_len: int = 50) -> str:
    """
    Clean text for safe use as a filename component.
    Removes special characters, collapses whitespace, truncates.

    TWEAK: Adjust max_len if output filenames are being truncated too aggressively.
        sanitize_filename(text, max_len=80)  # allow longer names

    FAILURE POINT: If two headings produce the same sanitized name,
    section_splitter.py will silently overwrite the first output file.
    Check split_manifest.csv for duplicate sub_doc_filename values.
    """
    clean = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", text)
    clean = re.sub(r"\s+", " ", clean).strip()
    if len(clean) > max_len:
        clean = clean[:max_len].rstrip()
    return clean



def ensure_output_dir(output_dir: str) -> None:
    """
    Create the output directory if it doesn't exist.
    Safe to call even if the directory already exists (exist_ok=True).

    FAILURE POINT: If the path is on a network drive or read-only location,
    os.makedirs will raise PermissionError. Run from a local drive if possible.
    """
    os.makedirs(output_dir, exist_ok=True)


# ---------------------------------------------------------------------------
# Consolidated Excel report utilities
# ---------------------------------------------------------------------------

def log_pipeline_issue(
    output_root: str,
    step_name: str,
    filename: str,
    issue_type: str,
    message: str,
) -> None:
    """
    Append one row to pipeline_issues.csv in the output root directory.

    This file accumulates issues from all pipeline steps and is included as
    the "Pipeline Issues" sheet in the consolidated DPS report.

    Args:
        output_root: The root output directory (e.g. "./output").
        step_name:   Name of the pipeline step (e.g. "Step 1 - Controls").
        filename:    The source .docx filename that had the issue (basename only).
        issue_type:  Short category: "ERROR", "WARNING", or "PASSWORD_PROTECTED".
        message:     Human-readable description of the issue.
    """
    import csv as csv_mod
    from datetime import datetime

    os.makedirs(output_root, exist_ok=True)
    csv_path = os.path.join(output_root, "pipeline_issues.csv")
    fieldnames = ["timestamp", "step", "filename", "issue_type", "message"]
    file_exists = os.path.isfile(csv_path) and os.path.getsize(csv_path) > 0

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv_mod.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "step": step_name,
            "filename": filename,
            "issue_type": issue_type,
            "message": message,
        })


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name (max 31 chars)."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    return name[:31]


def add_csv_as_sheet(wb, csv_path: str, sheet_name: str) -> bool:
    """
    Read a CSV file and add it as a styled worksheet in an openpyxl Workbook.

    Returns True if the sheet was added, False if the CSV is missing or empty.
    Light styling: blue header row, freeze pane, autofilter, auto column widths.
    """
    import csv as csv_mod
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    csv_path = Path(csv_path)
    if not csv_path.exists() or csv_path.stat().st_size == 0:
        return False

    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv_mod.reader(f)
        rows = list(reader)

    if not rows:
        return False

    safe_name = _sanitize_sheet_name(sheet_name)
    ws = wb.create_sheet(title=safe_name)

    # Write all rows
    for row in rows:
        ws.append(row)

    # --- Light styling ---
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Freeze top row and autofilter
    ws.freeze_panes = "A2"
    if ws.max_column and ws.max_row:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Auto column widths; wrap text in columns whose content exceeds the threshold
    WRAP_THRESHOLD = 50
    wrap_col_indices = set()
    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is not None:
                max_width = max(max_width, len(str(val)))
        letter = get_column_letter(col_idx)
        if max_width > WRAP_THRESHOLD:
            ws.column_dimensions[letter].width = 60
            wrap_col_indices.add(col_idx)
        else:
            ws.column_dimensions[letter].width = min(max_width + 2, 60)

    if wrap_col_indices:
        wrap_align = Alignment(vertical="top", wrap_text=True)
        for data_row in ws.iter_rows(min_row=2):
            for cell in data_row:
                if cell.column in wrap_col_indices:
                    cell.alignment = wrap_align

    return True


def add_xlsx_as_sheet(wb, xlsx_path: str, sheet_name: str, source_sheet_name: Optional[str] = None) -> bool:
    """
    Read an Excel file and copy its data as a styled worksheet in an openpyxl Workbook.
    Copies the first sheet by default, or a named sheet if source_sheet_name is given.
    Applies the same styling as CSV sheets.

    Returns True if the sheet was added, False if the Excel file is missing or empty.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False

    try:
        source_wb = load_workbook(xlsx_path)
        if source_sheet_name and source_sheet_name in source_wb.sheetnames:
            source_ws = source_wb[source_sheet_name]
        else:
            source_ws = source_wb.active
        if source_ws is None or source_ws.max_row == 0:
            return False
    except Exception:
        return False

    safe_name = _sanitize_sheet_name(sheet_name)
    ws = wb.create_sheet(title=safe_name)

    # Copy all rows from source sheet
    for row in source_ws.iter_rows():
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        ws.append(new_row)

    # --- Light styling ---
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Freeze top row and autofilter
    ws.freeze_panes = "A2"
    if ws.max_column and ws.max_row:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Auto column widths; wrap text in columns whose content exceeds the threshold
    WRAP_THRESHOLD = 50
    wrap_col_indices = set()
    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is not None:
                max_width = max(max_width, len(str(val)))
        letter = get_column_letter(col_idx)
        if max_width > WRAP_THRESHOLD:
            ws.column_dimensions[letter].width = 60
            wrap_col_indices.add(col_idx)
        else:
            ws.column_dimensions[letter].width = min(max_width + 2, 60)

    if wrap_col_indices:
        wrap_align = Alignment(vertical="top", wrap_text=True)
        for data_row in ws.iter_rows(min_row=2):
            for cell in data_row:
                if cell.column in wrap_col_indices:
                    cell.alignment = wrap_align

    return True


def build_consolidated_workbook(config: dict, timestamp_str: str) -> Optional[str]:
    """
    Build a single Excel workbook with one sheet per pipeline CSV output.

    Args:
        config: Parsed dps_config.yaml dict.
        timestamp_str: Timestamp for the filename (e.g. "2026-03-23_143052").

    Returns the saved file path on success, None on failure.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  WARNING: openpyxl not installed — skipping consolidated Excel report.")
        print("  FIX: pip install openpyxl")
        return None

    output_cfg = config.get("output", {})
    output_root = output_cfg.get("directory", "./output")
    config_dir = config.get("_config_dir", os.getcwd())
    if not os.path.isabs(output_root):
        output_root = os.path.normpath(os.path.join(config_dir, output_root))

    # Sheet manifest: (sheet_name, config_step_key, filename_config_key, source_sheet_name)
    # source_sheet_name is optional — used when pulling a specific sheet from a
    # multi-sheet .xlsx (e.g. the validation review workbook).
    sheet_manifest = [
        ("0 - Document Inventory", "profiler",          "inventory_file",  None),
        ("0 - Sections",           "profiler",          "sections_file",   None),
        ("0 - Tables",             "profiler",          "tables_file",     None),
        ("0 - CrossRefs",          "profiler",          "crossrefs_file",  None),
        ("1 - Controls",           "controls",          "output_file",     None),
        ("2 - Cross References",   "cross_references",  "output_file",     None),
        ("3 - Heading Changes",    "heading_fixes",     "changes_file",    None),
        ("4 - Split Manifest",     "split_documents",   "manifest_file",   None),
        ("5 - Metadata",           "metadata",          "manifest_file",   None),
        ("6 - Validation",         "validation",        "output_file",     None),
        ("6 - Validation Review",  "validation",        "review_file",     "Validation Review"),
        ("6 - Validation Summary", "validation",        "review_file",     "Summary"),
    ]

    wb = Workbook()
    # Remove the default empty sheet (will be re-added if no CSVs found)
    wb.remove(wb.active)

    sheets_added = 0

    # Pipeline Issues sheet — always first so problems are immediately visible
    issues_path = os.path.join(output_root, "pipeline_issues.csv")
    if add_csv_as_sheet(wb, issues_path, "⚠ Pipeline Issues"):
        sheets_added += 1

    for sheet_name, step_key, file_key, source_sheet in sheet_manifest:
        step_cfg = output_cfg.get(step_key, {})
        step_dir = step_cfg.get("directory", "")
        filename = step_cfg.get(file_key, "")
        file_path = os.path.join(output_root, step_dir, filename)

        # Determine file type and use appropriate function
        if filename.endswith(".xlsx"):
            if add_xlsx_as_sheet(wb, file_path, sheet_name, source_sheet):
                sheets_added += 1
        elif filename.endswith(".csv"):
            if add_csv_as_sheet(wb, file_path, sheet_name):
                sheets_added += 1

    if sheets_added == 0:
        print("  WARNING: No CSV files found — skipping consolidated Excel report.")
        return None

    report_cfg = output_cfg.get("consolidated_report", {})
    prefix = report_cfg.get("filename_prefix", "DPS_Report")
    filename = f"{prefix}_{timestamp_str}.xlsx"
    save_path = os.path.join(output_root, filename)

    try:
        wb.save(save_path)
        return save_path
    except PermissionError:
        print(f"  WARNING: Could not save {filename} — file may be open in another app.")
        return None
