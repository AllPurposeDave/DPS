#!/usr/bin/env python3
"""
Generate dps_config.xlsx — the Excel-based configuration template for DPS.

Usage:
  python generate_config_template.py                        # fresh template with defaults
  python generate_config_template.py --from-yaml dps_config_fallback.yaml  # populate from existing YAML
  python generate_config_template.py -o my_config.xlsx      # custom output path

The generated workbook has one sheet per configuration section, with
Setting/Value/Description columns, sub-headers, and data-validation dropdowns
for boolean and enum fields.
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Styling constants (matches existing DPS Excel styling) ──────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10, color="2F5496")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DESC_FONT = Font(name="Arial", size=9, color="666666")
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)


# ── Helper functions ────────────────────────────────────────────────────────

def _style_headers(ws, headers: List[str]):
    """Write header row and apply standard styling."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _add_subheader(ws, text: str, ncols: int = 3):
    """Add a sub-header row (# prefixed in column A, merged visually)."""
    row = [f"# {text}"] + [""] * (ncols - 1)
    ws.append(row)
    r = ws.max_row
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER


def _add_setting(ws, setting: str, value: Any, description: str = ""):
    """Add a Setting | Value | Description row."""
    ws.append([setting, value, description])
    # Style the description cell
    r = ws.max_row
    ws.cell(row=r, column=3).font = DESC_FONT


def _add_list_rows(ws, values: List[Any], descriptions: Optional[List[str]] = None):
    """Add Value | Description rows for list items."""
    for i, val in enumerate(values):
        desc = descriptions[i] if descriptions and i < len(descriptions) else ""
        ws.append([val, desc])
        ws.cell(row=ws.max_row, column=2).font = DESC_FONT


def _add_map_rows(ws, mapping: Dict[str, Any], descriptions: Optional[Dict[str, str]] = None):
    """Add Key | Value | Description rows for mappings."""
    for key, val in mapping.items():
        desc = descriptions.get(key, "") if descriptions else ""
        ws.append([key, val, desc])
        ws.cell(row=ws.max_row, column=3).font = DESC_FONT


def _add_bool_validation(ws, col_letter: str, min_row: int, max_row: int):
    """Add TRUE/FALSE dropdown validation to a column range."""
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv.error = "Please select TRUE or FALSE"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


def _add_enum_validation(ws, col_letter: str, row: int, options: List[str]):
    """Add enum dropdown validation to a specific cell."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = f"Please select one of: {', '.join(options)}"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{row}")


def _auto_column_widths(ws):
    """Auto-fit column widths with wrapping for wide columns."""
    WRAP_THRESHOLD = 50
    wrap_cols = set()
    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is not None:
                max_width = max(max_width, len(str(val)))
        letter = get_column_letter(col_idx)
        if max_width > WRAP_THRESHOLD:
            ws.column_dimensions[letter].width = 65
            wrap_cols.add(col_idx)
        else:
            ws.column_dimensions[letter].width = min(max_width + 4, 65)
    if wrap_cols:
        for data_row in ws.iter_rows(min_row=2):
            for cell in data_row:
                if cell.column in wrap_cols:
                    cell.alignment = WRAP_ALIGN


def _finalize_sheet(ws):
    """Apply auto-widths and autofilter."""
    _auto_column_widths(ws)
    if ws.max_column and ws.max_row:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


# ── Sheet builders ──────────────────────────────────────────────────────────

def _build_readme(wb):
    """Combined README — overview, quick-start checklist, sheet reference, and pipeline details."""
    ws = wb.create_sheet("README")
    TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
    SECTION_FONT = Font(name="Arial", bold=True, size=11, color="2F5496")
    SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    NOTE_FONT = Font(name="Arial", italic=True, size=9, color="666666")
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)

    def section_header(text):
        ws.append([text, "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER

    def body_row(a="", b="", c=""):
        ws.append([a, b, c])
        r = ws.max_row
        for col in range(1, 4):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).alignment = WRAP_ALIGN

    def note_row(text):
        ws.append([text, "", ""])
        ws.cell(row=ws.max_row, column=1).font = NOTE_FONT

    # Title
    ws.append(["DPS Configuration Workbook", "", ""])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])

    # Overview
    section_header("Overview")
    body_row(
        "This workbook is the central configuration file for the Document Processing System (DPS) pipeline.",
    )
    body_row(
        "Each worksheet controls a specific area of pipeline behavior. Edit the Value column on any sheet "
        "to customise settings; leave a cell blank to use the built-in default.",
    )
    ws.append([])

    # Quick-Start Checklist
    section_header("Quick-Start Checklist — Review Before First Run")
    ws.append(["Setting", "Where to Find It", "Default"])
    r = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row=r, column=col)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

    checklist = [
        ("Input directory",      "Input sheet  >  'directory' row",       "./input"),
        ("Output directory",     "Output sheet  >  'directory' row",      "./output"),
        ("Steps to run",         "Pipeline sheet  >  'Enabled' column",   "All enabled"),
        ("File pattern",         "Input sheet  >  'pattern' row",         "*.docx"),
        ("Max chunk size",       "Thresholds sheet  >  'max_characters'", "36 000 chars (~20 pages)"),
    ]
    for setting, location, default in checklist:
        body_row(setting, location, default)

    note_row("All other settings ship with sensible defaults. Adjust only when your environment requires it.")
    ws.append([])

    # How to Use This Workbook
    section_header("How to Use This Workbook")
    ws.append(["Guideline", "Details", ""])
    r = ws.max_row
    for col in range(1, 3):
        ws.cell(row=r, column=col).font = BOLD_FONT
        ws.cell(row=r, column=col).border = THIN_BORDER

    guidelines = [
        ("Column layout",       "Settings sheets use Setting | Value | Description columns. List sheets use Value | Description."),
        ("Sub-headers",         "Rows starting with '#' are section dividers. Do not delete or modify them."),
        ("Boolean fields",      "Use the TRUE / FALSE dropdown. Manual text entry is not accepted."),
        ("Blank values",        "Leaving a Value cell empty tells the pipeline to use its built-in default."),
        ("Adding list items",   "Insert a new row in the appropriate section and fill in the Value column."),
    ]
    for guideline, details in guidelines:
        body_row(guideline, details)
    ws.append([])

    # Pipeline Execution Order
    section_header("Pipeline Execution Order")
    ws.append(["Step", "Script", "Description"])
    r = ws.max_row
    for col in range(1, 4):
        ws.cell(row=r, column=col).font = BOLD_FONT
        ws.cell(row=r, column=col).border = THIN_BORDER

    steps = [
        ("0 - Document Profiler",         "policy_profiler.py",          "Scans all documents, extracts metadata, classifies document types, scores priority, and counts words."),
        ("1 - Acronym Finder",            "acronym_finder.py",           "Identifies acronym candidates and generates an audit workbook for human review. Confirmed acronyms feed into the metadata tagging stage."),
        ("2 - Control Extractor",         "extract_controls.py",         "Extracts structured control data from compliance documents into Excel using regex and heuristics."),
        ("3 - Cross-Reference Extractor", "cross_reference_extractor.py", "Captures cross-references including section references, URLs, and internal document heading links."),
        ("4 - Heading Style Fixer",       "heading_style_fixer.py",      "Converts visually bold text that appears to be a heading into proper Word heading styles."),
        ("5 - Section Splitter",          "section_splitter.py",         "Splits documents at Heading 1 boundaries once the character limit is exceeded."),
        ("6 - Metadata Injector",         "add_metadata.py",            "Stamps each sub-document with identity and classification metadata."),
        ("7 - DOCX to Markdown",          "docx2md.py",                 "Converts Word documents to clean Markdown with YAML front matter."),
        ("8 - DOCX to JSONL",             "docx2jsonl.py",              "Converts Word documents to chunked JSONL for RAG / vector-database ingestion."),
        ("9 - Control Validator",         "validate_controls.py",       "Validates that all controls extracted in Step 2 are present in the Step 5 split documents."),
    ]
    for step, script, desc in steps:
        body_row(step, script, desc)
    ws.append([])

    # Sheet Reference
    section_header("Sheet Reference")
    body_row(
        "Each worksheet below controls settings consumed by one or more pipeline scripts. "
        "The Reference Script column indicates which script reads the sheet at runtime.",
    )
    ws.append(["Sheet Name", "Reference Script", "Purpose"])
    r = ws.max_row
    for col in range(1, 4):
        ws.cell(row=r, column=col).font = BOLD_FONT
        ws.cell(row=r, column=col).border = THIN_BORDER

    sheets = [
        ("Input",              "run_pipeline.py",              "Input folder, file pattern, and exclude patterns."),
        ("Output",             "run_pipeline.py",              "Output folder structure and filenames for each step."),
        ("Pipeline",           "run_pipeline.py",              "Enable or disable individual pipeline steps."),
        ("Thresholds",         "section_splitter.py",          "Chunk-size limits and page-estimation parameters."),
        ("Sections",           "policy_profiler.py",           "Keywords used to classify Purpose, Scope, and Controls sections."),
        ("Headings",           "heading_style_fixer.py",       "Heading styles, custom style maps, and fake-heading detection rules."),
        ("Text Deletions",     "heading_style_fixer.py",       "Phrases or entire sections to strip during processing."),
        ("Cross References",   "cross_reference_extractor.py", "Cross-reference detection patterns."),
        ("Tables",             "policy_profiler.py",           "Table-classification keywords."),
        ("Classification",     "policy_profiler.py",           "Type A / B / C / D classification thresholds."),
        ("Profiling Flags",    "policy_profiler.py",           "Flags for control-dense and heading-variance documents."),
        ("Priority Scoring",   "policy_profiler.py",           "Weights for the priority-ranking algorithm."),
        ("Search Terms",       "policy_profiler.py",           "Custom terms to surface in profiler output."),
        ("Control Extraction", "extract_controls.py",          "Control-ID patterns, whitelist, and blacklist."),
        ("Metadata",           "add_metadata.py",              "Metadata fields, URL lookup table, and tag-generation rules."),
        ("Docx2md",            "docx2md.py",                   "Markdown conversion settings and front-matter field mapping."),
        ("Docx2jsonl",         "docx2jsonl.py",                "JSONL chunking parameters."),
        ("Acronym Finder",     "acronym_finder.py",            "Acronym detection patterns and ignore list."),
    ]
    for name, script, purpose in sheets:
        body_row(name, script, purpose)

    note_row("Most sheets work correctly with defaults. Edit only for custom heading styles, special control-ID formats, or organisation-specific requirements.")
    ws.append([])

    # Running the Pipeline
    section_header("Running the Pipeline")
    ws.append(["Command", "Description", ""])
    r = ws.max_row
    for col in range(1, 3):
        ws.cell(row=r, column=col).font = BOLD_FONT
        ws.cell(row=r, column=col).border = THIN_BORDER

    commands = [
        ("python run_pipeline.py",            "Execute the full pipeline (all enabled steps)."),
        ("python run_pipeline.py --step 0",   "Run a single step by its number."),
        ("python run_pipeline.py --step 1-3", "Run a contiguous range of steps."),
    ]
    for cmd, desc in commands:
        body_row(cmd, desc)

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 50
    ws.freeze_panes = "A2"


def _build_input_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Input")
    inp = cfg.get("input", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Directory Settings")
    _add_setting(ws, "directory", inp.get("directory", "./input"),
                 "Path to folder containing .docx files (absolute or relative) [default: ./input]")
    _add_setting(ws, "pattern", inp.get("pattern", "*.docx"),
                 "File glob pattern — almost always *.docx [default: *.docx]")
    _add_setting(ws, "recursive", inp.get("recursive", False),
                 "Scan sub-folders? TRUE if docs are in sub-folders [default: FALSE]")

    # Bool validation for recursive
    _add_bool_validation(ws, "B", ws.max_row, ws.max_row)

    _add_subheader(ws, "Exclude Patterns")
    ws.append(["## Column A = pattern to exclude (one per row). Files matching ANY pattern are skipped", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT

    # Switch to list layout for exclude patterns
    for pat in inp.get("exclude_patterns", ["~$", "_optimized", "_backup", "_fixed", "template"]):
        desc_map = {
            "~$": "Word temp/lock files",
            "_optimized": "Already-processed files",
            "_backup": "Backup copies",
            "_fixed": "Output from heading fixer",
            "template": "Template files, not actual policies",
        }
        ws.append([pat, desc_map.get(pat, ""), ""])
        ws.cell(row=ws.max_row, column=2).font = DESC_FONT

    _finalize_sheet(ws)


def _build_output_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Output")
    out = cfg.get("output", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Root Output Directory")
    _add_setting(ws, "directory", out.get("directory", "./output"),
                 "All step outputs go under this directory [default: ./output]")

    _add_subheader(ws, "Step 0 — Profiler")
    prof = out.get("profiler", {})
    _add_setting(ws, "profiler.directory", prof.get("directory", "0 - profiler"), "Sub-folder name [default: 0 - profiler]")
    _add_setting(ws, "profiler.inventory_file", prof.get("inventory_file", "document_inventory.xlsx"),
                 "Master spreadsheet [default: document_inventory.xlsx]")
    _add_setting(ws, "profiler.json_file", prof.get("json_file", "document_profiles.json"),
                 "Machine-readable profiles [default: document_profiles.json]")
    _add_setting(ws, "profiler.sections_file", prof.get("sections_file", "section_inventory.csv"),
                 "One row per section per doc [default: section_inventory.csv]")
    _add_setting(ws, "profiler.tables_file", prof.get("tables_file", "table_inventory.csv"),
                 "One row per table per doc [default: table_inventory.csv]")
    _add_setting(ws, "profiler.crossrefs_file", prof.get("crossrefs_file", "crossref_inventory.csv"),
                 "One row per cross-reference [default: crossref_inventory.csv]")

    _add_subheader(ws, "Step 1 — Acronyms")
    acr = out.get("acronyms", {})
    _add_setting(ws, "acronyms.directory", acr.get("directory", "1 - acronyms"), "Sub-folder name [default: 1 - acronyms]")
    _add_setting(ws, "acronyms.output_file", acr.get("output_file", "acronym_audit.xlsx"), "Audit Excel output [default: acronym_audit.xlsx]")

    _add_subheader(ws, "Step 2 — Controls")
    ctrl = out.get("controls", {})
    _add_setting(ws, "controls.directory", ctrl.get("directory", "2 - controls"), "Sub-folder name [default: 2 - controls]")
    _add_setting(ws, "controls.output_file", ctrl.get("output_file", "controls_output.csv"), "CSV output [default: controls_output.csv]")
    _add_setting(ws, "controls.output_file_xlsx", ctrl.get("output_file_xlsx", "controls_output.xlsx"),
                 "Excel output [default: controls_output.xlsx]")
    _add_setting(ws, "controls.checkpoint_file", ctrl.get("checkpoint_file", "checkpoint.json"),
                 "Resume progress tracker [default: checkpoint.json]")
    _add_setting(ws, "controls.error_log", ctrl.get("error_log", "errors.log"), "Error log file [default: errors.log]")

    _add_subheader(ws, "Step 3 — Cross References")
    xref = out.get("cross_references", {})
    _add_setting(ws, "cross_references.directory", xref.get("directory", "3 - cross_references"), "Sub-folder name [default: 3 - cross_references]")
    _add_setting(ws, "cross_references.output_file", xref.get("output_file", "cross_references.csv"), "CSV output [default: cross_references.csv]")

    _add_subheader(ws, "Step 4 — Heading Fixes")
    hfix = out.get("heading_fixes", {})
    _add_setting(ws, "heading_fixes.directory", hfix.get("directory", "4 - heading_fixes"), "Sub-folder name [default: 4 - heading_fixes]")
    _add_setting(ws, "heading_fixes.changes_file", hfix.get("changes_file", "heading_changes.csv"),
                 "Changes log CSV [default: heading_changes.csv]")

    _add_subheader(ws, "Step 5 — Split Documents")
    splt = out.get("split_documents", {})
    _add_setting(ws, "split_documents.directory", splt.get("directory", "5 - split_documents"), "Sub-folder name [default: 5 - split_documents]")
    _add_setting(ws, "split_documents.manifest_file", splt.get("manifest_file", "split_manifest.csv"),
                 "Split manifest CSV [default: split_manifest.csv]")

    _add_subheader(ws, "Step 6 — Metadata")
    meta = out.get("metadata", {})
    _add_setting(ws, "metadata.directory", meta.get("directory", "6 - metadata"), "Sub-folder name [default: 6 - metadata]")
    _add_setting(ws, "metadata.manifest_file", meta.get("manifest_file", "metadata_manifest.csv"),
                 "Metadata manifest CSV [default: metadata_manifest.csv]")

    _add_subheader(ws, "Step 7 — Markdown")
    md = out.get("markdown", {})
    _add_setting(ws, "markdown.directory", md.get("directory", "7 - markdown"), "Sub-folder name [default: 7 - markdown]")

    _add_subheader(ws, "Step 8 — JSONL")
    jl = out.get("jsonl", {})
    _add_setting(ws, "jsonl.directory", jl.get("directory", "8 - jsonl"), "Sub-folder name [default: 8 - jsonl]")

    _add_subheader(ws, "Step 9 — Validation")
    val = out.get("validation", {})
    _add_setting(ws, "validation.directory", val.get("directory", "9 - validation"), "Sub-folder name [default: 9 - validation]")
    _add_setting(ws, "validation.output_file", val.get("output_file", "control_validation.csv"), "CSV output [default: control_validation.csv]")
    _add_setting(ws, "validation.review_file", val.get("review_file", "validation_review.xlsx"),
                 "Human review workbook [default: validation_review.xlsx]")

    _add_subheader(ws, "Consolidated Report")
    consol = out.get("consolidated_report", {})
    _add_setting(ws, "consolidated_report.enabled", consol.get("enabled", True),
                 "Build single .xlsx workbook after all steps [default: TRUE]")
    _add_setting(ws, "consolidated_report.filename_prefix", consol.get("filename_prefix", "DPS_Report"),
                 "Prefix for timestamped filename [default: DPS_Report]")

    # Bool validation for consolidated_report.enabled
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "consolidated_report.enabled":
            _add_bool_validation(ws, "B", r, r)
            break

    _finalize_sheet(ws)


def _build_sections_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Sections")
    sections = cfg.get("sections", {})

    _style_headers(ws, ["Category", "Keyword", "Description"])

    defaults = {
        "purpose": ["purpose", "policy purpose", "1.0 purpose", "document purpose"],
        "scope": ["scope", "policy scope", "2.0 scope", "applicability", "applicability and scope"],
        "intent": ["intent", "policy intent", "3.0 intent", "objective", "policy objective", "security objective"],
        "controls": ["controls", "technical controls", "security controls", "4.0 controls",
                      "control requirements", "requirements", "policy requirements", "implementation requirements"],
        "appendix": ["appendix", "appendices", "5.0 appendix", "supplementary",
                      "reference", "references", "glossary", "definitions"],
    }

    desc = "Keywords the profiler matches against H1 headings to classify each section"
    for category in ["purpose", "scope", "intent", "controls", "appendix"]:
        keywords = sections.get(category, defaults.get(category, []))
        _add_subheader(ws, category.title())
        for kw in keywords:
            ws.append([category, kw, desc if kw == keywords[0] else ""])
            if kw == keywords[0]:
                ws.cell(row=ws.max_row, column=3).font = DESC_FONT
        desc = ""

    _finalize_sheet(ws)


def _build_headings_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Headings")
    hdg = cfg.get("headings", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Built-in Heading Styles")
    ws.append(["## Column A = style name (one per row). Standard Word heading style names", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for style in hdg.get("builtin_styles", ["Heading 1", "Heading 2", "Heading 3", "Heading 4",
                                              "heading 1", "heading 2", "heading 3", "heading 4"]):
        ws.append([style, "", ""])

    _add_subheader(ws, "Custom Heading Styles")
    ws.append(["## Column A = style name (one per row). Your org's custom heading style names", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for style in hdg.get("custom_heading_styles", ["Policy Heading 1", "Policy Heading 2",
                                                     "TOC Heading", "AppendixHeading"]):
        ws.append([style, "", ""])

    _add_subheader(ws, "Custom Style Map")
    ws.append(["## Column A = custom style name, Column B = standard heading it maps to", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    style_map = hdg.get("custom_style_map", {
        "policy heading 1": "Heading 1", "policy heading 2": "Heading 2",
        "policy heading 3": "Heading 3", "policyheading1": "Heading 1",
        "policyheading2": "Heading 2", "policyheading3": "Heading 3",
        "doc heading 1": "Heading 1", "doc heading 2": "Heading 2",
        "doc heading 3": "Heading 3", "heading1": "Heading 1",
        "heading2": "Heading 2", "heading3": "Heading 3",
        "title heading": "Heading 1", "section heading": "Heading 1",
        "subsection heading": "Heading 2",
    })
    for key, val in style_map.items():
        ws.append([key, val, ""])

    _add_subheader(ws, "Fake Heading Detection")
    _add_setting(ws, "fake_heading_min_font_size", hdg.get("fake_heading_min_font_size", 12),
                 "Min font size (pt) for bold text to be a fake heading. Lower=more detected [default: 12]")
    _add_setting(ws, "fake_heading_max_chars", hdg.get("fake_heading_max_chars", 200),
                 "Max char length for profiler fake heading candidates [default: 200]")
    _add_setting(ws, "fake_heading_max_chars_fixer", hdg.get("fake_heading_max_chars_fixer", 120),
                 "Tighter limit for heading fixer (Step 4) conversion [default: 120]")

    _add_subheader(ws, "Heading Level Patterns (Regex)")
    _add_setting(ws, "heading1_pattern", hdg.get("heading1_pattern", r"^(?:\d+\.0\s+|[IVXLC]+\.\s+)"),
                 'Matches H1 numbering, e.g. "1.0 Purpose", "IV. Controls"')
    _add_setting(ws, "heading2_pattern", hdg.get("heading2_pattern", r"^(?:\d+\.\d+\s+|[A-Z]\.\s+)"),
                 'Matches H2 numbering, e.g. "1.1 Access", "A. Overview"')
    _add_setting(ws, "heading3_pattern", hdg.get("heading3_pattern", r"^\d+\.\d+\.\d+\s+"),
                 'Matches H3 numbering, e.g. "1.1.1 Sub-section"')
    _add_setting(ws, "default_heading_level", hdg.get("default_heading_level", 2),
                 "Default level when no numbering pattern matches (1 or 2) [default: 2]")

    _finalize_sheet(ws)


def _build_text_deletions_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Text Deletions")
    td = cfg.get("text_deletions", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Settings")
    _add_setting(ws, "enabled", td.get("enabled", False),
                 "Set TRUE to activate text deletion during Step 4 [default: FALSE]")
    _add_setting(ws, "case_sensitive", td.get("case_sensitive", True),
                 "TRUE for exact case match, FALSE for case-insensitive [default: TRUE]")

    # Bool validations
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("enabled", "case_sensitive"):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Phrases to Delete")
    ws.append(["## Add exact phrases to delete (one per row in column A). Originals are never modified.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for phrase in (td.get("phrases") or []):
        ws.append([phrase, "", ""])

    _add_subheader(ws, "Sections to Delete")
    ws.append(["## Delete an entire section: heading + all content until the next heading of same or higher level.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## Match is case-insensitive substring on the heading text. Originals are never modified.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## Example: 'Appendix A' deletes the Appendix A heading and everything below it until the next H1/H2.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT

    # Column headers for the section deletion sub-table
    section_header_row = ws.max_row + 1
    ws.append(["Section Heading", "Delete (TRUE/FALSE)", "Description"])
    for cell in ws[ws.max_row]:
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    for entry in (td.get("section_deletions") or []):
        heading = entry.get("heading", "")
        delete = entry.get("delete", True)
        desc = entry.get("description", "")
        ws.append([heading, delete, desc])

    # Leave a few empty rows for user additions
    for _ in range(5):
        ws.append(["", "", ""])

    # Bool validation for Delete column (B) in section deletion rows
    data_start = section_header_row + 1
    _add_bool_validation(ws, "B", data_start, ws.max_row)

    _finalize_sheet(ws)


def _build_cross_references_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Cross References")
    xr = cfg.get("cross_references", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Detection Settings")
    _add_setting(ws, "detect_hyperlink_crossrefs", xr.get("detect_hyperlink_crossrefs", True),
                 "Detect hyperlinks whose text looks like a section/policy ref [default: TRUE]")
    _add_setting(ws, "detect_urls", xr.get("detect_urls", True),
                 "Extract URLs from hyperlinks and bare text [default: TRUE]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("detect_hyperlink_crossrefs", "detect_urls"):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Profiler Patterns (Regex — Step 0)")
    ws.append(["## Regex patterns for COUNTING cross-references in profiler", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for pat in xr.get("profiler_patterns", [
        r"see section\s+[\d\.]+[a-z]?", r"refer to\s+(section|the)",
        r"as described in\s+(the|section)", r"per section\s+[\d\.]+",
        r"in accordance with\s+(the|section)", r"as defined in\s+(the|section)",
        r"as outlined in\s+(the|section)", r"per the organization'?s\s+\w+",
        r"see the\s+\w+\s+policy", r"as specified in\s+(the|section)",
    ]):
        ws.append([pat, "", ""])

    _add_subheader(ws, "Extraction Patterns (Step 2)")
    ws.append(["## Phrase | Type (internal/external). Extractor auto-builds regex from phrase.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    default_patterns = [
        ("see", "internal"), ("refer to", "internal"), ("per", "internal"),
        ("as described in", "internal"), ("as defined in", "internal"),
        ("as described in", "external"), ("as defined in", "external"),
        ("refer to", "external"), ("in accordance with", "external"),
    ]
    raw_patterns = xr.get("extraction_patterns") or []
    if raw_patterns and isinstance(raw_patterns[0], dict):
        for entry in raw_patterns:
            ws.append([entry.get("phrase", ""), entry.get("type", ""), ""])
    elif raw_patterns:
        for phrase, ptype in raw_patterns:
            ws.append([phrase, ptype, ""])
    else:
        for phrase, ptype in default_patterns:
            ws.append([phrase, ptype, ""])

    _add_subheader(ws, "Document Name Keywords")
    ws.append(["## Column A = keyword (one per row). External patterns match doc names ending in these", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for kw in xr.get("document_name_keywords", [
        "Policy", "Standard", "Plan", "Procedure", "Guide",
        "Guideline", "Program", "Framework", "Charter",
    ]):
        ws.append([kw, "", ""])

    _finalize_sheet(ws)


def _build_tables_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Tables")
    tbl = cfg.get("tables", {}).get("classification", {})

    _style_headers(ws, ["Table Type", "Keyword", "Min Columns"])

    defaults = {
        "control_matrix": (["control id", "control", "requirement", "framework",
                            "nist", "status", "implementation", "responsible", "owner"], 3),
        "applicability_table": (["applies to", "applicability", "system", "environment",
                                  "in scope", "yes", "no", "n/a"], 2),
        "reference_table": (["term", "definition", "acronym", "glossary",
                              "abbreviation", "description"], 2),
        "crosswalk_table": (["nist", "cis", "cmmc", "iso", "mapping", "alignment", "framework"], 3),
        "role_responsibility": (["role", "responsibility", "responsible", "accountable", "raci"], 2),
    }

    for ttype in ["control_matrix", "applicability_table", "reference_table",
                   "crosswalk_table", "role_responsibility"]:
        _add_subheader(ws, ttype.replace("_", " ").title())
        data = tbl.get(ttype, {})
        keywords = data.get("keywords", defaults[ttype][0])
        min_cols = data.get("min_columns", defaults[ttype][1])
        for kw in keywords:
            ws.append([ttype, kw, min_cols])

    _finalize_sheet(ws)


def _build_classification_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Classification")
    cls = cfg.get("classification", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Type A — Table-Heavy")
    _add_setting(ws, "type_a.min_table_content_pct",
                 cls.get("type_a", {}).get("min_table_content_pct", 40),
                 ">X% of content in tables → Type A [default: 40]")

    _add_subheader(ws, "Type B — Prose-Heavy")
    _add_setting(ws, "type_b.max_table_content_pct",
                 cls.get("type_b", {}).get("max_table_content_pct", 10),
                 "<X% of content in tables → Type B [default: 10]")

    _add_subheader(ws, "Type C — Hybrid (Procedure Keywords)")
    ws.append(["## Column A = keyword (one per row). Keywords that signal embedded procedures", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for kw in cls.get("type_c", {}).get("procedure_keywords", [
        "step 1", "step 2", "procedure", "escalation",
        "workflow", "response process", "playbook",
    ]):
        ws.append([kw, "", ""])

    _add_subheader(ws, "Type D — Appendix-Dominant")
    _add_setting(ws, "type_d.min_appendix_content_pct",
                 cls.get("type_d", {}).get("min_appendix_content_pct", 60),
                 ">X% of content in appendix → Type D [default: 60]")

    _finalize_sheet(ws)


def _build_profiling_flags_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Profiling Flags")
    pf = cfg.get("profiling_flags", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "ControlDense")
    _add_setting(ws, "control_dense.min_controls_per_page",
                 pf.get("control_dense", {}).get("min_controls_per_page", 5.0),
                 "Flag if >= this many control IDs per approx page [default: 5.0]")

    _add_subheader(ws, "HeadingVariance")
    hv = pf.get("heading_variance", {})
    _add_setting(ws, "heading_variance.max_level_skips", hv.get("max_level_skips", 2),
                 "Flag if >= this many heading level jumps (e.g., H1→H3) [default: 2]")
    _add_setting(ws, "heading_variance.max_fake_ratio", hv.get("max_fake_ratio", 0.5),
                 "Flag if >= X% of headings are fake (0.5 = 50%) [default: 0.5]")

    _add_subheader(ws, "TableDense")
    _add_setting(ws, "table_dense.min_table_content_pct",
                 pf.get("table_dense", {}).get("min_table_content_pct", 30),
                 "Flag if >= X% table content (early warning) [default: 30]")

    _finalize_sheet(ws)


def _build_thresholds_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Thresholds")
    th = cfg.get("thresholds", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "RAG Chunk Size")
    _add_setting(ws, "max_characters", th.get("max_characters", 36000),
                 "Max chars per split sub-document. 36000≈20pp, 18000≈10pp, 72000≈40pp [default: 36000]")
    _add_setting(ws, "max_pages", th.get("max_pages", 20),
                 "Approximate page equivalent (reporting only) [default: 20]")

    _add_subheader(ws, "Estimation Ratios")
    _add_setting(ws, "paragraphs_per_page", th.get("paragraphs_per_page", 30),
                 "approx_pages = total_paragraphs / this. Lower=more pages [default: 30]")
    _add_setting(ws, "chars_per_page", th.get("chars_per_page", 1800),
                 "Chars per page for split manifest estimates [default: 1800]")

    _add_subheader(ws, "Flagging Thresholds")
    _add_setting(ws, "section_dominance_pct", th.get("section_dominance_pct", 50),
                 "Flag sections exceeding this % of total doc length [default: 50]")
    _add_setting(ws, "high_table_count", th.get("high_table_count", 10),
                 "Flag docs with more tables than this [default: 10]")

    _finalize_sheet(ws)


def _build_priority_scoring_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Priority Scoring")
    ps = cfg.get("priority_scoring", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Scoring Weights")
    ws.append(["## Higher weight = more impact on priority score. Adjust ratios to change emphasis.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    weights = ps.get("weights", {
        "table_count": 2.0, "cross_ref_count": 1.5, "fake_heading_count": 1.0,
        "page_count": 0.5, "missing_sections": 1.5, "merged_cells": 1.0,
        "over_size_limit": 3.0,
    })
    weight_descs = {
        "table_count": "More tables = harder to optimize [default: 2.0]",
        "cross_ref_count": "More cross-refs = more manual work [default: 1.5]",
        "fake_heading_count": "Fake headings break chunking [default: 1.0]",
        "page_count": "Longer docs need more splitting [default: 0.5]",
        "missing_sections": "Missing standard sections = structural problems [default: 1.5]",
        "merged_cells": "Merged cells complicate table flattening [default: 1.0]",
        "over_size_limit": "Over character limit is a hard retrieval problem [default: 3.0]",
    }
    for key, val in weights.items():
        _add_setting(ws, key, val, weight_descs.get(key, ""))

    _add_subheader(ws, "Usage Frequency")
    ws.append(["## Per-document usage (0-10). Multiplied by 2.0 and added to priority.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## Filename (exact match)", "Frequency (0-10)", ""])
    ws.cell(row=ws.max_row, column=1).font = SUBHEADER_FONT
    ws.cell(row=ws.max_row, column=2).font = SUBHEADER_FONT
    usage = ps.get("usage_frequency", {})
    for fname, freq in usage.items():
        ws.append([fname, freq, ""])
    # Leave a few empty rows for user to fill
    for _ in range(3):
        ws.append(["", "", ""])

    _finalize_sheet(ws)


def _build_search_terms_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Search Terms")
    st = cfg.get("search_terms", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Settings")
    _add_setting(ws, "enabled", st.get("enabled", True),
                 "Enable key term search in profiler [default: TRUE]")
    _add_setting(ws, "match_mode", st.get("match_mode", "word"),
                 '"word" = whole-word boundary, "substring" = anywhere in text [default: word]')
    _add_setting(ws, "show_counts", st.get("show_counts", True),
                 "TRUE = show occurrence count, FALSE = show YES/blank [default: TRUE]")

    # Validations
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("enabled", "show_counts"):
            _add_bool_validation(ws, "B", r, r)
        if val == "match_mode":
            _add_enum_validation(ws, "B", r, ["word", "substring"])

    _add_subheader(ws, "Terms")
    ws.append(["## Each term gets its own column in the Excel output. One per row.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for term in st.get("terms", ["AC", "Cloud", "NIST"]):
        ws.append([term, "", ""])
    # Empty rows for user
    for _ in range(5):
        ws.append(["", "", ""])

    _finalize_sheet(ws)


def _build_control_extraction_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Control Extraction")
    ce = cfg.get("control_extraction", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Common Settings")
    _add_setting(ws, "require_bold_control_id", ce.get("require_bold_control_id", True),
                 "Only extract control IDs that appear in bold text. Set FALSE if your IDs are not bold [default: TRUE]")
    _add_setting(ws, "bold_fallback_if_zero", ce.get("bold_fallback_if_zero", True),
                 "If bold-only extraction finds 0 controls in a doc, auto-retry without the bold requirement [default: TRUE]")
    _add_setting(ws, "tables_ignore_bold", ce.get("tables_ignore_bold", True),
                 "Always extract control IDs from table cells regardless of bold setting [default: TRUE]")
    _add_setting(ws, "control_id_anchor_start", ce.get("control_id_anchor_start", False),
                 "Only match control IDs near the start of a paragraph (filters out inline references) [default: FALSE]")
    _add_setting(ws, "control_id_start_chars", ce.get("control_id_start_chars", 25),
                 "How many chars from paragraph start to search when anchor_start is enabled [default: 25]")
    _add_setting(ws, "min_control_block_lines", ce.get("min_control_block_lines", 0),
                 "Discard control blocks with fewer than N content lines. 0 = keep all. 2 = drop inline refs [default: 0]")
    _add_setting(ws, "enable_checkpoint", ce.get("enable_checkpoint", True),
                 "Save progress so re-runs skip already-processed files. Delete checkpoint.json to force fresh run [default: TRUE]")
    _add_setting(ws, "output_format", ce.get("output_format", "both"),
                 '"csv", "xlsx", or "both" — controls the output file format for extracted controls [default: both]')

    bool_settings = (
        "require_bold_control_id", "bold_fallback_if_zero", "tables_ignore_bold",
        "control_id_anchor_start", "enable_checkpoint",
    )
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in bool_settings:
            _add_bool_validation(ws, "B", r, r)
        if val == "output_format":
            _add_enum_validation(ws, "B", r, ["csv", "xlsx", "both"])

    _add_subheader(ws, "Control ID Patterns (Regex)")
    ws.append(['## Regex patterns to match control IDs in document text. One pattern per row.', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Examples matched: AC-1.001, IR001.002, CFG.1.0042, AUP 1.001', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## If zero controls extracted, your IDs use a different format — add a pattern here.', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Test patterns at regex101.com before adding.', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for pat in ce.get("control_id_patterns", [
        r"\b[A-Z]{2,4}[-.]?\d{1,3}[-.]\d{2,4}\b",
        r"\b[A-Z]{2,4}\s+\d{1,3}\.\d{2,4}\b",
        r"\b[A-Z]{2,4}\d{2,3}\.\d{2,4}\b",
    ]):
        ws.append([pat, "", ""])

    _add_subheader(ws, "Whitelist / Blacklist")
    ws.append(['## One control ID per row. Type = "whitelist" or "blacklist".', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Whitelist: if ANY whitelist entries exist, ONLY matching controls are kept.', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Blacklist: matching controls are excluded (applied after whitelist).', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Supports exact IDs (e.g. AC-1.001) and prefix wildcards (e.g. AC-*).', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(['## Add as many rows as needed. Leave this section empty to keep all controls.', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for item in ce.get("whitelist", []):
        ws.append([f"whitelist:{item}", "", ""])
    for item in ce.get("blacklist", []):
        ws.append([f"blacklist:{item}", "", ""])

    _add_subheader(ws, "Guidance Keywords")
    ws.append(["## Keywords marking the boundary between control description and supplemental guidance.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## Column A = keyword (one per row). When a paragraph contains one of these, text after it is classified as guidance.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for kw in ce.get("guidance_keywords", [
        "implementation guidance", "implementation:", "guidelines:",
        "how to implement", "supplemental guidance",
    ]):
        ws.append([kw, "", ""])

    _add_subheader(ws, "Metadata Triggers")
    ws.append(["## Category | Keyword — triggers for extracting document-level metadata (purpose, scope, applicability).", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## Column A = category name, Column B = keyword. Multiple keywords per category are allowed (one per row).", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    ws.append(["## The extractor scans the first N paragraphs (see metadata_scan_paragraphs) for these keywords.", "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    triggers = ce.get("metadata_triggers", {
        "purpose": ["purpose", "objective", "intent"],
        "scope": ["scope", "coverage", "boundary"],
        "applicability": ["applicability", "applies to", "applies for"],
    })
    for category, keywords in triggers.items():
        for kw in keywords:
            ws.append([category, kw, ""])

    _add_subheader(ws, "Heading Detection (Advanced)")
    hd = ce.get("heading_detection", {})
    _add_setting(ws, "heading_detection.use_word_heading_style",
                 hd.get("use_word_heading_style", True), "Use Word heading styles for detection [default: TRUE]")
    _add_setting(ws, "heading_detection.section_keyword_pattern",
                 hd.get("section_keyword_pattern", r"^[Ss]ection\s+\d{1,2}"),
                 'Regex for section headings like "Section 4"')
    _add_setting(ws, "heading_detection.numbered_title_pattern",
                 hd.get("numbered_title_pattern", r"^\d{1,2}\.?\d{0,2}\s+[A-Z][a-zA-Z\s]{3,50}$"),
                 'Regex for numbered titles like "4.1 Access Control"')
    _add_setting(ws, "heading_detection.detect_allcaps", hd.get("detect_allcaps", True),
                 "Detect ALL CAPS headings [default: TRUE]")
    _add_setting(ws, "heading_detection.allcaps_max_length", hd.get("allcaps_max_length", 80),
                 "Max chars for all-caps heading [default: 80]")
    _add_setting(ws, "heading_detection.allcaps_min_words", hd.get("allcaps_min_words", 3),
                 "Min words for all-caps heading [default: 3]")
    _add_setting(ws, "heading_detection.detect_bold_short", hd.get("detect_bold_short", True),
                 "Detect short bold text as headings [default: TRUE]")
    _add_setting(ws, "heading_detection.bold_max_length", hd.get("bold_max_length", 60),
                 "Max chars for bold heading [default: 60]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and val.startswith("heading_detection.") and val.split(".")[-1] in (
            "use_word_heading_style", "detect_allcaps", "detect_bold_short"
        ):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Implementation Trigger (Regex)")
    _add_setting(ws, "implementation_trigger",
                 ce.get("implementation_trigger",
                        r"(?i)(implementation guidance|implementation:|guidelines:|how to implement|supplemental guidance)"),
                 "Regex for guidance boundary within a control block")

    _add_setting(ws, "metadata_scan_paragraphs", ce.get("metadata_scan_paragraphs", 40),
                 "How many paragraphs from top of doc to scan for metadata [default: 40]")

    _finalize_sheet(ws)


def _build_pipeline_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Pipeline")
    steps = cfg.get("pipeline", {}).get("steps", [])

    _style_headers(ws, ["Step", "Name", "Script", "Enabled", "Description"])

    defaults = [
        (0, "Step 0 - Document Profiler", "policy_profiler.py", True,
         "Scan all docs, extract metadata, classify types, score priority, count words"),
        (1, "Step 1 - Acronym Finder", "acronym_finder.py", True,
         "Scan all docs for acronym candidates and generate audit Excel"),
        (2, "Step 2 - Control Extractor", "extract_controls.py", True,
         "Extract structured control data from compliance docs"),
        (3, "Step 3 - Cross-Reference Extractor", "cross_reference_extractor.py", True,
         "Capture all cross-refs BEFORE any structural changes"),
        (4, "Step 4 - Heading Style Fixer", "heading_style_fixer.py", True,
         "Convert fake bold headings to real Word Heading styles"),
        (5, "Step 5 - Section Splitter", "section_splitter.py", True,
         "Split fixed docs at H1 boundaries into sub-documents"),
        (6, "Step 6 - Metadata Injector", "add_metadata.py", True,
         "Add identity metadata to sub-documents"),
        (7, "Step 7 - DOCX to Markdown", "docx2md.py", True,
         "Convert .docx files to clean Markdown with YAML frontmatter"),
        (8, "Step 8 - DOCX to JSONL", "docx2jsonl.py", True,
         "Convert .docx files to chunked JSONL for RAG/vector DB ingestion"),
        (9, "Step 9 - Control Validator", "validate_controls.py", True,
         "Validate all Step 2 controls are present in Step 5 split documents"),
    ]

    if steps:
        for i, step in enumerate(steps):
            ws.append([i, step.get("name", ""), step.get("script", ""),
                       step.get("enabled", True), step.get("description", "")])
    else:
        for row in defaults:
            ws.append(list(row))

    # Bool validation for Enabled column (D)
    _add_bool_validation(ws, "D", 2, ws.max_row)

    _finalize_sheet(ws)


def _build_metadata_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Metadata")
    meta = cfg.get("metadata", {})

    _style_headers(ws, ["Setting", "Value", "Description", "", ""])

    _add_subheader(ws, "General Settings", ncols=5)
    _add_setting(ws, "placement", meta.get("placement", "top"),
                 '"top", "top_and_bottom", or "each_page" [default: top]')
    _add_setting(ws, "add_separator", meta.get("add_separator", True),
                 "Horizontal rule after metadata block [default: TRUE]")
    _add_setting(ws, "font_size", meta.get("font_size", 8),
                 "Font size (pt) for metadata text [default: 8]")
    _add_setting(ws, "label_color", meta.get("label_color", "2F5496"),
                 "Hex RGB color for metadata labels [default: 2F5496]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == "placement":
            _add_enum_validation(ws, "B", r, ["top", "top_and_bottom", "each_page"])
        if val == "add_separator":
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Metadata Fields", ncols=5)
    col_headers = ["## Key", "Label", "Enabled", "Source", "Value (for static source)"]
    ws.append(col_headers)
    for col_idx in range(1, len(col_headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = DESC_FONT
    fields = meta.get("fields", [
        {"key": "name", "label": "Document", "enabled": True, "source": "auto"},
        {"key": "url", "label": "URL", "enabled": True, "source": "auto"},
        {"key": "scope", "label": "Scope", "enabled": True, "source": "auto"},
        {"key": "intent", "label": "Intent", "enabled": True, "source": "auto"},
        {"key": "tags", "label": "Tags", "enabled": True, "source": "auto"},
        {"key": "acronyms", "label": "Acronyms", "enabled": True, "source": "auto"},
    ])
    for field in fields:
        ws.append([field.get("key", ""), field.get("label", ""),
                   field.get("enabled", True), field.get("source", "auto"),
                   field.get("value", "")])

    _add_subheader(ws, "URL Resolution")
    url = meta.get("url", {})
    _add_setting(ws, "url.lookup_file", url.get("lookup_file", "./input/Doc_URL.xlsx"),
                 "Excel file mapping document names to URLs [default: ./input/Doc_URL.xlsx]")
    _add_setting(ws, "url.name_column", url.get("name_column", "Document_Name"),
                 "Column header for doc names in lookup file [default: Document_Name]")
    _add_setting(ws, "url.url_column", url.get("url_column", "URL"),
                 "Column header for URLs in lookup file [default: URL]")
    _add_setting(ws, "url.sheet", url.get("sheet", 0),
                 "Sheet name (text) or index (0 = first) [default: 0]")
    _add_setting(ws, "url.fallback_template", url.get("fallback_template", ""),
                 "URL template when no Excel match. Use {filename} placeholder [default: empty]")

    _add_subheader(ws, "Advanced Settings")
    _add_setting(ws, "max_scope_chars", meta.get("max_scope_chars", 300),
                 "Max chars for scope text extraction [default: 300]")
    _add_setting(ws, "max_intent_chars", meta.get("max_intent_chars", 300),
                 "Max chars for intent text extraction [default: 300]")

    _add_subheader(ws, "Tag Generation")
    tags = meta.get("tags", {})
    _add_setting(ws, "tags.include_doc_type", tags.get("include_doc_type", True),
                 'Add "Type-B" etc. from profiler [default: TRUE]')
    _add_setting(ws, "tags.include_sections_found", tags.get("include_sections_found", True),
                 'Add "has-scope", "has-controls" etc. [default: TRUE]')
    _add_setting(ws, "tags.acronym_definitions_file",
                 tags.get("acronym_definitions_file", "./input/Acronym_Definitions.xlsx"),
                 "Path to verified Acronym Definitions Excel (preferred). Leave empty to fall back to acronym_audit_file [default: ./input/Acronym_Definitions.xlsx]")
    _add_setting(ws, "tags.acronym_audit_file", tags.get("acronym_audit_file", ""),
                 "Path to raw Acronym Finder output Excel (fallback if no definitions file). Leave empty to skip [default: empty]")
    _add_setting(ws, "tags.max_acronym_tags", tags.get("max_acronym_tags", 15),
                 "Max unique-acronym tags per doc. 0 = unlimited [default: 15]")
    _add_setting(ws, "tags.max_tag_doc_count", tags.get("max_tag_doc_count", 3),
                 "Uniqueness threshold: acronyms in \u2264 N docs become tags [default: 3]")
    _add_setting(ws, "tags.max_acronym_definitions", tags.get("max_acronym_definitions", 0),
                 "Max acronym=definition pairs in Acronyms field. 0 = unlimited [default: 0]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("tags.include_doc_type", "tags.include_sections_found"):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Static Tags")
    ws.append(['## Tags added to ALL documents (one per row, e.g. "InfoSec", "GCC-High")', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    for tag in tags.get("static_tags", []):
        ws.append([tag, "", ""])
    for _ in range(3):
        ws.append(["", "", ""])

    _finalize_sheet(ws)


def _build_docx2md_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Docx2md")
    d2m = cfg.get("docx2md", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Input Mode")
    _add_setting(ws, "pure_conversion", d2m.get("pure_conversion", False),
                 "TRUE = Pure Conversion (read from input/), FALSE = Optimized (read from step output) [default: FALSE]")
    _add_setting(ws, "optimized_source_step", d2m.get("optimized_source_step", "heading_fixes"),
                 'Step output to read when not pure: "heading_fixes", "split_documents", "metadata" [default: heading_fixes]')
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "pure_conversion":
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Output")
    _add_setting(ws, "output_directory", d2m.get("output_directory", "./output/7 - markdown"),
                 "Where converted .md files are written [default: ./output/7 - markdown]")

    _add_subheader(ws, "Image Handling")
    _add_setting(ws, "image_handling", d2m.get("image_handling", "extract"),
                 '"extract", "placeholder", or "skip" [default: extract]')
    _add_setting(ws, "image_subfolder", d2m.get("image_subfolder", "images"),
                 "Subfolder name under each doc's image dir [default: images]")

    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "image_handling":
            _add_enum_validation(ws, "B", r, ["extract", "placeholder", "skip"])

    _add_subheader(ws, "Table Conversion")
    _add_setting(ws, "table_strategy", d2m.get("table_strategy", "auto"),
                 '"auto" (recommended), "markdown", or "html" [default: auto]')
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "table_strategy":
            _add_enum_validation(ws, "B", r, ["auto", "markdown", "html"])

    _add_subheader(ws, "Heading & Text Cleanup")
    _add_setting(ws, "heading_normalization", d2m.get("heading_normalization", True),
                 "Shift heading levels so doc starts at H1 [default: TRUE]")
    _add_setting(ws, "max_heading_level", d2m.get("max_heading_level", 2),
                 "Max heading depth in output (2 = only H1/H2). H3+ collapsed to this level. [default: 2]")
    _add_setting(ws, "max_consecutive_blank_lines", d2m.get("max_consecutive_blank_lines", 2),
                 "Collapse runs of blank lines [default: 2]")
    _add_setting(ws, "clean_smart_quotes", d2m.get("clean_smart_quotes", True),
                 "Convert smart quotes to ASCII [default: TRUE]")
    _add_setting(ws, "strip_zero_width_chars", d2m.get("strip_zero_width_chars", True),
                 "Remove zero-width spaces, BOM, etc. [default: TRUE]")
    _add_setting(ws, "extract_text_boxes", d2m.get("extract_text_boxes", True),
                 "Extract floating text box content inline [default: TRUE]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("heading_normalization", "clean_smart_quotes",
                    "strip_zero_width_chars", "extract_text_boxes"):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Doc URL Resolution")
    _add_setting(ws, "include_doc_url", d2m.get("include_doc_url", True),
                 "Resolve Doc URLs from metadata.url Excel lookup into frontmatter [default: TRUE]")
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "include_doc_url":
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Metadata Placement")
    _add_setting(ws, "metadata_placement", d2m.get("metadata_placement", "top"),
                 '"top" = YAML frontmatter only; "top_and_bottom" = also append readable block at end [default: top]')
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "metadata_placement":
            _add_enum_validation(ws, "B", r, ["top", "top_and_bottom"])

    _add_subheader(ws, "Control ID Headings")
    _add_setting(ws, "promote_control_ids_to_heading", d2m.get("promote_control_ids_to_heading", True),
                 "Promote paragraphs with control IDs to H2 headings (uses control_extraction patterns) [default: TRUE]")
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "promote_control_ids_to_heading":
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Logging")
    _add_setting(ws, "log_file_prefix", d2m.get("log_file_prefix", "docx2md_log"),
                 "Prefix for timestamped log Excel file [default: docx2md_log]")

    _add_subheader(ws, "Metadata Frontmatter")
    _add_setting(ws, "include_metadata_frontmatter", d2m.get("include_metadata_frontmatter", True),
                 "Add YAML frontmatter block at top of each .md file [default: TRUE]")
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "include_metadata_frontmatter":
            _add_bool_validation(ws, "B", r, r)

    # Metadata fields table — columns A-C are parsed by shared_utils; D-F are reference-only.
    _METADATA_FIELD_DEFAULTS = [
        {"name": "title",        "source": "core:title",        "default": "",
         "source_type": "Word core property",    "source_of_truth": "Document Title field in .docx",                         "example_output": '"Access Control Policy"'},
        {"name": "source_file",  "source": "filename",          "default": "",
         "source_type": "Filename",              "source_of_truth": "The .docx filename on disk",                            "example_output": '"Access_Control_Policy_POL-AC-2026-001.docx"'},
        {"name": "modified",     "source": "core:modified",     "default": "",
         "source_type": "Word core property",    "source_of_truth": "Last Modified date in .docx",                           "example_output": '"2026-03-15"'},
        {"name": "doc_id",       "source": r"filename_regex:([A-Z]+-[A-Z]+-\d{4}-\d+)", "default": "",
         "source_type": "Filename regex",        "source_of_truth": "ID pattern extracted from filename",                    "example_output": '"POL-AC-2026-001"'},
        {"name": "converted",    "source": "converted_date",    "default": "",
         "source_type": "Runtime",               "source_of_truth": "Timestamp when conversion runs",                        "example_output": '"2026-03-30T14:22:05"'},
        {"name": "PublishedURL", "source": "doc_url",           "default": "",
         "source_type": "Excel lookup",          "source_of_truth": "input/Doc_URL.xlsx (via metadata.url.* config)",        "example_output": '"https://contoso.sharepoint.com/..."'},
        {"name": "Acronyms",    "source": "excel_lookup_dict:./input/Acronym_Definitions.xlsx:Acronym Definitions:Document:Acronym:Definition", "default": "",
         "source_type": "Excel lookup (dict)",   "source_of_truth": "input/Acronym_Definitions.xlsx > Acronym Definitions",  "example_output": '["AC = Access Control", "MFA = Multi-Factor Auth"]'},
        {"name": "Tags",        "source": "excel_lookup_list:./input/Acronym_Definitions.xlsx:Custom Tags:Document_Name:Tags", "default": "",
         "source_type": "Excel lookup (list)",   "source_of_truth": "input/Acronym_Definitions.xlsx > Custom Tags",          "example_output": '["access control", "CUI", "FedRAMP-High"]'},
    ]

    ws.append(["## Metadata fields: Name | Source | Default", "", "",
               "Source Type", "Source of Truth", "Example Output"])
    r = ws.max_row
    ws.cell(row=r, column=1).font = DESC_FONT
    for col in range(4, 7):
        cell = ws.cell(row=r, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
    for field in d2m.get("metadata_fields", _METADATA_FIELD_DEFAULTS):
        # Match enrichment data from defaults by field name
        enrichment = next((f for f in _METADATA_FIELD_DEFAULTS if f["name"] == field.get("name")), {})
        ws.append([
            field.get("name", ""),
            field.get("source", ""),
            field.get("default", ""),
            enrichment.get("source_type", ""),
            enrichment.get("source_of_truth", ""),
            enrichment.get("example_output", ""),
        ])

    _finalize_sheet(ws)


def _build_docx2jsonl_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Docx2jsonl")
    d2j = cfg.get("docx2jsonl", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Input Mode")
    _add_setting(ws, "pure_conversion", d2j.get("pure_conversion", False),
                 "TRUE = Pure Conversion (read from input/), FALSE = Optimized (read from step output) [default: FALSE]")
    _add_setting(ws, "optimized_source_step", d2j.get("optimized_source_step", "heading_fixes"),
                 'Step output to read when not pure: "heading_fixes", "split_documents", "metadata" [default: heading_fixes]')
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "pure_conversion":
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Output")
    _add_setting(ws, "output_directory", d2j.get("output_directory", "./output/8 - jsonl"),
                 "Where converted .jsonl.txt files are written [default: ./output/8 - jsonl]")

    _add_subheader(ws, "Chunking")
    _add_setting(ws, "max_chunk_chars", d2j.get("max_chunk_chars", 1500),
                 "Maximum characters per chunk [default: 1500]")
    _add_setting(ws, "overlap_words", d2j.get("overlap_words", 30),
                 "Words to overlap between consecutive chunks [default: 30]")
    _add_setting(ws, "min_chunk_chars", d2j.get("min_chunk_chars", 100),
                 "Merge trailing chunks smaller than this into previous [default: 100]")

    _add_subheader(ws, "Acronym Definitions")
    _add_setting(ws, "acronym_definitions_file", d2j.get("acronym_definitions_file", "./input/Acronym_Definitions.xlsx"),
                 "Path to confirmed acronym definitions Excel (Per Document sheet) [default: ./input/Acronym_Definitions.xlsx]")

    _add_subheader(ws, "Tag Mapping")
    _add_setting(ws, "tag_file", d2j.get("tag_file", ""),
                 "Optional Excel file mapping documents to tags (leave empty to skip) [default: empty]")

    _finalize_sheet(ws)


def _build_acronym_finder_sheet(wb, cfg: dict):
    ws = wb.create_sheet("Acronym Finder")
    af = cfg.get("acronym_finder", {})

    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Output")
    _add_setting(ws, "output_file", af.get("output_file", "acronym_audit.xlsx"),
                 "Filename for the audit Excel report [default: acronym_audit.xlsx]")

    _add_subheader(ws, "Search Settings")
    search = af.get("search", {})
    _add_setting(ws, "search.min_length", search.get("min_length", 2),
                 "Minimum acronym length (2 = 'AC') [default: 2]")
    _add_setting(ws, "search.max_length", search.get("max_length", 8),
                 "Maximum acronym length [default: 8]")
    _add_setting(ws, "search.scan_tables", search.get("scan_tables", True),
                 "Include acronyms found inside tables [default: TRUE]")
    _add_setting(ws, "search.scan_headers_footers", search.get("scan_headers_footers", True),
                 "Include acronyms found in headers/footers [default: TRUE]")
    _add_setting(ws, "search.scan_textboxes", search.get("scan_textboxes", True),
                 "Include acronyms found in text boxes [default: TRUE]")
    _add_setting(ws, "search.min_global_occurrences", search.get("min_global_occurrences", 1),
                 "Min occurrences across ALL docs to appear in results [default: 1]")
    _add_setting(ws, "search.min_doc_occurrences", search.get("min_doc_occurrences", 1),
                 "Min occurrences within a single doc to appear [default: 1]")

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val in ("search.scan_tables", "search.scan_headers_footers", "search.scan_textboxes"):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Detection Patterns")
    patterns = af.get("patterns", {})
    _add_setting(ws, "patterns.pure_caps", patterns.get("pure_caps", True), "ABC, NIST, GCC [default: TRUE]")
    _add_setting(ws, "patterns.caps_with_numbers", patterns.get("caps_with_numbers", True), "AC-2, 800-53 [default: TRUE]")
    _add_setting(ws, "patterns.caps_with_hyphens", patterns.get("caps_with_hyphens", True), "FedRAMP, ATO-P [default: TRUE]")
    _add_setting(ws, "patterns.caps_with_slashes", patterns.get("caps_with_slashes", True), "IT/OT, CI/CD [default: TRUE]")
    _add_setting(ws, "patterns.parenthetical_defs", patterns.get("parenthetical_defs", True),
                 '"Multi-Factor Authentication (MFA)" [default: TRUE]')

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and val.startswith("patterns."):
            _add_bool_validation(ws, "B", r, r)

    _add_subheader(ws, "Ignore List (one per row)")
    ws.append(['## Acronyms and uppercase words to skip entirely', "", ""])
    ws.cell(row=ws.max_row, column=1).font = DESC_FONT
    default_ignore = [
        "THE", "AND", "FOR", "NOT", "BUT", "ALL", "ARE", "CAN", "HAS",
        "HER", "WAS", "ONE", "OUR", "OUT", "YOU", "USE", "MAY", "SHALL",
        "MUST", "WILL", "WITH", "THIS", "THAT", "FROM", "THEY", "BEEN",
        "HAVE", "EACH", "MAKE", "WHEN", "DOES", "INTO", "THEM", "THEN",
        "THAN", "ONLY", "OVER", "SUCH", "ALSO", "SOME", "THESE", "OTHER",
        "WHICH", "THEIR", "THERE", "WOULD", "ABOUT", "AFTER", "COULD",
        "WHERE", "SHOULD", "THOSE",
        "II", "III", "IV", "VI", "VII", "VIII", "IX", "XI", "XII",
        "PAGE", "DATE", "DRAFT", "FINAL", "NOTE", "TABLE", "FIGURE",
        "REV", "VER", "NA", "TBD", "TODO", "YES", "NO",
    ]
    for item in af.get("ignore_list", default_ignore):
        ws.append([item, "", ""])
    # Add some empty rows for user additions
    for _ in range(5):
        ws.append(["", "", ""])

    _finalize_sheet(ws)


# ── Main ────────────────────────────────────────────────────────────────────

def generate_config_workbook(cfg: dict, output_path: str):
    """Build and save the complete dps_config.xlsx workbook."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _build_readme(wb)
    _build_input_sheet(wb, cfg)
    _build_output_sheet(wb, cfg)
    _build_sections_sheet(wb, cfg)
    _build_headings_sheet(wb, cfg)
    _build_text_deletions_sheet(wb, cfg)
    _build_cross_references_sheet(wb, cfg)
    _build_tables_sheet(wb, cfg)
    _build_classification_sheet(wb, cfg)
    _build_profiling_flags_sheet(wb, cfg)
    _build_thresholds_sheet(wb, cfg)
    _build_priority_scoring_sheet(wb, cfg)
    _build_search_terms_sheet(wb, cfg)
    _build_control_extraction_sheet(wb, cfg)
    _build_pipeline_sheet(wb, cfg)
    _build_metadata_sheet(wb, cfg)
    _build_docx2md_sheet(wb, cfg)
    _build_docx2jsonl_sheet(wb, cfg)
    _build_acronym_finder_sheet(wb, cfg)

    wb.save(output_path)
    print(f"  Config template saved to: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")


def main():
    parser = argparse.ArgumentParser(description="Generate dps_config.xlsx template")
    parser.add_argument("--from-yaml", "-y", default=None,
                        help="Populate values from existing YAML config (e.g. dps_config_fallback.yaml)")
    parser.add_argument("-o", "--output", default="dps_config.xlsx",
                        help="Output path (default: dps_config.xlsx)")
    args = parser.parse_args()

    cfg = {}
    if args.from_yaml:
        import yaml
        if not os.path.isfile(args.from_yaml):
            print(f"ERROR: YAML file not found: {args.from_yaml}")
            sys.exit(1)
        with open(args.from_yaml, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
        print(f"  Loaded values from: {args.from_yaml}")

    generate_config_workbook(cfg, args.output)


if __name__ == "__main__":
    main()
