#!/usr/bin/env python3
"""Merge script: Combine batch result JSONs into final gap analysis Excel.

Handles both standard mode (one result per source control) and chunked mode
(multiple results per source control from different target family batches).

Usage:
    python scripts/merge.py
    python scripts/merge.py --config path/to/config.yaml
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

CONF_FILLS = {
    "high": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

CONF_RANK = {"high": 0, "medium": 1, "low": 2}


# ---------------------------------------------------------------------------
# Config and data loading
# ---------------------------------------------------------------------------

def load_config(config_path):
    with open(config_path) as f:
        return yaml.safe_load(f)


def extract_controls(filepath, sheet_name, col_cfg, header_row):
    """Extract controls from Excel (same logic as setup.py)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    controls = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ctrl_id = row[col_cfg["id"]] if len(row) > col_cfg["id"] else None
        if not ctrl_id:
            continue

        ctrl = {
            "id": str(ctrl_id).strip(),
            "description": str(row[col_cfg["description"]] or "").strip()
            if len(row) > col_cfg["description"] else "",
        }

        for sup in col_cfg.get("supplemental", []):
            idx = sup["index"]
            label = sup["label"]
            ctrl[label] = str(row[idx] or "").strip() if len(row) > idx else ""

        controls.append(ctrl)

    wb.close()
    return controls


def load_results(results_dir):
    """Load and merge all result JSON files from the results directory.

    In chunked mode, the same source control ID appears in multiple files
    (one per target chunk). This function merges matches across all files,
    deduplicates by target_id (keeping highest confidence), and determines
    unique_to_source only after all matches are combined.
    """
    # First pass: collect all raw entries per control ID
    raw_entries = {}  # ctrl_id -> list of result dicts from different files
    files_loaded = 0

    for json_file in sorted(results_dir.glob("*.json")):
        with open(json_file) as f:
            raw_text = f.read()

        # Strip markdown code fences that Claude sometimes wraps around JSON.
        # Handles: bare fences, fences with preamble text, and clean JSON.
        fence_match = re.search(r'```(?:json)?\s*\n(.*?)\n```', raw_text, re.DOTALL)
        if fence_match:
            stripped = fence_match.group(1)
        else:
            stripped = raw_text.strip()

        try:
            data = json.loads(stripped)
        except json.JSONDecodeError as e:
            print(f"  WARNING: Failed to parse {json_file.name}: {e}")
            continue

        files_loaded += 1
        for ctrl_id, result in data.items():
            normalized = normalize_result(result)
            raw_entries.setdefault(ctrl_id, []).append({
                "result": normalized,
                "source_file": json_file.name,
            })

    # Second pass: merge matches across chunks for each control
    merged = {}
    multi_chunk_ids = []

    for ctrl_id, entries in raw_entries.items():
        if len(entries) == 1:
            # Single result — use as-is
            merged[ctrl_id] = entries[0]["result"]
        else:
            # Multiple results (chunked mode) — merge matches
            merged[ctrl_id] = merge_chunked_results(ctrl_id, entries)
            multi_chunk_ids.append(ctrl_id)

    return merged, files_loaded, multi_chunk_ids


def merge_chunked_results(ctrl_id, entries):
    """Merge results for one source control from multiple target chunk batches.

    - Combines all matches, deduplicates by target_id (keeps highest confidence)
    - Determines unique_to_source based on combined matches
    - Combines gap rationales if any
    """
    all_matches = []
    gap_rationales = []

    for entry in entries:
        result = entry["result"]
        all_matches.extend(result.get("matches", []))
        gap = result.get("gap_rationale", "")
        if gap:
            gap_rationales.append(gap)

    # Deduplicate matches by target_id, keeping highest confidence
    best_by_target = {}
    for match in all_matches:
        tid = match.get("target_id", "UNKNOWN")
        conf = match.get("confidence", "low")
        existing = best_by_target.get(tid)

        if existing is None or CONF_RANK.get(conf, 3) < CONF_RANK.get(existing["confidence"], 3):
            best_by_target[tid] = match

    # Sort matches: high first, then medium, then low
    deduped = sorted(best_by_target.values(),
                     key=lambda m: CONF_RANK.get(m.get("confidence", "low"), 3))

    # Determine unique_to_source based on combined matches
    has_coverage = any(
        m.get("confidence") in ("high", "medium") for m in deduped
    )

    return {
        "matches": deduped,
        "unique_to_source": not has_coverage,
        "gap_rationale": " | ".join(gap_rationales) if not has_coverage and gap_rationales else "",
    }


def normalize_result(result):
    """Normalize field names across different result formats."""
    normalized = {
        "matches": result.get("matches", []),
        "unique_to_source": result.get(
            "unique_to_source",
            result.get("unique_to_sscf",
            result.get("unique", True))
        ),
        "gap_rationale": result.get("gap_rationale", ""),
    }

    # Normalize match field names (target_id vs ccm_id vs matched_id)
    for match in normalized["matches"]:
        if "target_id" not in match:
            match["target_id"] = match.get("ccm_id", match.get("matched_id", "UNKNOWN"))

    return normalized


# ---------------------------------------------------------------------------
# Gap analysis
# ---------------------------------------------------------------------------

def compute_gaps(source_controls, target_controls, results, coverage_levels):
    """Compute unique controls per framework."""
    # Source unique: source controls with no high/medium target matches
    source_unique = []
    for ctrl in source_controls:
        result = results.get(ctrl["id"], {})
        if result.get("unique_to_source", True):
            source_unique.append({
                **ctrl,
                "gap_rationale": result.get("gap_rationale", "No mapping data"),
            })

    # Target unique: target controls not matched by any high/medium source match
    matched_target_ids = set()
    for result in results.values():
        for match in result.get("matches", []):
            if match.get("confidence") in coverage_levels:
                matched_target_ids.add(match["target_id"])

    target_unique = [c for c in target_controls if c["id"] not in matched_target_ids]

    return source_unique, target_unique, matched_target_ids


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"


def write_source_unique_sheet(wb, source_unique, source_name, supplemental_labels):
    ws = wb.create_sheet(f"{source_name} Unique")
    headers = [f"{source_name} Control ID", "Description"]
    for label in supplemental_labels:
        if label not in ("Domain", "Title"):
            continue
        headers.append(label)
    headers.append("Gap Rationale")
    ws.append(headers)
    style_header(ws, len(headers))

    for ctrl in source_unique:
        row = [ctrl["id"], ctrl["description"]]
        for label in supplemental_labels:
            if label not in ("Domain", "Title"):
                continue
            row.append(ctrl.get(label, ""))
        row.append(ctrl.get("gap_rationale", ""))
        ws.append(row)

    # Auto-size columns
    widths = [16, 60] + [30] * (len(headers) - 3) + [50]
    for i, w in enumerate(widths[:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT


def write_target_unique_sheet(wb, target_unique, target_name, supplemental_labels):
    ws = wb.create_sheet(f"{target_name} Unique")
    headers = [f"{target_name} Control ID", "Description"]
    for label in supplemental_labels:
        if label not in ("Domain", "Title"):
            continue
        headers.append(label)
    ws.append(headers)
    style_header(ws, len(headers))

    for ctrl in target_unique:
        row = [ctrl["id"], ctrl["description"]]
        for label in supplemental_labels:
            if label not in ("Domain", "Title"):
                continue
            row.append(ctrl.get(label, ""))
        ws.append(row)

    widths = [16, 60] + [30] * (len(headers) - 2)
    for i, w in enumerate(widths[:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT


def write_full_mapping_sheet(wb, source_controls, results, source_name, target_name,
                             coverage_levels):
    ws = wb.create_sheet("Full Mapping")
    headers = [
        f"{source_name} ID", "Title", "Domain", "Description",
        f"Matched {target_name} IDs", "Confidence", "Rationale",
        "Match Count", "Unique",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for ctrl in source_controls:
        result = results.get(ctrl["id"], {})
        matches = result.get("matches", [])
        covered = [m for m in matches if m.get("confidence") in coverage_levels]

        matched_ids = "\n".join(m["target_id"] for m in covered) if covered else "—"
        confidences = "\n".join(m["confidence"] for m in covered) if covered else "—"
        rationales = (
            "\n".join(m.get("rationale", "") for m in covered)
            if covered
            else result.get("gap_rationale", "—")
        )
        match_count = len(covered)
        is_unique = "Yes" if result.get("unique_to_source", True) else "No"

        row_num = ws.max_row + 1
        ws.append([
            ctrl["id"],
            ctrl.get("Title", ""),
            ctrl.get("Domain", ""),
            ctrl["description"],
            matched_ids,
            confidences,
            rationales,
            match_count,
            is_unique,
        ])

        # Color confidence column
        if covered:
            top_conf = covered[0]["confidence"]
            ws.cell(row=row_num, column=6).fill = CONF_FILLS.get(top_conf, PatternFill())

    widths = [16, 28, 35, 50, 20, 12, 50, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Merge batch results into gap analysis Excel")
    parser.add_argument("--config", default=str(PROJECT_DIR / "config.yaml"),
                        help="Path to config.yaml")
    args = parser.parse_args()

    config = load_config(args.config)

    # --- Load results ---
    print("=" * 60)
    print("Step 1: Loading batch results")
    print("=" * 60)

    results_dir = PROJECT_DIR / "results"

    # --- Stale result detection ---
    batch_dir = PROJECT_DIR / "batches"
    expected_results = set()
    if batch_dir.exists():
        for md_file in batch_dir.glob("*.md"):
            expected_results.add(f"{md_file.stem}_result.json")

    actual_results = {f.name for f in results_dir.glob("*.json")}
    stale = actual_results - expected_results
    if stale:
        print(f"  WARNING: {len(stale)} result file(s) have no matching batch — possible stale data:")
        for s in sorted(stale):
            print(f"    - {s}")
        print(f"  Consider deleting stale files from results/ before merging.")

    results, files_loaded, multi_chunk_ids = load_results(results_dir)

    if not results:
        print("  ERROR: No result files found in results/")
        print("  Process batch files first, then save JSON output to results/")
        sys.exit(1)

    print(f"  Loaded {len(results)} control mappings from {files_loaded} files")

    if multi_chunk_ids:
        print(f"  Merged multi-chunk results for {len(multi_chunk_ids)} controls "
              f"(target chunking detected)")

    # --- Re-extract controls for metadata ---
    print(f"\n{'=' * 60}")
    print("Step 2: Loading control metadata from Excel")
    print("=" * 60)

    input_path = PROJECT_DIR / config["input_file"]
    if not input_path.exists():
        print(f"  ERROR: Input file not found: {input_path}")
        sys.exit(1)

    source_cfg = config["source"]
    target_cfg = config["target"]

    source_controls = extract_controls(
        input_path, source_cfg["sheet_name"],
        source_cfg["columns"], source_cfg.get("header_row", 1),
    )
    target_controls = extract_controls(
        input_path, target_cfg["sheet_name"],
        target_cfg["columns"], target_cfg.get("header_row", 1),
    )
    print(f"  {source_cfg['name']}: {len(source_controls)} controls")
    print(f"  {target_cfg['name']}: {len(target_controls)} controls")

    # --- Validate completeness ---
    expected_ids = {c["id"] for c in source_controls}
    mapped_ids = set(results.keys())
    missing = expected_ids - mapped_ids
    extra = mapped_ids - expected_ids

    if missing:
        print(f"\n  WARNING: {len(missing)} source controls not yet mapped:")
        for mid in sorted(missing):
            print(f"    - {mid}")

    if extra:
        print(f"\n  NOTE: {len(extra)} result IDs not in source controls:")
        for eid in sorted(extra):
            print(f"    - {eid}")

    # --- Compute gaps ---
    print(f"\n{'=' * 60}")
    print("Step 3: Computing gap analysis")
    print("=" * 60)

    coverage_levels = config["confidence"]["coverage_levels"]
    source_unique, target_unique, matched_target_ids = compute_gaps(
        source_controls, target_controls, results, coverage_levels,
    )

    print(f"  {source_cfg['name']} unique (not in {target_cfg['name']}):  {len(source_unique)}")
    print(f"  {target_cfg['name']} covered (by {source_cfg['name']}):     {len(matched_target_ids)}")
    print(f"  {target_cfg['name']} unique (not in {source_cfg['name']}):  {len(target_unique)}")

    # --- Write Excel ---
    print(f"\n{'=' * 60}")
    print("Step 4: Writing output Excel")
    print("=" * 60)

    output_dir = PROJECT_DIR / "output"
    output_dir.mkdir(exist_ok=True)

    project_name = config["project_name"].replace(" ", "_")
    output_file = output_dir / f"{project_name}_GapAnalysis.xlsx"

    source_sup_labels = [s["label"] for s in source_cfg["columns"].get("supplemental", [])]
    target_sup_labels = [s["label"] for s in target_cfg["columns"].get("supplemental", [])]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_source_unique_sheet(wb, source_unique, source_cfg["name"], source_sup_labels)
    write_target_unique_sheet(wb, target_unique, target_cfg["name"], target_sup_labels)
    write_full_mapping_sheet(wb, source_controls, results,
                            source_cfg["name"], target_cfg["name"], coverage_levels)

    wb.save(output_file)
    print(f"  Saved: {output_file}")

    # --- Save combined cache ---
    cache_file = output_dir / "mapping_cache.json"
    with open(cache_file, "w") as f:
        json.dump(results, f, indent=2)
    print(f"  Saved: {cache_file}")

    # --- Summary ---
    total_source = len(source_controls)
    mapped_count = len(mapped_ids & expected_ids)
    print(f"\n{'=' * 60}")
    print("Merge Complete")
    print("=" * 60)
    print(f"  Mapped: {mapped_count}/{total_source} source controls")
    if missing:
        print(f"  Missing: {len(missing)} (process remaining batches and re-run)")
    if multi_chunk_ids:
        # Show merge stats
        total_matches = sum(len(results[cid].get("matches", [])) for cid in multi_chunk_ids)
        print(f"  Multi-chunk merge: {len(multi_chunk_ids)} controls, "
              f"{total_matches} total matches (deduplicated)")
    print(f"  Output: {output_file}")


if __name__ == "__main__":
    main()
