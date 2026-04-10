"""
Generate xlsx2docx_config.xlsx — a config workbook template for the VBA macro.

The workbook contains a single "Config" sheet with Setting/Value columns.
Copy this sheet into any workbook before running the GenerateDocx macro,
or edit the values directly in this template and keep it alongside your data.

Usage:
    python create_config_workbook.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


def add_demo_controls_sheet(wb: Workbook):
    """Add a 'Controls' sheet with sample NIST SP 800-53 controls for demo/testing."""
    ws = wb.create_sheet("Controls")

    # Styles
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    body_font = Font(name="Calibri", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    band_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Column headers
    headers = [
        "Control ID",
        "Control Name",
        "Family",
        "Description",
        "Supplemental Guidance",
        "Priority",
        "Baseline (Low)",
        "Baseline (Moderate)",
        "Baseline (High)",
        "Status",
    ]
    col_widths = [14, 28, 22, 55, 55, 10, 14, 14, 14, 14]

    for c, (name, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    # Sample controls — representative subset across families
    # Descriptions are paraphrased summaries (not verbatim NIST text)
    controls = [
        {
            "id": "AC-1",
            "name": "Policy and Procedures",
            "family": "Access Control",
            "desc": (
                "Develop, document, and disseminate an access control policy "
                "that addresses purpose, scope, roles, responsibilities, management "
                "commitment, coordination among organizational entities, and compliance.\n"
                "Develop procedures to facilitate the implementation of the access "
                "control policy and the associated access controls."
            ),
            "guidance": (
                "Access control policy can be included as part of the general "
                "security policy. Access control procedures can be developed for "
                "the organization and for each system as needed.\n"
                "Review and update policy and procedures at an organization-defined frequency."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "AC-2",
            "name": "Account Management",
            "family": "Access Control",
            "desc": (
                "Define and document the types of accounts allowed and specifically "
                "prohibited for use within the system. Assign account managers. "
                "Require conditions for group and role membership.\n"
                "Authorize access based on a valid authorization, intended system "
                "usage, and other organization-defined attributes."
            ),
            "guidance": (
                "Account management includes the creation, activation, modification, "
                "disabling, and removal of accounts. Organizations include explicit "
                "authorization and conditions for the use of guest and temporary accounts.\n"
                "Conditions for account creation include constraints on group and role "
                "membership as well as identifying authorized users of the system."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Partially Implemented",
        },
        {
            "id": "AC-3",
            "name": "Access Enforcement",
            "family": "Access Control",
            "desc": (
                "Enforce approved authorizations for logical access to information "
                "and system resources in accordance with applicable access control policies."
            ),
            "guidance": (
                "Access control policies control access between active entities "
                "(users or processes acting on behalf of users) and passive entities "
                "(devices, files, records, domains) in organizational systems.\n"
                "Access enforcement mechanisms can be employed at the application and "
                "service level to provide increased security for the organization."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "AC-6",
            "name": "Least Privilege",
            "family": "Access Control",
            "desc": (
                "Employ the principle of least privilege, allowing only authorized "
                "accesses for users and processes which are necessary to accomplish "
                "assigned organizational tasks."
            ),
            "guidance": (
                "Organizations employ least privilege for specific duties and systems. "
                "The principle of least privilege is also applied to system processes, "
                "ensuring that the processes have access to systems and operate at "
                "privilege levels no higher than necessary to accomplish organizational missions."
            ),
            "priority": "P1",
            "low": "", "mod": "Selected", "high": "Selected",
            "status": "Planned",
        },
        {
            "id": "AT-1",
            "name": "Policy and Procedures",
            "family": "Awareness and Training",
            "desc": (
                "Develop, document, and disseminate a security and privacy awareness "
                "and training policy that addresses purpose, scope, roles, "
                "responsibilities, management commitment, and compliance.\n"
                "Develop procedures to facilitate the implementation of the "
                "awareness and training policy."
            ),
            "guidance": (
                "Awareness and training policy and procedures address the controls "
                "in the AT family implemented within systems and organizations. "
                "The risk management strategy guides and informs the development of "
                "the policy and procedures."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "AT-2",
            "name": "Literacy Training and Awareness",
            "family": "Awareness and Training",
            "desc": (
                "Provide security and privacy literacy training to system users "
                "(including managers, senior executives, and contractors) as part "
                "of initial training for new users and at an organization-defined "
                "frequency thereafter.\n"
                "Training includes awareness of security risks associated with "
                "user activities and applicable policies, standards, and procedures."
            ),
            "guidance": (
                "Organizations determine the content and frequency of training "
                "based on the specific organizational requirements and the systems "
                "to which personnel have authorized access. Training can be role-based "
                "and include foundational concepts for new employees, periodic refreshers, "
                "and updates when significant changes occur."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "AU-1",
            "name": "Policy and Procedures",
            "family": "Audit and Accountability",
            "desc": (
                "Develop, document, and disseminate an audit and accountability "
                "policy that addresses purpose, scope, roles, responsibilities, "
                "management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the audit "
                "and accountability policy and the associated audit controls."
            ),
            "guidance": (
                "Audit and accountability policy and procedures address the controls "
                "in the AU family. The risk management strategy is a key factor in "
                "establishing policy and procedures."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "AU-2",
            "name": "Event Logging",
            "family": "Audit and Accountability",
            "desc": (
                "Identify the types of events that the system is capable of logging "
                "in support of the audit function. Coordinate the event logging "
                "function with other entities requiring audit-related information.\n"
                "Event types include password changes, failed logons, failed accesses "
                "to security objects, administrative privilege usage, and third-party "
                "credential usage."
            ),
            "guidance": (
                "An event is any observable occurrence in a system. The specific "
                "events that an organization audits are typically a subset of all "
                "possible events. Organizations should consider the auditing of events "
                "identified by type, location, or subject."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Partially Implemented",
        },
        {
            "id": "AU-6",
            "name": "Audit Record Review, Analysis, and Reporting",
            "family": "Audit and Accountability",
            "desc": (
                "Review and analyze system audit records at an organization-defined "
                "frequency for indications of inappropriate or unusual activity. "
                "Report findings to designated organizational officials.\n"
                "Adjust the level of audit record review, analysis, and reporting "
                "when there is a change in risk based on law enforcement information, "
                "intelligence information, or other credible sources."
            ),
            "guidance": (
                "Audit record review, analysis, and reporting covers information "
                "security and privacy-related logging performed by organizations. "
                "Findings can be reported to organizational elements that have a need "
                "to know and can include indicators of compromise or concern."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Planned",
        },
        {
            "id": "CA-1",
            "name": "Policy and Procedures",
            "family": "Assessment, Authorization, and Monitoring",
            "desc": (
                "Develop, document, and disseminate an assessment, authorization, "
                "and monitoring policy that addresses purpose, scope, roles, "
                "responsibilities, management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the policy."
            ),
            "guidance": (
                "Assessment, authorization, and monitoring policy addresses the "
                "controls in the CA family. The policy can be included as part of "
                "the general security and privacy policy. Procedures can be developed "
                "for the organization and for each system."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "CM-1",
            "name": "Policy and Procedures",
            "family": "Configuration Management",
            "desc": (
                "Develop, document, and disseminate a configuration management "
                "policy that addresses purpose, scope, roles, responsibilities, "
                "management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the "
                "configuration management policy and the associated controls."
            ),
            "guidance": (
                "Configuration management policy and procedures address the "
                "controls in the CM family. The risk management strategy informs "
                "the development of the policy and procedures."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "CM-6",
            "name": "Configuration Settings",
            "family": "Configuration Management",
            "desc": (
                "Establish and document configuration settings for components "
                "employed within the system that reflect the most restrictive mode "
                "consistent with operational requirements.\n"
                "Implement the configuration settings and identify, document, and "
                "approve any deviations from established settings."
            ),
            "guidance": (
                "Configuration settings are the parameters that can be changed "
                "in hardware, software, or firmware components that affect the "
                "security and privacy posture of a system. Security-related "
                "parameters are those parameters impacting the security state of "
                "systems including the parameters required to satisfy other controls.\n"
                "Organizations establish organization-wide configuration settings "
                "and subsequently derive specific configuration settings for systems."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Partially Implemented",
        },
        {
            "id": "IA-1",
            "name": "Policy and Procedures",
            "family": "Identification and Authentication",
            "desc": (
                "Develop, document, and disseminate an identification and "
                "authentication policy that addresses purpose, scope, roles, "
                "responsibilities, management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the "
                "identification and authentication policy and the associated controls."
            ),
            "guidance": (
                "Identification and authentication policy and procedures address "
                "the controls in the IA family. The risk management strategy is "
                "a key factor in the development of the policy and procedures."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "IA-2",
            "name": "Identification and Authentication (Organizational Users)",
            "family": "Identification and Authentication",
            "desc": (
                "Uniquely identify and authenticate organizational users, and "
                "associate that unique identification with processes acting on "
                "behalf of those users."
            ),
            "guidance": (
                "Organizations can satisfy the identification and authentication "
                "requirements by complying with the requirements in HSPD-12. "
                "Organizational users include employees, contractors, or individuals "
                "deemed to have equivalent status.\n"
                "Unique identification of individuals in group accounts may be needed "
                "for detailed accountability of individual activity."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "IR-1",
            "name": "Policy and Procedures",
            "family": "Incident Response",
            "desc": (
                "Develop, document, and disseminate an incident response policy "
                "that addresses purpose, scope, roles, responsibilities, management "
                "commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the incident "
                "response policy and the associated controls."
            ),
            "guidance": (
                "Incident response policy and procedures address the controls in "
                "the IR family. Policy can be included as part of the general "
                "security and privacy policy."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "IR-4",
            "name": "Incident Handling",
            "family": "Incident Response",
            "desc": (
                "Implement an incident handling capability for incidents that "
                "includes preparation, detection and analysis, containment, "
                "eradication, and recovery.\n"
                "Coordinate incident handling activities with contingency "
                "planning activities and incorporate lessons learned from "
                "ongoing incident handling into procedures and training."
            ),
            "guidance": (
                "Organizations recognize that incident response capabilities "
                "are dependent on the capabilities of systems and the mission "
                "and business processes being supported. Incident-related "
                "information can be obtained from a variety of sources including "
                "audit monitoring, physical access monitoring, and network monitoring."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Partially Implemented",
        },
        {
            "id": "RA-1",
            "name": "Policy and Procedures",
            "family": "Risk Assessment",
            "desc": (
                "Develop, document, and disseminate a risk assessment policy "
                "that addresses purpose, scope, roles, responsibilities, "
                "management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the risk "
                "assessment policy and the associated controls."
            ),
            "guidance": (
                "Risk assessment policy and procedures address the controls in "
                "the RA family. The risk management strategy informs the "
                "development of the risk assessment policy."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "RA-5",
            "name": "Vulnerability Monitoring and Scanning",
            "family": "Risk Assessment",
            "desc": (
                "Monitor and scan for vulnerabilities in the system and hosted "
                "applications at an organization-defined frequency and when new "
                "vulnerabilities potentially affecting the system are identified.\n"
                "Employ vulnerability monitoring tools and techniques that "
                "facilitate interoperability among tools. Analyze vulnerability "
                "scan reports and results from vulnerability monitoring."
            ),
            "guidance": (
                "Organizations determine the required vulnerability scanning "
                "for all system components, including potential vulnerabilities "
                "in scanning activities themselves. Vulnerability analysis includes "
                "consideration of trend data from security advisories, vulnerability "
                "databases, and other sources of vulnerability information."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Planned",
        },
        {
            "id": "SC-1",
            "name": "Policy and Procedures",
            "family": "System and Communications Protection",
            "desc": (
                "Develop, document, and disseminate a system and communications "
                "protection policy that addresses purpose, scope, roles, "
                "responsibilities, management commitment, coordination, and compliance.\n"
                "Develop procedures to facilitate the implementation of the system "
                "and communications protection policy and the associated controls."
            ),
            "guidance": (
                "System and communications protection policy and procedures "
                "address the controls in the SC family. Policy can be included "
                "as part of the general security and privacy policy."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Implemented",
        },
        {
            "id": "SC-7",
            "name": "Boundary Protection",
            "family": "System and Communications Protection",
            "desc": (
                "Monitor and control communications at the external managed "
                "interfaces to the system and at key internal managed interfaces.\n"
                "Implement subnetworks for publicly accessible system components "
                "that are physically or logically separated from internal networks.\n"
                "Connect to external networks or systems only through managed "
                "interfaces consisting of boundary protection devices."
            ),
            "guidance": (
                "Managed interfaces include gateways, routers, firewalls, guards, "
                "network-based malicious code analysis, virtualization systems, "
                "or encrypted tunnels implemented within a security architecture.\n"
                "Restricting external web traffic only to organizational web servers "
                "within managed interfaces and prohibiting external traffic that "
                "appears to be spoofing internal addresses are examples of restricting "
                "and prohibiting communications."
            ),
            "priority": "P1",
            "low": "Selected", "mod": "Selected", "high": "Selected",
            "status": "Partially Implemented",
        },
    ]

    # Write data rows
    for i, ctrl in enumerate(controls, start=2):
        ws.cell(row=i, column=1, value=ctrl["id"]).font = body_font
        ws.cell(row=i, column=2, value=ctrl["name"]).font = body_font
        ws.cell(row=i, column=3, value=ctrl["family"]).font = body_font
        ws.cell(row=i, column=4, value=ctrl["desc"]).font = body_font
        ws.cell(row=i, column=5, value=ctrl["guidance"]).font = body_font
        ws.cell(row=i, column=6, value=ctrl["priority"]).font = body_font
        ws.cell(row=i, column=7, value=ctrl["low"]).font = body_font
        ws.cell(row=i, column=8, value=ctrl["mod"]).font = body_font
        ws.cell(row=i, column=9, value=ctrl["high"]).font = body_font
        ws.cell(row=i, column=10, value=ctrl["status"]).font = body_font

        # Wrap text and align top for all cells
        for c in range(1, 11):
            ws.cell(row=i, column=c).alignment = wrap

        # Alternate row banding
        if i % 2 == 0:
            for c in range(1, 11):
                ws.cell(row=i, column=c).fill = band_fill

    # Freeze header row
    ws.freeze_panes = "A2"


def create_config_workbook(output_path: str = None):
    if output_path is None:
        output_path = str(Path(__file__).parent / "xlsx2docx_config.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 55

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    key_font = Font(name="Calibri", size=10)
    value_font = Font(name="Calibri", size=10)
    note_font = Font(name="Calibri", size=9, italic=True, color="666666")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Header row
    row = 1
    for col, text in [(1, "Setting"), (2, "Value"), (3, "Description")]:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    def add_section(title):
        nonlocal row
        row += 1
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
        ws.cell(row=row, column=1, value=title).font = section_font

    def add_setting(key, default, description):
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=key).font = key_font
        val_cell = ws.cell(row=row, column=2, value=default)
        val_cell.font = value_font
        val_cell.border = thin_border
        ws.cell(row=row, column=3, value=description).font = note_font
        # Light border on all cells in the row
        ws.cell(row=row, column=1).border = thin_border

    # ── Data Layout ──
    add_section("Data Layout")
    add_setting("heading_column", "Control ID",
                "Column name whose values become Word headings (case-insensitive)")
    add_setting("heading_level", 2,
                "Heading level: 1=H1, 2=H2, 3=H3, 4=H4")
    add_setting("header_row", 1,
                "Row number containing column headers (1-based)")
    add_setting("data_start_row", 2,
                "First row of data (1-based)")
    add_setting("skip_empty_heading", "TRUE",
                "Skip rows where heading column is empty (TRUE/FALSE)")
    add_setting("exclude_columns", "",
                "Comma-separated column names to exclude (e.g. Status,Notes)")
    add_setting("include_columns", "",
                "Comma-separated columns to include (empty = all). Overrides exclude.")
    add_setting("rename_columns", "",
                "OldName=NewLabel pairs, comma-separated (e.g. Desc=Description)")

    # ── Document Identity ──
    add_section("Document Identity")
    add_setting("document_title", "",
                "Document title — used in header/footer {document_title} token")
    add_setting("organization_name", "",
                "Organization name — {organization_name} token")
    add_setting("classification", "",
                "Classification label — {classification} token")
    add_setting("author", "",
                "Author name — {author} token")
    add_setting("version", "1.0",
                "Document version — {version} token")
    add_setting("date", "",
                "Date (YYYY-MM-DD). Leave blank for today's date. — {date} token")

    # ── Page Layout ──
    add_section("Page Layout")
    add_setting("page_size", "Letter",
                "Letter or A4")
    add_setting("orientation", "Portrait",
                "Portrait or Landscape")
    add_setting("margin_top", 1.0,
                "Top margin in inches")
    add_setting("margin_bottom", 1.0,
                "Bottom margin in inches")
    add_setting("margin_left", 1.25,
                "Left margin in inches")
    add_setting("margin_right", 1.25,
                "Right margin in inches")

    # ── Body Text ──
    add_section("Body Text")
    add_setting("body_font", "Calibri",
                "Body text font name")
    add_setting("body_size", 11,
                "Body text font size (pt)")
    add_setting("body_line_spacing", 1.15,
                "Line spacing multiplier (1.0=single, 1.5, 2.0=double)")

    # ── Heading Style ──
    add_section("Heading Style")
    add_setting("heading_font", "Calibri",
                "Heading font name")
    add_setting("heading_size", 14,
                "Heading font size (pt)")
    add_setting("heading_bold", "TRUE",
                "Bold headings (TRUE/FALSE)")
    add_setting("heading_color", "2F5496",
                "Heading color — 6-digit hex, no # (e.g. 2F5496)")

    # ── Label Style (key-value labels) ──
    add_section("Label Style (key-value bold prefix)")
    add_setting("label_font", "Calibri",
                "Label font name")
    add_setting("label_size", 11,
                "Label font size (pt)")
    add_setting("label_bold", "TRUE",
                "Bold labels (TRUE/FALSE)")
    add_setting("label_color", "2F5496",
                "Label color — 6-digit hex")

    # ── Header ──
    add_section("Page Header")
    add_setting("header_left", "{organization_name}",
                "Left header text (supports tokens)")
    add_setting("header_center", "",
                "Center header text")
    add_setting("header_right", "{document_title}",
                "Right header text")
    add_setting("header_font", "Arial",
                "Header font name")
    add_setting("header_size", 8,
                "Header font size (pt)")
    add_setting("header_color", "666666",
                "Header text color — 6-digit hex")

    # ── Footer ──
    add_section("Page Footer")
    add_setting("footer_left", "{classification}",
                "Left footer text (supports tokens)")
    add_setting("footer_center", "Page {page} of {pages}",
                "Center footer — {page} and {pages} become live Word fields")
    add_setting("footer_right", "{date}",
                "Right footer text")
    add_setting("footer_font", "Arial",
                "Footer font name")
    add_setting("footer_size", 8,
                "Footer font size (pt)")
    add_setting("footer_color", "666666",
                "Footer text color — 6-digit hex")

    # ── Misc ──
    add_section("Miscellaneous")
    add_setting("add_sheet_heading", "FALSE",
                "Insert sheet name as H1 before controls (TRUE/FALSE)")

    # Freeze top row
    ws.freeze_panes = "A2"

    # Protect column A (setting names) from accidental edits — light protection
    # Users edit column B only. Column C is descriptions.
    ws.protection.sheet = False  # Not locked — just visual guidance

    # --- Update Config sheet to reflect the demo data ---
    # Set heading_column to match the demo "Controls" sheet
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "heading_column":
            ws.cell(row=r, column=2).value = "Control ID"
            break
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "document_title":
            ws.cell(row=r, column=2).value = "NIST SP 800-53 Security Controls (Sample)"
            break
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "organization_name":
            ws.cell(row=r, column=2).value = "Demo Organization"
            break
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "classification":
            ws.cell(row=r, column=2).value = "Unclassified"
            break

    # ── Add demo NIST 800-53 Controls sheet ──
    add_demo_controls_sheet(wb)

    wb.save(output_path)
    print(f"Config workbook created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_config_workbook()
