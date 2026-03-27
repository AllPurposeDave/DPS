#!/usr/bin/env python3
"""
create_acronym_template.py — Generate a blank acronym intake Excel template.

Creates a pre-formatted .xlsx file with two sheets:
  1. "Acronym Definitions"  — master glossary for human reviewers to fill in.
  2. "Per Document"         — per-doc mapping consumed by docx2md's
                              excel_lookup_list source type.

Run this once to create the template, fill it in, then point docx2md at it.

USAGE:
  python "Misc/Acronym Finder/create_acronym_template.py"
  python "Misc/Acronym Finder/create_acronym_template.py" --output ./output/acronym_intake.xlsx
  python "Misc/Acronym Finder/create_acronym_template.py" --config dps_config.yaml
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Style constants (match acronym_finder.py palette)
# ---------------------------------------------------------------------------

_BLUE_FILL   = PatternFill("solid", fgColor="2F5496")
_GREEN_FILL  = PatternFill("solid", fgColor="375623")
_GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
_YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")

_WHITE_BOLD  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_BODY_FONT   = Font(name="Arial", size=10)
_NOTE_FONT   = Font(name="Arial", size=9, italic=True, color="595959")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP   = Alignment(vertical="top", wrap_text=True)
_LEFT   = Alignment(vertical="top")


def _style_header(ws, row: int, col_count: int, fill=_BLUE_FILL):
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.font      = _WHITE_BOLD
        c.fill      = fill
        c.alignment = _CENTER
        c.border    = _THIN_BORDER


def _style_data(ws, row: int, col_count: int, fill=None):
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.font      = _BODY_FONT
        c.border    = _THIN_BORDER
        c.alignment = _LEFT
        if fill:
            c.fill = fill


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Sheet 1 — Acronym Definitions (master glossary)
# ---------------------------------------------------------------------------

_DEFINITIONS_HEADERS = [
    "Acronym",
    "Full Name / Definition",
    "Category",
    "Notes",
]

_SAMPLE_DEFINITIONS = [
    ("MFA",     "Multi-Factor Authentication",                    "Security",    ""),
    ("NIST",    "National Institute of Standards and Technology", "Compliance",  ""),
    ("DPS",     "Document Processing System",                     "Internal",    "Internal tool name"),
    ("RAG",     "Retrieval-Augmented Generation",                 "AI/ML",       ""),
    ("FedRAMP", "Federal Risk and Authorization Management Program", "Compliance", ""),
]

_CATEGORY_EXAMPLES = (
    "Security, Compliance, AI/ML, Infrastructure, Internal, Regulatory, "
    "Networking, Cloud, HR, Legal, Finance"
)


def _build_definitions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Acronym Definitions"

    # --- Instruction banner (row 1) ---
    ws.merge_cells("A1:D1")
    banner = ws.cell(row=1, column=1,
                     value=(
                         "ACRONYM INTAKE TEMPLATE  —  Fill in the rows below. "
                         "Add one acronym per row. The 'Per Document' sheet maps "
                         "acronyms to specific source files for docx2md metadata."
                     ))
    banner.font      = Font(name="Arial", bold=True, size=10, color="1F3864")
    banner.fill      = PatternFill("solid", fgColor="DCE6F1")
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # --- Column headers (row 2) ---
    for col, h in enumerate(_DEFINITIONS_HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(_DEFINITIONS_HEADERS))

    # --- Sample rows ---
    for i, (acr, defn, cat, note) in enumerate(_SAMPLE_DEFINITIONS, 3):
        ws.cell(row=i, column=1, value=acr)
        ws.cell(row=i, column=2, value=defn)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=note)
        fill = _GRAY_FILL if i % 2 == 0 else None
        _style_data(ws, i, len(_DEFINITIONS_HEADERS), fill=fill)

    # --- Empty input rows ---
    start_empty = 3 + len(_SAMPLE_DEFINITIONS)
    for i in range(start_empty, start_empty + 50):
        _style_data(ws, i, len(_DEFINITIONS_HEADERS),
                    fill=_GRAY_FILL if i % 2 == 0 else None)

    # --- Note row below the data ---
    note_row = start_empty + 50
    ws.merge_cells(f"A{note_row}:D{note_row}")
    note_cell = ws.cell(row=note_row, column=1,
                        value=f"Suggested categories: {_CATEGORY_EXAMPLES}")
    note_cell.font      = _NOTE_FONT
    note_cell.alignment = _WRAP

    # --- Column widths ---
    _set_col_width(ws, 1, 14)
    _set_col_width(ws, 2, 48)
    _set_col_width(ws, 3, 18)
    _set_col_width(ws, 4, 36)

    # --- Freeze panes & filter ---
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:D{start_empty + 49}"


# ---------------------------------------------------------------------------
# Sheet 2 — Per Document (consumed by docx2md excel_lookup_list)
# ---------------------------------------------------------------------------

_PER_DOC_HEADERS = [
    "Document",
    "Acronym",
    "Occurrences",
    "Found In",
    "Definition(s) Detected",
]

_SAMPLE_PER_DOC = [
    ("PolicyDoc-SEC-2024-001.docx", "MFA",     3, "body, table", "Multi-Factor Authentication"),
    ("PolicyDoc-SEC-2024-001.docx", "NIST",    7, "body",        "National Institute of Standards and Technology"),
    ("PolicyDoc-IT-2024-042.docx",  "FedRAMP", 2, "body",        "Federal Risk and Authorization Management Program"),
    ("PolicyDoc-IT-2024-042.docx",  "RAG",     1, "body",        ""),
]

_PER_DOC_NOTE = (
    "This sheet is read by docx2md when you use:\n"
    "  source: \"excel_lookup_list:<this_file>:Per Document:Document:Acronym\"\n"
    "Column A must match the exact .docx filename. "
    "You can paste rows from the acronym_finder output Excel or fill in manually. "
    "Rows with no Definition are flagged yellow by the acronym finder."
)


def _build_per_doc_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Per Document")

    # --- Instruction banner (row 1) ---
    ws.merge_cells("A1:E1")
    banner = ws.cell(row=1, column=1, value=_PER_DOC_NOTE)
    banner.font      = Font(name="Arial", size=9, italic=True, color="1F3864")
    banner.fill      = PatternFill("solid", fgColor="DCE6F1")
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 60

    # --- Headers (row 2) ---
    for col, h in enumerate(_PER_DOC_HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(_PER_DOC_HEADERS), fill=_GREEN_FILL)

    # --- Sample rows ---
    for i, row_data in enumerate(_SAMPLE_PER_DOC, 3):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        fill = _YELLOW_FILL if not row_data[4] else (_GRAY_FILL if i % 2 == 0 else None)
        _style_data(ws, i, len(_PER_DOC_HEADERS), fill=fill)

    # --- Empty input rows ---
    start_empty = 3 + len(_SAMPLE_PER_DOC)
    for i in range(start_empty, start_empty + 100):
        _style_data(ws, i, len(_PER_DOC_HEADERS),
                    fill=_GRAY_FILL if i % 2 == 0 else None)

    # --- Column widths ---
    _set_col_width(ws, 1, 42)   # Document
    _set_col_width(ws, 2, 14)   # Acronym
    _set_col_width(ws, 3, 14)   # Occurrences
    _set_col_width(ws, 4, 22)   # Found In
    _set_col_width(ws, 5, 50)   # Definition(s)

    # --- Freeze & filter ---
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{start_empty + 99}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _resolve_output_path(args) -> str:
    if args.output:
        return os.path.abspath(args.output)

    # Place next to dps_config.yaml in output/
    if args.config and os.path.isfile(args.config):
        config_dir = os.path.dirname(os.path.abspath(args.config))
        return os.path.join(config_dir, "output", "acronym_intake_template.xlsx")

    # Default: output/ next to this script's project root
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent.parent
    return str(project_root / "output" / "acronym_intake_template.xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="Generate a blank acronym intake Excel template."
    )
    parser.add_argument(
        "--output", "-o",
        metavar="PATH",
        help="Where to write the .xlsx file (default: output/acronym_intake_template.xlsx)",
    )
    parser.add_argument(
        "--config", "-c",
        metavar="CONFIG",
        help="Path to dps_config.yaml (used to resolve output directory)",
    )
    args = parser.parse_args()

    out_path = _resolve_output_path(args)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    _build_definitions_sheet(wb)
    _build_per_doc_sheet(wb)

    wb.save(out_path)
    print(f"Template created: {out_path}")
    print()
    print("Sheets:")
    print("  'Acronym Definitions'  — master glossary (Acronym, Definition, Category, Notes)")
    print("  'Per Document'         — per-doc mapping for docx2md excel_lookup_list")
    print()
    print("To use with docx2md, add to dps_config.yaml > docx2md > metadata_fields:")
    print(f'  - name: "acronyms"')
    print(f'    source: "excel_lookup_list:{os.path.relpath(out_path)}:Per Document:Document:Acronym"')
    print(f'    default: ""')


if __name__ == "__main__":
    main()
