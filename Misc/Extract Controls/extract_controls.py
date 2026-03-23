"""
extract_controls.py -- Compliance Document Control Extractor
=============================================================

WHAT IT DOES:
    Reads .docx compliance/security policy documents from ./input/,
    extracts structured control data (IDs, descriptions, supplemental
    guidance, and document metadata), then writes everything into a
    single consolidated Excel file at ./output/4 - controls/controls_output.xlsx.

PIPELINE:
    .docx file
      -> python-docx paragraph extraction (text, bold flag, style name)
      -> metadata scan (purpose, scope, applicability from first 40 paragraphs)
      -> section header detection (heading styles, numbered titles, all-caps, bold)
      -> control block segmentation (regex-matched IDs delimit blocks)
      -> text classification (description / supplemental guidance / miscellaneous)
      -> whitelist/blacklist filtering
      -> Excel row assembly
      -> consolidated output

HOW TO RUN:
    python extract_controls.py

DIRECTORY STRUCTURE:
    ./input/               Place your .docx files here
    ./output/4 - controls/ Created automatically. Excel, checkpoint, and error log land here.

DEPENDENCIES:
    pip install python-docx pyyaml openpyxl
    Everything else is Python stdlib. No API keys. Fully offline.

CONFIGURATION:
    Edit config-control-extractor.yaml to customize control ID patterns,
    guidance keywords, whitelist/blacklist, heading detection, and metadata triggers.
"""

import re
import os
import json
import logging
import fnmatch
from pathlib import Path
from dataclasses import dataclass, asdict, fields

import yaml
from docx import Document
from openpyxl import Workbook, load_workbook


# ============================================================
# SECTION: CONFIGURATION
# Loads settings from config-control-extractor.yaml.
# Falls back to hardcoded defaults if the file is missing.
# ============================================================

DEFAULT_CONFIG = {
    "control_id_patterns": [
        r'\b[A-Z]{2,4}[-.]?\d{1,3}[-.]\d{2,4}\b',
    ],
    "whitelist": [],
    "blacklist": [],
    "guidance_keywords": [
        "implementation guidance",
        "implementation:",
        "guidelines:",
        "how to implement",
        "supplemental guidance",
    ],
    "metadata_triggers": {
        "purpose": ["purpose", "objective", "intent"],
        "scope": ["scope", "coverage", "boundary"],
        "applicability": ["applicability", "applies to", "applies for"],
    },
    "heading_detection": {
        "use_word_heading_style": True,
        "section_keyword_pattern": r'^[Ss]ection\s+\d{1,2}',
        "numbered_title_pattern": r'^\d{1,2}\.?\d{0,2}\s+[A-Z][a-zA-Z\s]{3,50}$',
        "detect_allcaps": True,
        "allcaps_max_length": 80,
        "allcaps_min_words": 3,
        "detect_bold_short": True,
        "bold_max_length": 60,
    },
}


def load_config(config_path=None):
    """Load configuration from YAML file, falling back to defaults.

    Args:
        config_path: Path to the YAML config file. If None, looks for
                     config-control-extractor.yaml next to this script.

    Returns:
        dict: Merged configuration with compiled regex patterns.
    """
    if config_path is None:
        config_path = Path(__file__).parent / "config-control-extractor.yaml"
    else:
        config_path = Path(config_path)

    config = dict(DEFAULT_CONFIG)

    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            user_config = yaml.safe_load(f) or {}
        # Merge user config over defaults
        for key, value in user_config.items():
            if key == "heading_detection" and isinstance(value, dict):
                config["heading_detection"] = {**DEFAULT_CONFIG["heading_detection"], **value}
            elif key == "metadata_triggers" and isinstance(value, dict):
                config["metadata_triggers"] = {**DEFAULT_CONFIG["metadata_triggers"], **value}
            else:
                config[key] = value
        print(f"Config loaded from: {config_path}")
    else:
        print(f"Config not found at {config_path}, using defaults.")

    # Compile regex patterns from config
    config["_control_id_regexes"] = [
        re.compile(p) for p in config["control_id_patterns"]
    ]

    # Build combined control ID regex for use in text matching
    combined_pattern = "|".join(f"({p})" for p in config["control_id_patterns"])
    config["_control_id_combined"] = re.compile(combined_pattern)

    # Build guidance trigger regex
    escaped_keywords = [re.escape(kw) for kw in config["guidance_keywords"]]
    guidance_pattern = "|".join(escaped_keywords)
    config["_guidance_regex"] = re.compile(f"(?i)({guidance_pattern})")

    # Build metadata trigger regexes
    config["_metadata_regexes"] = {}
    all_metadata_words = []
    for field_name, keywords in config["metadata_triggers"].items():
        escaped = [re.escape(kw) for kw in keywords]
        pattern = r'\b(' + "|".join(escaped) + r')\b'
        config["_metadata_regexes"][field_name] = re.compile(pattern, re.IGNORECASE)
        all_metadata_words.extend(keywords)

    # Build metadata label strip regex
    escaped_all = [re.escape(w) for w in all_metadata_words]
    strip_pattern = r'(?i)^\s*(' + "|".join(escaped_all) + r')\s*[:\-]?\s*'
    config["_metadata_label_strip"] = re.compile(strip_pattern)

    # Compile heading detection regexes
    hd = config["heading_detection"]
    config["_section_keyword_regex"] = re.compile(hd["section_keyword_pattern"])
    config["_numbered_title_regex"] = re.compile(hd["numbered_title_pattern"])

    return config


# ============================================================
# SECTION: DATA MODEL
# One instance per extracted control row in the output.
# ============================================================

@dataclass
class ControlRow:
    """Represents a single row in the output Excel file.

    Fields map 1:1 to Excel columns. Default empty strings keep the
    output clean when a field isn't found in the source document.
    """
    source_file: str = ""
    section_header: str = ""
    control_id: str = ""
    control_description: str = ""
    supplemental_guidance: str = ""
    miscellaneous: str = ""
    purpose: str = ""
    scope: str = ""
    applicability: str = ""


# ============================================================
# SECTION: LOGGING SETUP
# Errors go to both stderr and ./output/errors.log so you can
# review failures after a batch run.
# ============================================================

def setup_logging():
    """Configure logging to write errors to ./output/errors.log.

    Returns:
        logging.Logger: Configured logger instance.
    """
    output_dir = Path("./output/4 - controls")
    output_dir.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("extract_controls")
    logger.setLevel(logging.DEBUG)

    # File handler -- captures warnings and errors for post-run review
    file_handler = logging.FileHandler(output_dir / "errors.log", mode="a")
    file_handler.setLevel(logging.WARNING)
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    )
    logger.addHandler(file_handler)

    # Stream handler -- so errors also show up in your terminal
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.WARNING)
    stream_handler.setFormatter(
        logging.Formatter("%(levelname)s: %(message)s")
    )
    logger.addHandler(stream_handler)

    return logger


logger = setup_logging()


# ============================================================
# SECTION: PARAGRAPH EXTRACTION
# Turns a .docx file into a flat list of paragraph dicts that
# downstream functions can reason about without touching python-docx.
# ============================================================

def extract_paragraphs_from_docx(filepath):
    """Open a .docx file and return a list of paragraph metadata dicts.

    Each dict contains the raw text, whether the first run is bold,
    and the Word style name. Downstream functions use these signals
    to detect headers, metadata labels, and control blocks.

    Args:
        filepath: Path (str or Path) to the .docx file.

    Returns:
        List[dict]: Each dict has keys:
            - "text" (str): The paragraph's full text.
            - "bold" (bool): True if the first run is bold.
            - "style" (str): The Word style name (e.g. "Heading 1").
    """
    document = Document(str(filepath))
    paragraphs = []

    for paragraph in document.paragraphs:
        text = paragraph.text.strip()

        # Determine bold status from the first run.
        # Some paragraphs have zero runs (empty or whitespace-only).
        first_run_bold = False
        if paragraph.runs:
            first_run_bold = bool(paragraph.runs[0].bold)

        style_name = paragraph.style.name if paragraph.style else ""

        paragraphs.append({
            "text": text,
            "bold": first_run_bold,
            "style": style_name,
        })

    return paragraphs


# ============================================================
# SECTION: SECTION HEADER DETECTION
# Identifies paragraph-level section boundaries so we can tag
# each control with the section it belongs to.
# ============================================================

def is_section_header(paragraph_dict, config):
    """Determine whether a paragraph is a section header.

    Checks multiple signals in priority order and returns on the
    first match. Each signal can be toggled via config.

    Args:
        paragraph_dict: Dict with keys "text", "bold", "style".
        config: Loaded configuration dict.

    Returns:
        Tuple[bool, str]: (is_header, cleaned_header_text).
            If not a header, returns (False, "").
    """
    text = paragraph_dict["text"]
    style = paragraph_dict["style"]
    bold = paragraph_dict["bold"]
    hd = config["heading_detection"]

    # Skip empty paragraphs and lines that contain a control ID --
    # a control ID line is data, not a section boundary
    if not text or config["_control_id_combined"].search(text):
        return (False, "")

    # Signal 1: Word style explicitly says "Heading"
    if hd["use_word_heading_style"] and "Heading" in style:
        return (True, text)

    # Signal 2: Starts with "Section N"
    if config["_section_keyword_regex"].match(text):
        return (True, text)

    # Signal 3: Numbered title like "2.1 Access Control Policy"
    if config["_numbered_title_regex"].match(text):
        return (True, text)

    # Signal 4: All-caps, short, no trailing period, at least N words
    if hd["detect_allcaps"]:
        word_count = len(text.split())
        if (text == text.upper()
                and len(text) < hd["allcaps_max_length"]
                and not text.endswith(".")
                and word_count >= hd["allcaps_min_words"]):
            return (True, text)

    # Signal 5: Bold, short, no trailing period
    if hd["detect_bold_short"]:
        if bold and len(text) < hd["bold_max_length"] and not text.endswith("."):
            return (True, text)

    return (False, "")


# ============================================================
# SECTION: METADATA EXTRACTION
# Pulls Purpose, Scope, and Applicability from the document's
# opening pages. These repeat on every row for that document.
# ============================================================

def extract_metadata(paragraphs, config):
    """Scan the first 40 paragraphs for document-level metadata.

    Looks for trigger keywords (purpose, scope, applicability) and
    captures all paragraph text from the trigger through to the next
    trigger or section header.

    Args:
        paragraphs: List of paragraph dicts from extract_paragraphs_from_docx.
        config: Loaded configuration dict.

    Returns:
        dict: Keys "purpose", "scope", "applicability" with string values.
              Empty string if not found.
    """
    metadata = {field: "" for field in config["_metadata_regexes"]}
    scan_limit = min(40, len(paragraphs))

    trigger_positions = []  # (index, field_name)

    for index in range(scan_limit):
        text = paragraphs[index]["text"]
        if not text:
            continue

        for field_name, pattern in config["_metadata_regexes"].items():
            if pattern.search(text):
                trigger_positions.append((index, field_name))
                break

    for position_index, (para_index, field_name) in enumerate(trigger_positions):
        if metadata[field_name]:
            continue

        collected_lines = []
        if position_index + 1 < len(trigger_positions):
            stop_index = trigger_positions[position_index + 1][0]
        else:
            stop_index = scan_limit

        for capture_index in range(para_index, stop_index):
            paragraph_text = paragraphs[capture_index]["text"]
            if not paragraph_text:
                continue

            if capture_index > para_index:
                is_header, _ = is_section_header(paragraphs[capture_index], config)
                if is_header:
                    break

            collected_lines.append(paragraph_text)

        raw_metadata = " ".join(collected_lines)
        cleaned = config["_metadata_label_strip"].sub("", raw_metadata, count=1)
        metadata[field_name] = clean_text(cleaned)

    return metadata


# ============================================================
# SECTION: CONTROL BLOCK SEGMENTATION
# Walks all paragraphs, splits them into blocks delimited by
# control IDs, then separates description / guidance / misc.
# ============================================================

def find_control_blocks(paragraphs, config):
    """Identify and segment individual control blocks from paragraphs.

    A control block starts at any paragraph containing a control ID regex
    match and extends until the next control ID or section header. After
    block boundaries are established, each block is split into:
      - control_description (text before guidance keywords)
      - supplemental_guidance (text after guidance keywords)
      - miscellaneous (unclassified text)

    Args:
        paragraphs: List of paragraph dicts from extract_paragraphs_from_docx.
        config: Loaded configuration dict.

    Returns:
        List[dict]: Each dict has keys:
            - "control_id" (str)
            - "control_description" (str)
            - "supplemental_guidance" (str)
            - "miscellaneous" (str)
            - "section_header" (str)
    """
    control_id_regex = config["_control_id_combined"]
    blocks = []
    current_section_header = ""
    active_block = None

    for paragraph_dict in paragraphs:
        text = paragraph_dict["text"]

        is_header, header_text = is_section_header(paragraph_dict, config)
        if is_header:
            if active_block is not None:
                blocks.append(active_block)
                active_block = None
            current_section_header = header_text
            continue

        if not text:
            continue

        control_id_matches = control_id_regex.findall(text)
        # findall with groups returns tuples; flatten to get the matched strings
        if control_id_matches:
            # Flatten: each match is a tuple of groups, take the first non-empty
            if isinstance(control_id_matches[0], tuple):
                control_id_matches = [
                    next(g for g in m if g) for m in control_id_matches
                ]

            if active_block is not None:
                blocks.append(active_block)

            for match_index, control_id in enumerate(control_id_matches):
                if match_index == 0:
                    active_block = {
                        "control_id": control_id,
                        "raw_lines": [text],
                        "section_header": current_section_header,
                    }
                else:
                    blocks.append({
                        "control_id": control_id,
                        "raw_lines": [text],
                        "section_header": current_section_header,
                    })
        elif active_block is not None:
            active_block["raw_lines"].append(text)

    if active_block is not None:
        blocks.append(active_block)

    # Split each block's raw text into description / guidance / miscellaneous
    guidance_regex = config["_guidance_regex"]
    processed_blocks = []

    for block in blocks:
        joined_text = "\n".join(block["raw_lines"])
        split_result = guidance_regex.split(joined_text, maxsplit=1)

        if len(split_result) >= 3:
            # split_result = [before, matched_trigger, after]
            description_text = clean_text(split_result[0])
            guidance_text = clean_text(split_result[2])

            # Check if there's a second guidance trigger in the "after" portion
            # Text after a second trigger goes to miscellaneous
            second_split = guidance_regex.split(split_result[2], maxsplit=1)
            if len(second_split) >= 3:
                guidance_text = clean_text(second_split[0])
                miscellaneous_text = clean_text(second_split[2])
            else:
                miscellaneous_text = ""
        else:
            # No guidance trigger found
            # The first sentence is the description; if there are multiple
            # sentences, remaining text goes to miscellaneous
            sentences = joined_text.split(".", 1)
            if len(sentences) > 1 and sentences[1].strip():
                description_text = clean_text(sentences[0] + ".")
                miscellaneous_text = clean_text(sentences[1])
            else:
                description_text = clean_text(joined_text)
                miscellaneous_text = ""
            guidance_text = ""

        processed_blocks.append({
            "control_id": block["control_id"],
            "control_description": description_text,
            "supplemental_guidance": guidance_text,
            "miscellaneous": miscellaneous_text,
            "section_header": block["section_header"],
        })

    return processed_blocks


# ============================================================
# SECTION: WHITELIST / BLACKLIST FILTERING
# Filters control blocks by ID using exact match or prefix wildcard.
# ============================================================

def matches_filter(control_id, filter_list):
    """Check if a control ID matches any entry in a filter list.

    Supports exact matches and prefix wildcards (e.g., "AC-*").

    Args:
        control_id: The control ID string to check.
        filter_list: List of filter patterns (exact IDs or "PREFIX-*").

    Returns:
        bool: True if the control ID matches any filter entry.
    """
    for pattern in filter_list:
        if pattern.endswith("*"):
            if control_id.startswith(pattern[:-1]):
                return True
        else:
            if control_id == pattern:
                return True
    return False


def apply_filters(control_blocks, config):
    """Apply whitelist and blacklist filters to control blocks.

    Args:
        control_blocks: List of control block dicts.
        config: Loaded configuration dict.

    Returns:
        List[dict]: Filtered control blocks.
    """
    whitelist = config.get("whitelist", [])
    blacklist = config.get("blacklist", [])

    filtered = control_blocks

    # Whitelist: if non-empty, keep only matching controls
    if whitelist:
        filtered = [b for b in filtered if matches_filter(b["control_id"], whitelist)]

    # Blacklist: remove matching controls
    if blacklist:
        filtered = [b for b in filtered if not matches_filter(b["control_id"], blacklist)]

    return filtered


# ============================================================
# SECTION: TEXT CLEANING
# Normalizes whitespace and strips non-printable characters.
# ============================================================

def clean_text(text):
    """Normalize whitespace and remove non-printable characters.

    Compliance docs often have leftover formatting artifacts --
    non-breaking spaces, control characters, extra whitespace from
    copy-paste. This function produces clean strings.

    Args:
        text: Raw string to clean.

    Returns:
        str: Cleaned string with collapsed whitespace and only
             printable ASCII plus newlines.
    """
    text = text.strip()
    # Remove non-printable characters but keep newlines for readability
    text = re.sub(r'[^\x20-\x7E\n]', '', text)
    # Collapse runs of whitespace (including newlines) to a single space
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


# ============================================================
# SECTION: SINGLE DOCUMENT PROCESSING
# Orchestrates the full pipeline for one .docx file.
# ============================================================

def process_single_document(filepath, config):
    """Run the full extraction pipeline on a single .docx file.

    Coordinates paragraph extraction, metadata scanning, section
    header tracking, control block segmentation, and filtering.

    Args:
        filepath: Path to the .docx file.
        config: Loaded configuration dict.

    Returns:
        List[ControlRow]: One row per extracted control.
    """
    filepath = Path(filepath)
    source_file = filepath.name

    paragraphs = extract_paragraphs_from_docx(filepath)

    if not paragraphs:
        logger.warning("Empty document (0 paragraphs): %s", source_file)
        print(f"  WARNING: {source_file} has 0 paragraphs, skipping.")
        return []

    # Pull document-level metadata from the opening pages
    metadata = extract_metadata(paragraphs, config)

    purpose_found = "Yes" if metadata.get("purpose") else "No"
    scope_found = "Yes" if metadata.get("scope") else "No"
    applicability_found = "Yes" if metadata.get("applicability") else "No"
    print(f"  Metadata found -- Purpose: {purpose_found} | "
          f"Scope: {scope_found} | Applicability: {applicability_found}")

    # Segment paragraphs into control blocks
    control_blocks = find_control_blocks(paragraphs, config)
    print(f"  Controls extracted: {len(control_blocks)}")

    # Apply whitelist/blacklist filtering
    control_blocks = apply_filters(control_blocks, config)
    if config.get("whitelist") or config.get("blacklist"):
        print(f"  Controls after filtering: {len(control_blocks)}")

    # Assemble ControlRow instances
    rows = []
    for block in control_blocks:
        row = ControlRow(
            source_file=source_file,
            section_header=block["section_header"],
            control_id=block["control_id"],
            control_description=block["control_description"],
            supplemental_guidance=block["supplemental_guidance"],
            miscellaneous=block["miscellaneous"],
            purpose=metadata.get("purpose", ""),
            scope=metadata.get("scope", ""),
            applicability=metadata.get("applicability", ""),
        )
        rows.append(row)

    return rows


# ============================================================
# SECTION: CHECKPOINTING
# Tracks which files have been successfully processed so you can
# re-run after fixing errors without re-processing everything.
# ============================================================

def load_checkpoint(checkpoint_path):
    """Load the list of already-processed filenames from disk.

    Args:
        checkpoint_path: Path to checkpoint.json.

    Returns:
        list: Filenames that have already been processed.
    """
    if checkpoint_path.exists():
        with open(checkpoint_path, "r") as checkpoint_file:
            return json.load(checkpoint_file)
    return []


def save_checkpoint(checkpoint_path, completed_files):
    """Write the updated list of processed filenames to disk.

    Args:
        checkpoint_path: Path to checkpoint.json.
        completed_files: List of filenames to persist.
    """
    with open(checkpoint_path, "w") as checkpoint_file:
        json.dump(completed_files, checkpoint_file, indent=2)


# ============================================================
# SECTION: EXCEL OUTPUT
# Writes control rows to an .xlsx file using openpyxl.
# ============================================================

EXCEL_COLUMNS = [
    "source_file", "section_header", "control_id",
    "control_description", "supplemental_guidance", "miscellaneous",
    "purpose", "scope", "applicability",
]


def write_rows_to_excel(output_path, rows):
    """Append control rows to an Excel file, creating it if needed.

    Args:
        output_path: Path to the .xlsx output file.
        rows: List of ControlRow instances to write.
    """
    if output_path.exists() and output_path.stat().st_size > 0:
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Controls"
        ws.append(EXCEL_COLUMNS)

    for row in rows:
        row_dict = asdict(row)
        ws.append([row_dict[col] for col in EXCEL_COLUMNS])

    wb.save(output_path)


# ============================================================
# SECTION: MAIN ENTRY POINT
# Scans ./input/, processes each file, writes the consolidated Excel.
# ============================================================

def main():
    """Entry point. Scans ./input/ for .docx files, extracts controls,
    writes ./output/4 - controls/controls_output.xlsx.

    Respects checkpointing -- previously processed files are skipped.
    Each file is wrapped in try/except so one bad document won't kill
    the entire batch.
    """
    # Load configuration
    config = load_config()

    input_dir = Path("./input")
    output_dir = Path("./output/4 - controls")
    output_dir.mkdir(parents=True, exist_ok=True)

    checkpoint_path = output_dir / "checkpoint.json"
    output_xlsx_path = output_dir / "controls_output.xlsx"

    if not input_dir.exists():
        print("ERROR: ./input/ directory not found. Create it and add .docx files.")
        return

    # Gather all .docx files, sorted for deterministic ordering
    # Filter out Word temp files (~$*.docx) created when a doc is open
    docx_files = sorted(
        f for f in input_dir.glob("*.docx") if not f.name.startswith("~$")
    )

    if not docx_files:
        print("No .docx files found in ./input/. Nothing to do.")
        return

    completed_files = load_checkpoint(checkpoint_path)
    total_controls = 0
    files_processed = 0
    files_skipped = 0
    error_count = 0

    for docx_path in docx_files:
        filename = docx_path.name

        if filename in completed_files:
            print(f"Skipping (already processed): {filename}")
            files_skipped += 1
            continue

        print(f"Processing: {filename}")

        try:
            rows = process_single_document(docx_path, config)

            if rows:
                write_rows_to_excel(output_xlsx_path, rows)

            total_controls += len(rows)

            # Mark this file as done AFTER writing Excel
            completed_files.append(filename)
            save_checkpoint(checkpoint_path, completed_files)
            files_processed += 1

        except Exception as error:
            error_count += 1
            logger.error("Failed to process %s: %s", filename, error, exc_info=True)
            print(f"  ERROR processing {filename}: {error}")

    print(
        f"\nDone. {files_processed} files | {total_controls} controls | "
        f"{error_count} errors. Output: {output_xlsx_path}"
    )
    if files_skipped:
        print(f"  ({files_skipped} files skipped via checkpoint)")


# ============================================================
# SECTION: SCRIPT EXECUTION
# ============================================================

if __name__ == "__main__":
    main()
