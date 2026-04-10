#!/usr/bin/env python3
"""
dataverse_export.py — Export control catalog data to Dataverse-ready CSV.

Reads one or two Excel workbooks (Control Catalog with NIST-style IDs and/or
Standard Controls with org-style IDs), normalizes them into a unified schema,
builds a composite RAG-content column for M365 CoPilot Studio, and writes a
Dataverse-importable CSV plus a schema reference document.

All settings are driven by dataverse_export_config.yaml.

Usage:
  python dataverse_export.py
  python dataverse_export.py --config dataverse_export_config.yaml
  python dataverse_export.py --catalog-xlsx ../../input/nist_catalog.xlsx
  python dataverse_export.py --standard-xlsx ../../input/org_controls.xlsx
  python dataverse_export.py --catalog-xlsx catalog.xlsx --standard-xlsx standard.xlsx
  python dataverse_export.py -v
"""

from __future__ import annotations

import argparse
import codecs
import copy
import csv
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


# ── Default configuration ────────────────────────────────────────────────────

DEFAULT_CONFIG: dict[str, Any] = {
    "input": {
        "catalog_xlsx": "",
        "catalog_sheet": "Controls",
        "standard_xlsx": "",
        "standard_sheet": "Controls",
        "column_map": {
            "Control ID": "control_id",
            "Control Name": "control_name",
            "Description": "description",
            "Supplemental Guidance": "supplemental_guidance",
            "Baseline": "baseline",
            "Purpose": "purpose",
            "Scope": "scope",
            "Applicability": "applicability",
            "Compliance Date": "compliance_date",
            "Published URL": "published_url",
            "Family": "family_name",
        },
        "header_row": 1,
        "data_start_row": 2,
    },
    "output": {
        "directory": "../../output/dataverse_export/",
        "controls_csv": "dps_controls_import.csv",
        "schema_reference": "dataverse_schema_reference.md",
    },
    "family_map": {
        "AC": "Access Control",
        "AT": "Awareness and Training",
        "AU": "Audit and Accountability",
        "CA": "Assessment, Authorization, and Monitoring",
        "CM": "Configuration Management",
        "CP": "Contingency Planning",
        "IA": "Identification and Authentication",
        "IR": "Incident Response",
        "MA": "Maintenance",
        "MP": "Media Protection",
        "PE": "Physical and Environmental Protection",
        "PL": "Planning",
        "PM": "Program Management",
        "PS": "Personnel Security",
        "PT": "PII Processing and Transparency",
        "RA": "Risk Assessment",
        "SA": "System and Services Acquisition",
        "SC": "System and Communications Protection",
        "SI": "System and Information Integrity",
        "SR": "Supply Chain Risk Management",
    },
    "rag_content": {
        "fields": [
            "control_id",
            "control_name",
            "family_name",
            "baseline",
            "description",
            "supplemental_guidance",
            "purpose",
            "scope",
        ],
    },
}

# Internal field names — defines the CSV column order
SCHEMA_FIELDS: list[str] = [
    "control_id",
    "control_name",
    "family",
    "family_name",
    "description",
    "supplemental_guidance",
    "baseline",
    "purpose",
    "scope",
    "applicability",
    "compliance_date",
    "published_url",
    "source",
    "rag_content",
]

# Labels for the RAG content block
RAG_LABELS: dict[str, str] = {
    "control_id": "Control",
    "control_name": "",          # appended to control_id line
    "family_name": "Family",
    "baseline": "Baseline",
    "description": "Description",
    "supplemental_guidance": "Supplemental Guidance",
    "purpose": "Purpose",
    "scope": "Scope",
}


# ── Configuration helpers ────────────────────────────────────────────────────

def _deep_merge(base: dict, override: dict) -> dict:
    """Recursively merge *override* into a copy of *base*."""
    merged = copy.deepcopy(base)
    for key, val in override.items():
        if key in merged and isinstance(merged[key], dict) and isinstance(val, dict):
            merged[key] = _deep_merge(merged[key], val)
        else:
            merged[key] = copy.deepcopy(val)
    return merged


def load_config(config_path: str) -> dict[str, Any]:
    """Load YAML config and deep-merge with defaults."""
    config_path = os.path.abspath(config_path)
    if not os.path.isfile(config_path):
        logging.warning("Config not found at '%s' — using defaults.", config_path)
        return copy.deepcopy(DEFAULT_CONFIG)

    with open(config_path, encoding="utf-8") as fh:
        loaded = yaml.safe_load(fh) or {}

    config = _deep_merge(DEFAULT_CONFIG, loaded)
    logging.info("Loaded config: %s", config_path)
    return config


# ── Excel reader ─────────────────────────────────────────────────────────────

def _build_reverse_column_map(column_map: dict[str, str]) -> dict[str, str]:
    """Invert {Excel header: internal field} and lowercase the keys for matching."""
    return {k.strip().lower(): v for k, v in column_map.items()}


def load_controls(
    xlsx_path: str,
    sheet_name: str,
    source_label: str,
    config: dict[str, Any],
) -> list[dict[str, str]]:
    """Read an Excel sheet and return a list of control dicts."""
    xlsx_path = os.path.abspath(xlsx_path)
    if not os.path.isfile(xlsx_path):
        logging.error("File not found: %s", xlsx_path)
        return []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        logging.error("Sheet '%s' not found in '%s'. Available: %s",
                       sheet_name, xlsx_path, wb.sheetnames)
        wb.close()
        return []

    ws = wb[sheet_name]
    input_cfg = config["input"]
    header_row = int(input_cfg.get("header_row", 1))
    data_start = int(input_cfg.get("data_start_row", 2))
    column_map = input_cfg.get("column_map", {})
    reverse_map = _build_reverse_column_map(column_map)
    family_map = config.get("family_map", {})

    # Read header row to build column-index -> field-name mapping
    col_field: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        header_lower = str(cell.value).strip().lower()
        if header_lower in reverse_map:
            col_field[cell.column] = reverse_map[header_lower]

    if not col_field:
        logging.warning("No mapped columns found in '%s' sheet '%s'. "
                        "Check column_map in config.", xlsx_path, sheet_name)
        wb.close()
        return []

    logging.info("Mapped %d columns in '%s' / '%s': %s",
                 len(col_field), os.path.basename(xlsx_path), sheet_name,
                 {c: f for c, f in sorted(col_field.items())})

    controls: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=data_start):
        record: dict[str, str] = {f: "" for f in SCHEMA_FIELDS}
        record["source"] = source_label

        for cell in row:
            field = col_field.get(cell.column)
            if field and cell.value is not None:
                record[field] = str(cell.value).strip()

        # Skip rows with no control_id
        if not record["control_id"]:
            continue

        # Normalize multiline fields: strip \r to avoid delimiter issues
        for field in ("description", "supplemental_guidance", "purpose",
                      "scope", "rag_content"):
            if record[field]:
                record[field] = record[field].replace("\r\n", "\n").replace("\r", "\n")

        # Derive family code and name for catalog controls
        if not record["family"]:
            record["family"] = _derive_family_code(record["control_id"])
        if not record["family_name"] and record["family"]:
            record["family_name"] = family_map.get(record["family"], "")

        controls.append(record)

    wb.close()
    logging.info("Loaded %d controls from '%s' / '%s' (source: %s)",
                 len(controls), os.path.basename(xlsx_path), sheet_name, source_label)
    return controls


def _derive_family_code(control_id: str) -> str:
    """Extract the alpha family prefix from a NIST-style control ID.

    Examples: "AC-2" -> "AC", "CM-6(1)" -> "CM", "IA-2" -> "IA".
    Returns "" if the ID doesn't match NIST format.
    """
    m = re.match(r"^([A-Z]{2,4})[-.]", control_id.upper())
    return m.group(1) if m else ""


# ── RAG content builder ──────────────────────────────────────────────────────

def build_rag_content(controls: list[dict[str, str]], config: dict[str, Any]) -> None:
    """Populate the rag_content field on each control dict (in-place)."""
    fields = config.get("rag_content", {}).get("fields", list(RAG_LABELS.keys()))

    for ctrl in controls:
        parts: list[str] = []

        for field in fields:
            value = ctrl.get(field, "").strip()
            if not value:
                continue

            label = RAG_LABELS.get(field, field)

            if field == "control_id":
                # First line: "Control AC-2: Account Management"
                name = ctrl.get("control_name", "").strip()
                line = f"Control {value}"
                if name:
                    line += f": {name}"
                parts.append(line)
            elif field == "control_name":
                # Already included with control_id above
                continue
            elif field in ("description", "supplemental_guidance", "purpose", "scope"):
                # Block fields — label on its own line
                parts.append(f"\n{label}:\n{value}")
            else:
                # Inline fields
                parts.append(f"{label}: {value}")

        ctrl["rag_content"] = "\n".join(parts).strip()


# ── CSV writer ───────────────────────────────────────────────────────────────

def write_dataverse_csv(
    controls: list[dict[str, str]],
    output_path: str,
) -> str:
    """Write controls to a Dataverse-importable CSV (UTF-8 with BOM)."""
    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=SCHEMA_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for ctrl in controls:
            writer.writerow(ctrl)

    logging.info("Wrote %d rows to: %s", len(controls), output_path)
    return output_path


# ── Schema reference generator ───────────────────────────────────────────────

SCHEMA_REFERENCE = """\
# Dataverse Table Schema: dps_controls

Use this reference to create the Dataverse table manually.
The CSV import file (`dps_controls_import.csv`) maps 1:1 to these columns.

## Table: dps_controls

| Column | Display Name | Dataverse Type | Max Length | Required | Notes |
|--------|-------------|---------------|-----------|----------|-------|
| `control_id` | Control ID | Single Line Text | 50 | Yes | Primary lookup key. NIST: "AC-2", Org: "ABCD02.003" |
| `control_name` | Control Name | Single Line Text | 255 | No | e.g. "Account Management" |
| `family` | Family Code | Choice | — | No | NIST family code: AC, AT, AU, CA, CM, CP, IA, IR, MA, MP, PE, PL, PM, PS, PT, RA, SA, SC, SI, SR. Blank for Standard Controls. |
| `family_name` | Family Name | Single Line Text | 100 | No | Full name, e.g. "Access Control" |
| `description` | Description | Multiline Text | 100,000 | No | Control description |
| `supplemental_guidance` | Supplemental Guidance | Multiline Text | 100,000 | No | Additional guidance text |
| `baseline` | Baseline | Single Line Text | 100 | No | e.g. "Low, Moderate, High" |
| `purpose` | Purpose | Multiline Text | 10,000 | No | Document/control purpose |
| `scope` | Scope | Multiline Text | 10,000 | No | Scope of applicability |
| `applicability` | Applicability | Single Line Text | 500 | No | Applicability statement |
| `compliance_date` | Compliance Date | Single Line Text | 50 | No | Target compliance date |
| `published_url` | Published URL | Single Line Text (URL) | 2,000 | No | Reference URL |
| `source` | Source | Choice | — | No | "Control Catalog" or "Standard Controls" |
| `rag_content` | RAG Content | Multiline Text | 100,000 | No | Composite search column for CoPilot Studio |

## Choice Field: Family Code

| Value | Label |
|-------|-------|
| AC | AC |
| AT | AT |
| AU | AU |
| CA | CA |
| CM | CM |
| CP | CP |
| IA | IA |
| IR | IR |
| MA | MA |
| MP | MP |
| PE | PE |
| PL | PL |
| PM | PM |
| PS | PS |
| PT | PT |
| RA | RA |
| SA | SA |
| SC | SC |
| SI | SI |
| SR | SR |

## Choice Field: Source

| Value | Label |
|-------|-------|
| Control Catalog | Control Catalog |
| Standard Controls | Standard Controls |

## CSV-to-Dataverse Column Mapping

The CSV uses short internal names. Map them to Dataverse display names during import:

| CSV Header | Dataverse Display Name |
|------------|----------------------|
| `control_id` | Control ID |
| `control_name` | Control Name |
| `family` | Family Code |
| `family_name` | Family Name |
| `description` | Description |
| `supplemental_guidance` | Supplemental Guidance |
| `baseline` | Baseline |
| `purpose` | Purpose |
| `scope` | Scope |
| `applicability` | Applicability |
| `compliance_date` | Compliance Date |
| `published_url` | Published URL |
| `source` | Source |
| `rag_content` | RAG Content |

> **Note:** Dataverse prepends your solution publisher prefix to logical column
> names (e.g., `cr_control_id` or `new_control_id`). The display names above
> are what you'll see in the UI and what the CSV import wizard maps against.

## Setup Instructions

1. **Create the table** in your Dataverse environment:
   - Go to Power Apps > Tables > New table
   - Name: `dps_controls`, Display name: "DPS Controls"
   - Primary column: `control_id` (rename from default "Name")

2. **Add columns** per the schema above:
   - Single Line Text columns: set max length as noted
   - Multiline Text columns: use "Text Area" format, set max length
   - Choice columns: **create the option sets listed above before importing**
     — unmapped Choice values are silently dropped during import
   - URL column: use "URL" format for `published_url`

3. **Set up an alternate key** on `control_id`:
   - Go to the table > Keys > New key
   - Name: "Control ID Key", Column: `control_id`
   - This enables upsert (update-or-insert) for future re-imports,
     so updated controls overwrite existing rows instead of creating duplicates

4. **Import the CSV**:
   - Go to the table > Import > Import data
   - Select `dps_controls_import.csv`
   - Map CSV columns to Dataverse columns using the mapping table above
   - The CSV uses UTF-8 with BOM encoding for Dataverse compatibility

5. **Configure CoPilot Studio**:
   - Add the `dps_controls` table as a knowledge source
   - Ensure the `rag_content` column is indexed for search
   - The `rag_content` column contains all relevant control information
     in a single text block optimized for natural-language retrieval
"""


def write_schema_reference(output_path: str) -> str:
    """Write the Dataverse schema reference Markdown file."""
    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as fh:
        fh.write(SCHEMA_REFERENCE)

    logging.info("Schema reference written to: %s", output_path)
    return output_path


# ── CLI & main ───────────────────────────────────────────────────────────────

def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(
        description="Export control catalog data to Dataverse-ready CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python dataverse_export.py
  python dataverse_export.py --catalog-xlsx ../../input/nist_catalog.xlsx
  python dataverse_export.py --standard-xlsx ../../input/org_controls.xlsx
  python dataverse_export.py --catalog-xlsx cat.xlsx --standard-xlsx std.xlsx -v
""",
    )
    parser.add_argument(
        "--config", "-c",
        default=None,
        help="Path to config YAML (default: dataverse_export_config.yaml in script dir)",
    )
    parser.add_argument(
        "--catalog-xlsx",
        default=None,
        help="Path to Control Catalog Excel file (NIST-style IDs)",
    )
    parser.add_argument(
        "--standard-xlsx",
        default=None,
        help="Path to Standard Controls Excel file (org-style IDs)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output directory (overrides config)",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose (DEBUG) logging",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Resolve config
    if args.config:
        config_path = os.path.abspath(args.config)
    else:
        config_path = os.path.join(os.path.dirname(__file__),
                                   "dataverse_export_config.yaml")
    config = load_config(config_path)

    # CLI overrides
    input_cfg = config["input"]
    catalog_xlsx = args.catalog_xlsx or input_cfg.get("catalog_xlsx", "")
    standard_xlsx = args.standard_xlsx or input_cfg.get("standard_xlsx", "")

    if not catalog_xlsx and not standard_xlsx:
        print("No input files specified. Use --catalog-xlsx and/or --standard-xlsx,\n"
              "or set input.catalog_xlsx / input.standard_xlsx in the config.",
              file=sys.stderr)
        sys.exit(1)

    # Resolve relative paths:
    #   CLI args → relative to CWD
    #   Config values → relative to config file's directory
    config_dir = os.path.dirname(config_path)

    def _resolve(path: str, from_cli: bool) -> str:
        if not path or os.path.isabs(path):
            return path
        base = os.getcwd() if from_cli else config_dir
        return os.path.normpath(os.path.join(base, path))

    catalog_from_cli = args.catalog_xlsx is not None
    standard_from_cli = args.standard_xlsx is not None
    catalog_xlsx = _resolve(catalog_xlsx, catalog_from_cli)
    standard_xlsx = _resolve(standard_xlsx, standard_from_cli)

    # Output directory
    output_cfg = config["output"]
    output_dir = args.output or output_cfg.get("directory", "output/dataverse_export/")
    output_from_cli = args.output is not None
    output_dir = _resolve(output_dir, output_from_cli)

    # Load controls from both sources
    all_controls: list[dict[str, str]] = []

    if catalog_xlsx:
        catalog_sheet = input_cfg.get("catalog_sheet", "Controls")
        rows = load_controls(catalog_xlsx, catalog_sheet, "Control Catalog", config)
        all_controls.extend(rows)
        print(f"  Catalog:  {len(rows)} controls from {os.path.basename(catalog_xlsx)}")

    if standard_xlsx:
        standard_sheet = input_cfg.get("standard_sheet", "Controls")
        rows = load_controls(standard_xlsx, standard_sheet, "Standard Controls", config)
        all_controls.extend(rows)
        print(f"  Standard: {len(rows)} controls from {os.path.basename(standard_xlsx)}")

    if not all_controls:
        print("No controls loaded from any source.", file=sys.stderr)
        sys.exit(1)

    # Validate: warn on duplicate control_ids (primary key in Dataverse)
    seen_ids: dict[str, int] = {}
    for ctrl in all_controls:
        cid = ctrl["control_id"]
        seen_ids[cid] = seen_ids.get(cid, 0) + 1
    dupes = {cid: count for cid, count in seen_ids.items() if count > 1}
    if dupes:
        logging.warning("Duplicate control_id values found (Dataverse primary key "
                        "must be unique): %s", dupes)

    # Build RAG content
    build_rag_content(all_controls, config)

    # Write outputs
    csv_path = os.path.join(output_dir, output_cfg.get("controls_csv",
                                                         "dps_controls_import.csv"))
    schema_path = os.path.join(output_dir, output_cfg.get("schema_reference",
                                                            "dataverse_schema_reference.md"))

    write_dataverse_csv(all_controls, csv_path)
    write_schema_reference(schema_path)

    print(f"\nDone. {len(all_controls)} controls exported.")
    print(f"  CSV:    {os.path.relpath(csv_path)}")
    print(f"  Schema: {os.path.relpath(schema_path)}")


if __name__ == "__main__":
    main()
