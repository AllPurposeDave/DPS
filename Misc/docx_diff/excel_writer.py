"""Write the docx-diff Excel workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from comments import Comment, CommentsDiff
from metadata import Metadata, flatten_for_diff
from pairing import (
    MATCHED, MATCHED_FUZZY, MATCHED_NORMALIZED,
    ORPHAN_DIFF1, ORPHAN_DIFF2, Pair,
)
from signals import DIFF1, DIFF2, MISSING, NA, TIE, Signal
from text_diff import ParagraphDiff
from tracked_changes import TrackedChange, TrackedDiff

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
ORPHAN_FILL = PatternFill("solid", fgColor="FFC7CE")
TIE_FILL = PatternFill("solid", fgColor="E7E6E6")
DIFF1_FILL = PatternFill("solid", fgColor="C6EFCE")
DIFF2_FILL = PatternFill("solid", fgColor="BDD7EE")

DEL_FONT = InlineFont(color="C00000", strike=True)
INS_FONT = InlineFont(color="006100", b=True)
EQ_FONT = InlineFont(color="000000")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


@dataclass
class PairResult:
    pair: Pair
    meta1: Metadata | None = None
    meta2: Metadata | None = None
    para_diff: list[ParagraphDiff] = field(default_factory=list)
    tracked: TrackedDiff | None = None
    comments: CommentsDiff | None = None
    signals: list[Signal] = field(default_factory=list)


def _fmt(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value
    if isinstance(value, (set, frozenset)):
        return ", ".join(sorted(str(v) for v in value))
    return value


def _write_header(ws, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is None:
                continue
            if isinstance(cell.value, CellRichText):
                text = "".join(
                    block.text if isinstance(block, TextBlock) else str(block)
                    for block in cell.value
                )
            elif isinstance(cell.value, datetime):
                text = cell.value.isoformat(sep=" ", timespec="seconds")
            else:
                text = str(cell.value)
            for line in text.splitlines() or [""]:
                if len(line) > longest:
                    longest = len(line)
        ws.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _apply_autofilter(ws) -> None:
    if ws.max_row > 1 and ws.max_column > 0:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        )


def _inline_cell(inline_ops: list[tuple[str, str]]) -> CellRichText:
    blocks: list[TextBlock] = []
    for op, text in inline_ops:
        if not text:
            continue
        if op == "equal":
            blocks.append(TextBlock(EQ_FONT, text))
        elif op == "delete":
            blocks.append(TextBlock(DEL_FONT, text))
        elif op == "insert":
            blocks.append(TextBlock(INS_FONT, text))
    if not blocks:
        blocks.append(TextBlock(EQ_FONT, ""))
    return CellRichText(*blocks)


def _points_fill(points_to: str) -> PatternFill | None:
    if points_to == DIFF1:
        return DIFF1_FILL
    if points_to == DIFF2:
        return DIFF2_FILL
    if points_to == TIE:
        return TIE_FILL
    return None


def _write_summary(ws, results: list[PairResult]) -> None:
    headers = [
        "File", "Match", "Status", "Archive bytes equal?", "document.xml equal?",
        "# Metadata fields differ", "# Paragraphs changed",
        "# Tracked revisions Δ", "# Comments Δ", "Signal tally",
    ]
    _write_header(ws, headers)
    for r, result in enumerate(results, start=2):
        pair = result.pair
        ws.cell(row=r, column=1, value=pair.name).alignment = TOP
        ws.cell(row=r, column=2, value=pair.match_kind).alignment = TOP
        orphan = pair.match_kind in (ORPHAN_DIFF1, ORPHAN_DIFF2)
        if orphan:
            status = "ORPHAN"
        elif result.meta1 and result.meta2:
            archive_eq = result.meta1.hashes.get("archive_sha256") == result.meta2.hashes.get("archive_sha256")
            doc_eq = result.meta1.hashes.get("document_xml_sha256") == result.meta2.hashes.get("document_xml_sha256")
            changed = (not archive_eq) or result.para_diff or (
                result.tracked and (result.tracked.only_in_1 or result.tracked.only_in_2)
            ) or (
                result.comments and (result.comments.only_in_1 or result.comments.only_in_2)
            )
            status = "IDENTICAL" if (archive_eq and not changed) else ("CHANGED" if changed else "METADATA ONLY")
        else:
            status = "ERROR"
        ws.cell(row=r, column=3, value=status).alignment = TOP

        if result.meta1 and result.meta2:
            archive_eq = result.meta1.hashes.get("archive_sha256") == result.meta2.hashes.get("archive_sha256")
            doc_eq = result.meta1.hashes.get("document_xml_sha256") == result.meta2.hashes.get("document_xml_sha256")
            ws.cell(row=r, column=4, value="YES" if archive_eq else "NO").alignment = TOP
            ws.cell(row=r, column=5, value="YES" if doc_eq else "NO").alignment = TOP
        else:
            ws.cell(row=r, column=4, value="").alignment = TOP
            ws.cell(row=r, column=5, value="").alignment = TOP

        meta_diffs = 0
        if result.meta1 and result.meta2:
            f1 = flatten_for_diff(result.meta1)
            f2 = flatten_for_diff(result.meta2)
            for key in set(f1) | set(f2):
                if f1.get(key) != f2.get(key):
                    meta_diffs += 1
        ws.cell(row=r, column=6, value=meta_diffs).alignment = TOP
        ws.cell(row=r, column=7, value=len(result.para_diff)).alignment = TOP
        tc_delta = 0
        if result.tracked:
            tc_delta = len(result.tracked.only_in_1) + len(result.tracked.only_in_2)
        ws.cell(row=r, column=8, value=tc_delta).alignment = TOP
        com_delta = 0
        if result.comments:
            com_delta = len(result.comments.only_in_1) + len(result.comments.only_in_2)
        ws.cell(row=r, column=9, value=com_delta).alignment = TOP

        tally: dict[str, int] = {}
        for s in result.signals:
            tally[s.points_to] = tally.get(s.points_to, 0) + 1
        parts = []
        for key in (DIFF1, DIFF2, TIE, MISSING, NA):
            if tally.get(key):
                parts.append(f"{key}: {tally[key]}")
        ws.cell(row=r, column=10, value=" · ".join(parts) or "—").alignment = TOP

        if orphan:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = ORPHAN_FILL
    _apply_autofilter(ws)
    _autosize(ws)


def _write_signals(ws, results: list[PairResult]) -> None:
    _write_header(ws, ["File", "Signal", "Diff 1 value", "Diff 2 value", "Points to", "Notes"])
    row = 2
    for result in results:
        for s in result.signals:
            ws.cell(row=row, column=1, value=result.pair.name).alignment = TOP
            ws.cell(row=row, column=2, value=s.name).alignment = TOP
            ws.cell(row=row, column=3, value=_fmt(s.diff1_value)).alignment = TOP
            ws.cell(row=row, column=4, value=_fmt(s.diff2_value)).alignment = TOP
            pts = ws.cell(row=row, column=5, value=s.points_to)
            pts.alignment = TOP
            fill = _points_fill(s.points_to)
            if fill is not None:
                pts.fill = fill
            ws.cell(row=row, column=6, value=s.note).alignment = WRAP
            row += 1
    _apply_autofilter(ws)
    _autosize(ws)


def _write_metadata(ws, results: list[PairResult]) -> None:
    _write_header(ws, ["File", "Category", "Field", "Diff 1", "Diff 2", "Equal?", "Δ"])
    row = 2
    for result in results:
        if not (result.meta1 and result.meta2):
            continue
        f1 = flatten_for_diff(result.meta1)
        f2 = flatten_for_diff(result.meta2)
        all_keys = sorted(set(f1) | set(f2))
        for cat, field_name in all_keys:
            v1 = f1.get((cat, field_name))
            v2 = f2.get((cat, field_name))
            equal = v1 == v2
            delta: Any = ""
            if isinstance(v1, datetime) and isinstance(v2, datetime):
                diff = (v2 - v1).total_seconds()
                days = diff / 86400.0
                delta = f"{days:+.2f} d"
            elif isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
                delta = v2 - v1
            ws.cell(row=row, column=1, value=result.pair.name).alignment = TOP
            ws.cell(row=row, column=2, value=cat).alignment = TOP
            ws.cell(row=row, column=3, value=field_name).alignment = TOP
            ws.cell(row=row, column=4, value=_fmt(v1)).alignment = TOP
            ws.cell(row=row, column=5, value=_fmt(v2)).alignment = TOP
            ws.cell(row=row, column=6, value="YES" if equal else "NO").alignment = TOP
            ws.cell(row=row, column=7, value=delta).alignment = TOP
            if not equal:
                for col in (4, 5):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
            row += 1
    _apply_autofilter(ws)
    _autosize(ws)


def _write_text_diff(ws, results: list[PairResult]) -> None:
    _write_header(ws, [
        "File", "Location", "¶ # (Diff 1)", "¶ # (Diff 2)", "Kind",
        "Diff 1 text", "Diff 2 text", "Inline diff",
    ])
    row = 2
    for result in results:
        for d in result.para_diff:
            ws.cell(row=row, column=1, value=result.pair.name).alignment = TOP
            ws.cell(row=row, column=2, value=d.location).alignment = TOP
            ws.cell(row=row, column=3, value=d.index1 if d.index1 is not None else "").alignment = TOP
            ws.cell(row=row, column=4, value=d.index2 if d.index2 is not None else "").alignment = TOP
            ws.cell(row=row, column=5, value=d.kind).alignment = TOP
            ws.cell(row=row, column=6, value=d.text1).alignment = WRAP
            ws.cell(row=row, column=7, value=d.text2).alignment = WRAP
            inline = ws.cell(row=row, column=8, value=_inline_cell(d.inline_ops))
            inline.alignment = WRAP
            row += 1
    _apply_autofilter(ws)
    _autosize(ws, max_width=80)


def _write_tracked(ws, results: list[PairResult]) -> None:
    _write_header(ws, ["File", "Source", "Type", "Author", "Date", "Location", "Text"])
    row = 2
    for result in results:
        if not result.tracked:
            continue
        for src_label, changes in (
            ("Diff 1 only", result.tracked.only_in_1),
            ("Diff 2 only", result.tracked.only_in_2),
            ("Both", result.tracked.in_both),
        ):
            for c in changes:
                ws.cell(row=row, column=1, value=result.pair.name).alignment = TOP
                ws.cell(row=row, column=2, value=src_label).alignment = TOP
                ws.cell(row=row, column=3, value=c.rev_type).alignment = TOP
                ws.cell(row=row, column=4, value=c.author).alignment = TOP
                ws.cell(row=row, column=5, value=_fmt(c.date)).alignment = TOP
                ws.cell(row=row, column=6, value=c.location).alignment = TOP
                ws.cell(row=row, column=7, value=c.text).alignment = WRAP
                row += 1
    _apply_autofilter(ws)
    _autosize(ws, max_width=80)


def _write_comments(ws, results: list[PairResult]) -> None:
    _write_header(ws, [
        "File", "Source", "Comment ID", "Parent ID", "Author", "Date",
        "Anchor text", "Comment text", "Resolved?",
    ])
    row = 2
    for result in results:
        if not result.comments:
            continue
        for src_label, comments in (
            ("Diff 1 only", result.comments.only_in_1),
            ("Diff 2 only", result.comments.only_in_2),
            ("Both", result.comments.in_both),
        ):
            for c in comments:
                ws.cell(row=row, column=1, value=result.pair.name).alignment = TOP
                ws.cell(row=row, column=2, value=src_label).alignment = TOP
                ws.cell(row=row, column=3, value=c.comment_id).alignment = TOP
                ws.cell(row=row, column=4, value=c.parent_id or "").alignment = TOP
                ws.cell(row=row, column=5, value=c.author).alignment = TOP
                ws.cell(row=row, column=6, value=_fmt(c.date)).alignment = TOP
                ws.cell(row=row, column=7, value=c.anchor_text).alignment = WRAP
                ws.cell(row=row, column=8, value=c.text).alignment = WRAP
                ws.cell(row=row, column=9, value="YES" if c.resolved else "NO").alignment = TOP
                row += 1
    _apply_autofilter(ws)
    _autosize(ws, max_width=80)


def _write_orphans(ws, results: list[PairResult]) -> None:
    _write_header(ws, ["Filename", "Side", "Size (bytes)", "mtime", "Notes"])
    row = 2
    for result in results:
        pair = result.pair
        if pair.match_kind == ORPHAN_DIFF1:
            side = "Diff 1"
            meta = result.meta1
            name = pair.path1.name if pair.path1 else pair.name
        elif pair.match_kind == ORPHAN_DIFF2:
            side = "Diff 2"
            meta = result.meta2
            name = pair.path2.name if pair.path2 else pair.name
        else:
            continue
        ws.cell(row=row, column=1, value=name).alignment = TOP
        ws.cell(row=row, column=2, value=side).alignment = TOP
        if meta:
            ws.cell(row=row, column=3, value=meta.fs.get("size", "")).alignment = TOP
            ws.cell(row=row, column=4, value=_fmt(meta.fs.get("mtime"))).alignment = TOP
            ws.cell(row=row, column=5, value=meta.error or "").alignment = WRAP
        row += 1
    _apply_autofilter(ws)
    _autosize(ws)


def write_workbook(results: list[PairResult], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_summary(wb.create_sheet("Summary"), results)
    _write_signals(wb.create_sheet("Signals"), results)
    _write_metadata(wb.create_sheet("Metadata"), results)
    _write_text_diff(wb.create_sheet("Text Diff"), results)
    _write_tracked(wb.create_sheet("Tracked Changes"), results)
    _write_comments(wb.create_sheet("Comments"), results)
    _write_orphans(wb.create_sheet("Orphans"), results)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
