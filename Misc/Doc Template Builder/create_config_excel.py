#!/usr/bin/env python3
"""
create_config_excel.py — Generate doc_template_config.xlsx

Creates a fresh Excel config workbook for the Doc Template Builder.
Fill in the values, then run build_doc_from_config.py to produce the Word doc.

Usage:
  python create_config_excel.py                          # default: doc_template_config.xlsx
  python create_config_excel.py -o my_template_config.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Styling constants (matches DPS Excel styling) ───────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10, color="2F5496")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DESC_FONT = Font(name="Arial", size=9, color="666666")
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)


# ── Helper functions ─────────────────────────────────────────────────────────

def _style_headers(ws, headers: List[str]):
    """Write header row and apply standard styling."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _add_subheader(ws, text: str, ncols: int = 3):
    """Add a sub-header row (# prefixed in column A)."""
    row = [f"# {text}"] + [""] * (ncols - 1)
    ws.append(row)
    r = ws.max_row
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER


def _add_setting(ws, setting: str, value: Any, description: str = ""):
    """Add a Setting | Value | Description row."""
    ws.append([setting, value, description])
    ws.cell(row=ws.max_row, column=3).font = DESC_FONT


def _add_bool_validation(ws, col_letter: str, min_row: int, max_row: int):
    """Add TRUE/FALSE dropdown validation to a column range."""
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv.error = "Please select TRUE or FALSE"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


def _add_enum_validation(ws, col_letter: str, row: int, options: List[str]):
    """Add enum dropdown validation to a specific cell."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = f"Please select one of: {', '.join(options)}"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{row}")


def _auto_column_widths(ws):
    """Auto-fit column widths with wrapping for wide columns."""
    WRAP_THRESHOLD = 50
    wrap_cols = set()
    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is not None:
                max_width = max(max_width, len(str(val)))
        letter = get_column_letter(col_idx)
        if max_width > WRAP_THRESHOLD:
            ws.column_dimensions[letter].width = 65
            wrap_cols.add(col_idx)
        else:
            ws.column_dimensions[letter].width = min(max_width + 4, 65)
    if wrap_cols:
        for data_row in ws.iter_rows(min_row=2):
            for cell in data_row:
                if cell.column in wrap_cols:
                    cell.alignment = WRAP_ALIGN


def _finalize_sheet(ws):
    """Apply auto-widths and autofilter."""
    _auto_column_widths(ws)
    if ws.max_column and ws.max_row:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


# ── Heading defaults ─────────────────────────────────────────────────────────

# (key, default_value, description)
H1_DEFAULTS = [
    ("placeholder_text", "Section Title",    "Placeholder text inserted as a demo H1 paragraph"),
    ("font_name",        "Calibri",           "Font family name"),
    ("font_size",        16,                  "Font size in pt"),
    ("bold",             "TRUE",              "Bold? TRUE or FALSE"),
    ("italic",           "FALSE",             "Italic? TRUE or FALSE"),
    ("underline",        "FALSE",             "Underline? TRUE or FALSE"),
    ("all_caps",         "FALSE",             "ALL CAPS? TRUE or FALSE"),
    ("color",            "2F5496",            "Hex RGB color — 6 digits, no # prefix (e.g. 2F5496)"),
    ("alignment",        "left",              "Paragraph alignment: left or center"),
    ("space_before",     12,                  "Space before paragraph in pt"),
    ("space_after",      6,                   "Space after paragraph in pt"),
    ("keep_with_next",   "TRUE",              "Keep heading with the next paragraph? TRUE or FALSE"),
    ("numbering_prefix", "N.0",              "Numbering pattern hint. H1 uses N.0 (1.0, 2.0) or Roman (I., II.)"),
]

H2_DEFAULTS = [
    ("placeholder_text", "Subsection Title",  "Placeholder text inserted as a demo H2 paragraph"),
    ("font_name",        "Calibri",           "Font family name"),
    ("font_size",        14,                  "Font size in pt"),
    ("bold",             "TRUE",              "Bold? TRUE or FALSE"),
    ("italic",           "FALSE",             "Italic? TRUE or FALSE"),
    ("underline",        "FALSE",             "Underline? TRUE or FALSE"),
    ("all_caps",         "FALSE",             "ALL CAPS? TRUE or FALSE"),
    ("color",            "2F5496",            "Hex RGB color — 6 digits, no # prefix (e.g. 2F5496)"),
    ("alignment",        "left",              "Paragraph alignment: left or center"),
    ("space_before",     10,                  "Space before paragraph in pt"),
    ("space_after",      4,                   "Space after paragraph in pt"),
    ("keep_with_next",   "TRUE",              "Keep heading with the next paragraph? TRUE or FALSE"),
    ("numbering_prefix", "N.N",              "Numbering pattern hint. H2 uses N.N (1.1, 2.3) or letter (A., B.)"),
]

H3_DEFAULTS = [
    ("placeholder_text", "Sub-subsection Title", "Placeholder text inserted as a demo H3 paragraph"),
    ("font_name",        "Calibri",           "Font family name"),
    ("font_size",        12,                  "Font size in pt"),
    ("bold",             "TRUE",              "Bold? TRUE or FALSE"),
    ("italic",           "FALSE",             "Italic? TRUE or FALSE"),
    ("underline",        "FALSE",             "Underline? TRUE or FALSE"),
    ("all_caps",         "FALSE",             "ALL CAPS? TRUE or FALSE"),
    ("color",            "2F5496",            "Hex RGB color — 6 digits, no # prefix (e.g. 2F5496)"),
    ("alignment",        "left",              "Paragraph alignment: left or center"),
    ("space_before",     8,                   "Space before paragraph in pt"),
    ("space_after",      4,                   "Space after paragraph in pt"),
    ("keep_with_next",   "TRUE",              "Keep heading with the next paragraph? TRUE or FALSE"),
    ("numbering_prefix", "N.N.N",            "Numbering pattern hint. H3 uses N.N.N (1.1.1, 2.3.1)"),
]

H4_DEFAULTS = [
    ("placeholder_text", "Detail Heading",    "Placeholder text inserted as a demo H4 paragraph"),
    ("font_name",        "Calibri",           "Font family name"),
    ("font_size",        11,                  "Font size in pt"),
    ("bold",             "TRUE",              "Bold? TRUE or FALSE"),
    ("italic",           "TRUE",              "Italic? TRUE or FALSE"),
    ("underline",        "FALSE",             "Underline? TRUE or FALSE"),
    ("all_caps",         "FALSE",             "ALL CAPS? TRUE or FALSE"),
    ("color",            "2F5496",            "Hex RGB color — 6 digits, no # prefix (e.g. 2F5496)"),
    ("alignment",        "left",              "Paragraph alignment: left or center"),
    ("space_before",     6,                   "Space before paragraph in pt"),
    ("space_after",      2,                   "Space after paragraph in pt"),
    ("keep_with_next",   "TRUE",              "Keep heading with the next paragraph? TRUE or FALSE"),
    ("numbering_prefix", "",                  "Numbering pattern hint (blank = unnumbered)"),
]

BOOL_KEYS = {"bold", "italic", "underline", "all_caps", "keep_with_next"}
ALIGNMENT_KEYS = {"alignment"}


# ── Standard section defaults ────────────────────────────────────────────────

# (section_name, include, heading_text, placeholder_body)
STANDARD_SECTIONS = [
    ("Purpose",                  "TRUE",  "1.0 Purpose",
     "This policy establishes the requirements for..."),
    ("Scope",                    "TRUE",  "2.0 Scope",
     "This policy applies to all information systems, personnel, and third parties..."),
    ("Roles and Responsibilities", "TRUE", "3.0 Roles and Responsibilities",
     "The following roles and responsibilities are defined for this policy..."),
    ("Policy Requirements",      "TRUE",  "4.0 Policy Requirements",
     "[Add structured control requirements — use Control ID format: XX-N.NNN]"),
    ("Compliance and Enforcement", "TRUE", "5.0 Compliance and Enforcement",
     "Violations of this policy may result in disciplinary action..."),
    ("References",               "TRUE",  "6.0 References",
     "[List referenced policies, standards, and frameworks]"),
    ("Definitions",              "FALSE", "7.0 Definitions",
     "[Add terms and definitions relevant to this policy]"),
    ("Revision History",         "TRUE",  "8.0 Revision History",
     "(Inserts a 4-column table: Version | Date | Author | Description)"),
    ("Appendix",                 "FALSE", "Appendix A",
     "[Add supplementary material, crosswalk tables, or control matrices]"),
]


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_readme(wb):
    ws = wb.create_sheet("README")
    ws.append(["Doc Template Config Workbook"])
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws.append([])
    instructions = [
        "This workbook controls the heading styles, document settings, and section structure",
        "used by build_doc_from_config.py to generate NIST/RMF-compliant policy document templates.",
        "",
        "SHEETS:",
        "  README           — This help page",
        "  Document         — Body text defaults, margins, line spacing",
        "  Heading Styles   — Font, size, color, spacing for Heading 1–4",
        "  Document Identity— Title, doc ID, classification, owner, dates",
        "  Page Layout      — Page size, orientation",
        "  Header Footer    — Header/footer text, fonts, tokens ({doc_id}, {classification}, etc.)",
        "  Standard Sections— Pre-built NIST policy sections (Purpose, Scope, etc.)",
        "  Table Styles     — Table header/body formatting, borders, banding",
        "  Cover Page       — Optional cover/title page settings",
        "",
        "HOW TO USE:",
        "  1. Edit values in any sheet — each uses:  Setting | Value | Description  columns.",
        "  2. Sub-headers (rows starting with #) separate groups — do NOT delete them.",
        "  3. Boolean fields have TRUE/FALSE dropdowns. Enum fields have option dropdowns.",
        "  4. Empty Value cells use the built-in default.",
        "",
        "HOW TO RUN:",
        "  1. Generate this config (already done):",
        "       python create_config_excel.py",
        "  2. Edit values in this workbook as needed.",
        "  3. Generate the Word doc template:",
        "       python build_doc_from_config.py",
        "       python build_doc_from_config.py --config doc_template_config.xlsx -o my_template.docx",
        "",
        "COLOR FORMAT:",
        "  Enter hex RGB colors as 6 uppercase digits with NO # prefix.",
        "  Examples:  2F5496  (DPS blue)   |   000000  (black)   |   FFFFFF  (white)",
        "",
        "HEADING STYLE KEYS (in 'Heading Styles' sheet):",
        "  h1.*  — controls the Heading 1 style      h3.*  — controls the Heading 3 style",
        "  h2.*  — controls the Heading 2 style      h4.*  — controls the Heading 4 style",
        "",
        "STANDARD SECTIONS (in 'Standard Sections' sheet):",
        "  Heading text uses numbering patterns (1.0, 2.0...) that align with the DPS pipeline's",
        "  heading detection regex. Section names match profiler keywords (purpose, scope, etc.).",
        "  Set Include=FALSE to skip a section. Edit Heading Text and Placeholder Body as needed.",
        "",
        "HEADER/FOOTER TOKENS:",
        "  Use {document_title}, {doc_id}, {classification}, {organization_name}, {effective_date}",
        "  in header/footer text — they resolve from Document Identity values at build time.",
        "  Use {page} and {pages} for dynamic Word page number fields.",
        "",
        "DPS PIPELINE INTEGRATION:",
        "  Templates are optimized for the DPS pipeline (profiler, control extraction, heading fixer,",
        "  section splitter, docx2md, docx2jsonl). The builder writes Document Identity values into",
        "  Word core_properties so docx2md.py extracts correct YAML frontmatter automatically.",
        "  Heading styles use Word built-in styles (Heading 1–4) for pipeline compatibility.",
        "  If your org uses custom heading style names, map them in dps_config.xlsx (custom_style_map).",
    ]
    for line in instructions:
        ws.append([line])
    ws.column_dimensions["A"].width = 90


def _build_document_sheet(wb):
    ws = wb.create_sheet("Document")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Document Defaults")
    _add_setting(ws, "output_filename", "doc_template.docx",
                 "Output .docx filename (relative to where you run the script)")
    _add_setting(ws, "body_font_name",  "Calibri",
                 "Default body text font (applied to Normal style)")
    _add_setting(ws, "body_font_size",  11,
                 "Default body font size in pt")

    _add_subheader(ws, "Body Text Formatting")
    _add_setting(ws, "line_spacing", 1.15,
                 "Line spacing multiplier (1.0, 1.15, 1.5, 2.0)")
    line_spacing_row = ws.max_row
    _add_setting(ws, "paragraph_space_before", 0,
                 "Body paragraph space before in pt")
    _add_setting(ws, "paragraph_space_after", 8,
                 "Body paragraph space after in pt")
    _add_setting(ws, "first_line_indent", 0,
                 "First line indent in inches (0 = none)")
    _add_setting(ws, "text_alignment", "left",
                 "Body text alignment: left or justified")
    text_align_row = ws.max_row

    _add_subheader(ws, "Page Margins (inches)")
    _add_setting(ws, "margin_top",    1.0,  "Top margin in inches")
    _add_setting(ws, "margin_bottom", 1.0,  "Bottom margin in inches")
    _add_setting(ws, "margin_left",   1.25, "Left margin in inches")
    _add_setting(ws, "margin_right",  1.25, "Right margin in inches")

    _add_enum_validation(ws, "B", line_spacing_row, ["1.0", "1.15", "1.5", "2.0"])
    _add_enum_validation(ws, "B", text_align_row, ["left", "justified"])

    _finalize_sheet(ws)


def _build_heading_styles_sheet(wb):
    ws = wb.create_sheet("Heading Styles")
    _style_headers(ws, ["Setting", "Value", "Description"])

    bool_rows = []
    alignment_rows = []

    for level_label, defaults in [
        ("Heading 1", H1_DEFAULTS),
        ("Heading 2", H2_DEFAULTS),
        ("Heading 3", H3_DEFAULTS),
        ("Heading 4", H4_DEFAULTS),
    ]:
        prefix = level_label.replace("Heading ", "h")
        _add_subheader(ws, level_label)
        for key, val, desc in defaults:
            _add_setting(ws, f"{prefix}.{key}", val, desc)
            if key in BOOL_KEYS:
                bool_rows.append(ws.max_row)
            if key in ALIGNMENT_KEYS:
                alignment_rows.append(ws.max_row)

    for r in bool_rows:
        _add_bool_validation(ws, "B", r, r)
    for r in alignment_rows:
        _add_enum_validation(ws, "B", r, ["left", "center"])

    _finalize_sheet(ws)


def _build_document_identity_sheet(wb):
    ws = wb.create_sheet("Document Identity")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Document Identification")
    _add_setting(ws, "document_title", "",
                 "Full policy document title (written to Word core_properties.title)")
    _add_setting(ws, "document_id", "",
                 "Document identifier, e.g. POL-AC-2026-001 (format: [A-Z]+-[A-Z]+-YYYY-NNN)")
    _add_setting(ws, "version", "1.0",
                 "Version number")
    _add_setting(ws, "classification", "Unclassified",
                 "Classification level")
    classification_row = ws.max_row
    _add_setting(ws, "organization_name", "",
                 "Organization name")
    _add_setting(ws, "department", "",
                 "Department or division")

    _add_subheader(ws, "Dates")
    _add_setting(ws, "effective_date", "",
                 "Date policy takes effect (YYYY-MM-DD)")
    _add_setting(ws, "review_date", "",
                 "Last review date (YYYY-MM-DD)")
    _add_setting(ws, "next_review_date", "",
                 "Next scheduled review (YYYY-MM-DD)")

    _add_subheader(ws, "Ownership")
    _add_setting(ws, "document_status", "Draft",
                 "Current document status")
    status_row = ws.max_row
    _add_setting(ws, "document_owner", "",
                 "Owner/author name or role (written to Word core_properties.author)")
    _add_setting(ws, "approving_authority", "",
                 "Approving official name or role")
    _add_setting(ws, "supersedes", "",
                 "Previous document ID/version this replaces")

    _add_enum_validation(ws, "B", classification_row,
                         ["Unclassified", "CUI", "FOUO", "Confidential", "Internal Use Only"])
    _add_enum_validation(ws, "B", status_row,
                         ["Draft", "Under Review", "Approved", "Superseded", "Retired"])

    _finalize_sheet(ws)


def _build_page_layout_sheet(wb):
    ws = wb.create_sheet("Page Layout")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Page Setup")
    _add_setting(ws, "page_size", "Letter",
                 "Page size: Letter (8.5x11) or A4 (8.27x11.69)")
    size_row = ws.max_row
    _add_setting(ws, "orientation", "Portrait",
                 "Page orientation: Portrait or Landscape")
    orient_row = ws.max_row

    _add_enum_validation(ws, "B", size_row, ["Letter", "A4"])
    _add_enum_validation(ws, "B", orient_row, ["Portrait", "Landscape"])

    _finalize_sheet(ws)


def _build_header_footer_sheet(wb):
    ws = wb.create_sheet("Header Footer")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Header")
    _add_setting(ws, "header_left", "{organization_name}",
                 "Left header text. Tokens: {doc_id}, {classification}, {organization_name}")
    _add_setting(ws, "header_center", "",
                 "Center header text")
    _add_setting(ws, "header_right", "{doc_id}",
                 "Right header text")
    _add_setting(ws, "header_font_name", "Arial",
                 "Header font family")
    _add_setting(ws, "header_font_size", 8,
                 "Header font size in pt")
    _add_setting(ws, "header_color", "666666",
                 "Header font color (hex RGB, no # prefix)")

    _add_subheader(ws, "Footer")
    _add_setting(ws, "footer_left", "{classification}",
                 "Left footer text")
    _add_setting(ws, "footer_center", "Page {page} of {pages}",
                 "Center footer. {page}/{pages} become Word field codes")
    _add_setting(ws, "footer_right", "{effective_date}",
                 "Right footer text")
    _add_setting(ws, "footer_font_name", "Arial",
                 "Footer font family")
    _add_setting(ws, "footer_font_size", 8,
                 "Footer font size in pt")
    _add_setting(ws, "footer_color", "666666",
                 "Footer font color (hex RGB, no # prefix)")

    _add_subheader(ws, "Options")
    _add_setting(ws, "different_first_page", "TRUE",
                 "Suppress header/footer on cover/first page")
    _add_bool_validation(ws, "B", ws.max_row, ws.max_row)

    _finalize_sheet(ws)


def _build_standard_sections_sheet(wb):
    ws = wb.create_sheet("Standard Sections")
    headers = ["Section", "Include", "Heading Text", "Placeholder Body"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"

    include_rows = []
    for section_name, include, heading_text, placeholder in STANDARD_SECTIONS:
        ws.append([section_name, include, heading_text, placeholder])
        include_rows.append(ws.max_row)

    for r in include_rows:
        _add_bool_validation(ws, "B", r, r)

    # Auto-fit columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 65

    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = WRAP_ALIGN

    if ws.max_column and ws.max_row:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _build_table_styles_sheet(wb):
    ws = wb.create_sheet("Table Styles")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Header Row")
    _add_setting(ws, "table_header_font_name", "Arial",
                 "Table header font")
    _add_setting(ws, "table_header_font_size", 10,
                 "Table header font size in pt")
    _add_setting(ws, "table_header_bold", "TRUE",
                 "Bold header text")
    bold_row = ws.max_row
    _add_setting(ws, "table_header_font_color", "FFFFFF",
                 "Header text color (hex RGB)")
    _add_setting(ws, "table_header_fill_color", "2F5496",
                 "Header background color (hex RGB)")

    _add_subheader(ws, "Body Rows")
    _add_setting(ws, "table_body_font_name", "Calibri",
                 "Table body font")
    _add_setting(ws, "table_body_font_size", 10,
                 "Table body font size in pt")
    _add_setting(ws, "table_row_banding", "TRUE",
                 "Alternate row shading")
    banding_row = ws.max_row
    _add_setting(ws, "table_banding_color", "D6E4F0",
                 "Banding fill color (hex RGB)")

    _add_subheader(ws, "Borders")
    _add_setting(ws, "table_border_style", "single",
                 "Border style: single or none")
    border_row = ws.max_row
    _add_setting(ws, "table_border_color", "CCCCCC",
                 "Border color (hex RGB)")

    _add_subheader(ws, "Layout")
    _add_setting(ws, "table_width", "full",
                 "Table width: full (page width)")
    _add_setting(ws, "table_autofit", "TRUE",
                 "Auto-fit column widths")
    autofit_row = ws.max_row

    _add_bool_validation(ws, "B", bold_row, bold_row)
    _add_bool_validation(ws, "B", banding_row, banding_row)
    _add_bool_validation(ws, "B", autofit_row, autofit_row)
    _add_enum_validation(ws, "B", border_row, ["single", "none"])

    _finalize_sheet(ws)


def _build_cover_page_sheet(wb):
    ws = wb.create_sheet("Cover Page")
    _style_headers(ws, ["Setting", "Value", "Description"])

    _add_subheader(ws, "Cover Page")
    _add_setting(ws, "enable_cover_page", "FALSE",
                 "Generate a cover/title page")
    enable_row = ws.max_row
    _add_setting(ws, "cover_title", "{document_title}",
                 "Main title (resolves tokens from Document Identity)")
    _add_setting(ws, "cover_subtitle", "{organization_name}",
                 "Subtitle line")

    _add_subheader(ws, "Classification Banner")
    _add_setting(ws, "cover_classification", "{classification}",
                 "Classification banner text")
    _add_setting(ws, "cover_classification_color", "FF0000",
                 "Classification banner color (hex RGB)")

    _add_subheader(ws, "Signatories")
    _add_setting(ws, "cover_prepared_by", "",
                 '"Prepared By" field')
    _add_setting(ws, "cover_approved_by", "",
                 '"Approved By" field')
    _add_setting(ws, "cover_date", "{effective_date}",
                 "Date shown on cover")

    _add_subheader(ws, "Options")
    _add_setting(ws, "cover_logo_placeholder", "FALSE",
                 'Insert "[Organization Logo]" placeholder')
    logo_row = ws.max_row

    _add_bool_validation(ws, "B", enable_row, enable_row)
    _add_bool_validation(ws, "B", logo_row, logo_row)

    _finalize_sheet(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def generate_config_workbook(output_path: str):
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _build_readme(wb)
    _build_document_sheet(wb)
    _build_heading_styles_sheet(wb)
    _build_document_identity_sheet(wb)
    _build_page_layout_sheet(wb)
    _build_header_footer_sheet(wb)
    _build_standard_sections_sheet(wb)
    _build_table_styles_sheet(wb)
    _build_cover_page_sheet(wb)

    wb.save(output_path)
    print(f"  Config workbook saved: {output_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print()
    print("  Edit values in sheets, then run:")
    print("    python build_doc_from_config.py")


def main():
    parser = argparse.ArgumentParser(
        description="Generate doc_template_config.xlsx — heading style config for build_doc_from_config.py"
    )
    parser.add_argument(
        "-o", "--output",
        default="doc_template_config.xlsx",
        help="Output path for the config workbook (default: doc_template_config.xlsx)",
    )
    args = parser.parse_args()

    try:
        generate_config_workbook(args.output)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
